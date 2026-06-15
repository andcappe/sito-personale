"""
Dashboard Analisi Portafoglio — versione client standalone
Replica esatta della Tab 1 di ir_fe_14.py
"""

import io
import json
import base64
import time
import threading
import os
import sys
import uuid
import pickle
import requests
from pathlib import Path
from datetime import datetime

# Assicura che la cartella superiore sia disponibile per l'import di navbar.py
PARENT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if PARENT_DIR not in sys.path:
    sys.path.insert(0, PARENT_DIR)

import sessions_manager as _sm
import data_core as dc          # logica dati condivisa (un solo posto)

from style_analysis import get_style_analysis_tab, register_style_analysis_callbacks

import numpy as np
import pandas as pd
import yfinance as yf
import plotly.graph_objects as go
from plotly.subplots import make_subplots

from dash import Dash, html, dcc, dash_table, Input, Output, State, ALL, callback_context, no_update
from dash.exceptions import PreventUpdate
from flask import send_file as flask_send_file

# ─────────────────────────────────────────────────────────────────────────────
# App
# ─────────────────────────────────────────────────────────────────────────────
_EXTERNAL_STYLESHEETS = [
    'https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700&family=Inter:wght@400;600;700&display=swap',
    'https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.0/css/all.min.css',
]

app = Dash(__name__, suppress_callback_exceptions=True,
           external_stylesheets=_EXTERNAL_STYLESHEETS,
           requests_pathname_prefix='/portafoglio/',
           routes_pathname_prefix='/portafoglio/')
app.server.config['MAX_CONTENT_LENGTH'] = 256 * 1024 * 1024

app.index_string = '''
<!DOCTYPE html>
<html>
<head>
{%metas%}
<title>Analisi di Portafoglio</title>
{%favicon%}
{%css%}
<style>
  html { font-size: 16px; }
  @keyframes spin {
    0%   { transform: rotate(0deg); }
    100% { transform: rotate(360deg); }
  }
</style>
</head>
<body>
{%app_entry%}
<footer>{%config%}{%scripts%}{%renderer%}</footer>
<script>
(function(){
  var _tip=null;
  function _pos(el){
    var r=el.getBoundingClientRect();
    var color=el.getAttribute('data-tooltip-color')||'#1a3a6b';
    var text=el.getAttribute('data-tooltip')||'';
    if(!text)return;
    if(_tip){_tip.remove();_tip=null;}
    _tip=document.createElement('div');
    _tip.textContent=text;
    _tip.style.cssText='position:fixed;background:#fff;color:#1a2a4a;border:2px solid '+color+';border-radius:5px;padding:4px 10px;font-size:11px;font-family:Inter,sans-serif;font-weight:600;z-index:99999;pointer-events:none;white-space:nowrap;box-shadow:0 2px 8px rgba(0,0,0,0.13);';
    document.body.appendChild(_tip);
    var tx=r.right+8, ty=r.top+r.height/2-_tip.offsetHeight/2;
    if(tx+_tip.offsetWidth>window.innerWidth-8) tx=r.left-_tip.offsetWidth-8;
    _tip.style.left=Math.max(4,tx)+'px';
    _tip.style.top=Math.max(4,Math.min(ty,window.innerHeight-_tip.offsetHeight-4))+'px';
  }
  function _find(e){
    var el=e.target;
    while(el&&el!==document.body){if(el.hasAttribute&&el.hasAttribute('data-tooltip'))return el;el=el.parentNode;}
    return null;
  }
  document.addEventListener('mouseover',function(e){var el=_find(e);if(el)_pos(el);},true);
  document.addEventListener('mouseout',function(e){
    var el=_find(e);
    if(el&&!el.contains(e.relatedTarget)){if(_tip){_tip.remove();_tip=null;}}
  },true);
})();
</script>
</body>
</html>
'''

# Percorsi file
_XLSX        = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ETF.xlsx')
_PROFILO_HTML = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                              'profilo.html'))
_FOTO_PNG = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                          'assets', 'foto.png'))

# Cartella condivisa Files/ (un livello sopra rispetto a portafoglio/)
_FILES_DIR = Path(os.path.dirname(os.path.abspath(__file__))).parent / 'Files'

_FILE_ORDER = ['ETF', 'CRIPTO', 'COMMODITIES']
_PERSONALE_OPT = {'label': '👤 Personale', 'value': '__personale__'}

def _list_files():
    """Restituisce lista di opzioni dcc.Dropdown dai .xlsx in Files/."""
    if not _FILES_DIR.exists():
        return [{'label': 'ETF', 'value': 'ETF.xlsx'}]
    files = list(_FILES_DIR.glob('*.xlsx'))
    files.sort(key=lambda f: _FILE_ORDER.index(f.stem.upper()) if f.stem.upper() in _FILE_ORDER else 99)
    return [{'label': f.stem, 'value': f.name} for f in files] or \
           [{'label': 'ETF', 'value': 'ETF.xlsx'}]

def _list_files_with_personale():
    return [_PERSONALE_OPT] + _list_files()

def _xlsx_path(filename='ETF.xlsx'):
    """Percorso assoluto del file xlsx nella cartella Files/."""
    fp = _FILES_DIR / filename
    if fp.exists():
        return str(fp)
    return _XLSX  # fallback al file locale


_HEADER_KEYS = {'TICKER', 'ISIN', 'SIMBOLO', 'SYMBOL', 'CODICE', 'CODE',
                'TITOLO', 'CUSIP'}

def _read_asset_excel(source):
    """
    Legge un Excel di asset gestendo eventuali righe-titolo sopra le intestazioni.
    Se le intestazioni reali (TICKER/ISIN/…) non sono nella prima riga ma più sotto
    (es. una riga 'Tabella 1' iniziale), trova la riga di intestazione corretta e
    scarta tutto ciò che la precede. Per i file prezzi (date + numeri) o con header
    già corretto ricade sulla lettura standard.
    source: bytes/bytearray oppure percorso/file-like.
    """
    import io as _io_ra

    def _mk():
        return _io_ra.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source

    try:
        raw = pd.read_excel(_mk(), header=None)
    except Exception:
        return pd.read_excel(_mk())

    hdr_idx = None
    for i in range(min(6, len(raw))):
        vals = [str(x).strip().upper() for x in raw.iloc[i].tolist()]
        if any(v in _HEADER_KEYS for v in vals):
            hdr_idx = i
            break

    if hdr_idx is None:
        # Nessuna intestazione asset riconosciuta → lettura standard (es. file prezzi)
        return pd.read_excel(_mk())

    cols = [str(x).strip() for x in raw.iloc[hdr_idx].tolist()]
    df = raw.iloc[hdr_idx + 1:].copy()
    df.columns = cols
    df = df.dropna(how='all').reset_index(drop=True)
    # Scarta colonne senza nome / 'nan'
    keep = [c for c in df.columns if c and str(c).strip().lower() != 'nan']
    df = df.loc[:, keep]
    # Sicurezza: rimuovi eventuali righe-intestazione duplicate tra i dati
    if len(df.columns):
        c0 = df[df.columns[0]].astype(str).str.strip().str.upper()
        df = df[~c0.isin(_HEADER_KEYS)].reset_index(drop=True)
    return df

# Rotta Flask per servire la pagina profilo
@app.server.route('/health')
def health_check():
    return 'OK', 200

@app.server.route('/')
@app.server.route('/sito')
def serve_profilo():
    return flask_send_file(_PROFILO_HTML)

@app.server.route('/foto.png')
def serve_foto():
    return flask_send_file(_FOTO_PNG, mimetype='image/png')

@app.server.route('/clienti')
def pagina_clienti():
    from flask import Response
    files = sorted(SESSIONS_DIR.glob('tickers_*.xlsx'))
    righe = ''
    for i, f in enumerate(files, 1):
        ts_raw = f.stem.replace('tickers_', '')
        try:
            ts_fmt = datetime.strptime(ts_raw, '%Y%m%d_%H%M%S').strftime('%d/%m/%Y %H:%M:%S')
        except Exception:
            ts_fmt = ts_raw
        try:
            df_tmp = pd.read_excel(f)
            n_asset = len(df_tmp)
        except Exception:
            n_asset = '?'
        righe += (
            f'<tr>'
            f'<td style="padding:8px 16px;font-weight:700;color:#1a3a6b">#{i:03d}</td>'
            f'<td style="padding:8px 16px">{ts_fmt}</td>'
            f'<td style="padding:8px 16px">{n_asset} asset</td>'
            f'<td style="padding:8px 16px">'
            f'<a href="/clienti/download/{f.name}" '
            f'style="color:#2554a0;font-weight:600">⬇ Scarica</a>'
            f'</td>'
            f'</tr>\n'
        )
    if not righe:
        righe = '<tr><td colspan="4" style="padding:20px;color:#888;text-align:center">Nessun file caricato ancora.</td></tr>'
    html = f'''<!DOCTYPE html>
<html lang="it">
<head>
  <meta charset="UTF-8"/>
  <title>File Clienti — Andrea Cappelletti</title>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet"/>
  <style>
    body{{font-family:Inter,sans-serif;background:#fff;color:#1a2a4a;margin:0;padding:40px 5%}}
    h1{{font-size:1.6rem;color:#1a3a6b;margin-bottom:4px}}
    p{{color:#5a7099;font-size:0.9rem;margin-bottom:2rem}}
    table{{width:100%;border-collapse:collapse;box-shadow:0 2px 12px rgba(26,58,107,.08)}}
    th{{background:#1a3a6b;color:#fff;padding:10px 16px;text-align:left;font-size:0.82rem;letter-spacing:.05em;text-transform:uppercase}}
    tr:nth-child(even){{background:#f5f8fe}}
    tr:hover{{background:#e8f0fb}}
    a{{text-decoration:none}}
    .back{{display:inline-block;margin-top:2rem;color:#2554a0;font-weight:600;font-size:.85rem}}
  </style>
</head>
<body>
  <h1>File Clienti Caricati</h1>
  <p>Ogni file ticker caricato dal portale è elencato qui con data, ora e numero di asset.</p>
  <table>
    <thead><tr><th>#</th><th>Data/Ora</th><th>Asset</th><th>File</th></tr></thead>
    <tbody>{righe}</tbody>
  </table>
  <a class="back" href="/">← Torna al sito</a>
</body>
</html>'''
    return Response(html, mimetype='text/html')

@app.server.route('/clienti/download/<filename>')
def download_client_file(filename):
    from flask import abort
    if not filename.startswith('tickers_') or not filename.endswith('.xlsx'):
        abort(404)
    filepath = SESSIONS_DIR / filename
    if not filepath.exists():
        abort(404)
    return flask_send_file(str(filepath),
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True,
                           download_name=filename)

@app.server.route('/test-yf')
def test_yf():
    import json as _json
    done = threading.Event()
    result = [None]
    def _go():
        try:
            df = yf.download('AAPL', start='2024-01-01', auto_adjust=True,
                             progress=False, threads=False)
            result[0] = {'ok': True, 'rows': len(df), 'cols': list(df.columns)}
        except Exception as e:
            result[0] = {'ok': False, 'error': str(e)}
        finally:
            done.set()
    threading.Thread(target=_go, daemon=True).start()
    done.wait(timeout=30)
    return _json.dumps(result[0] or {'ok': False, 'error': 'timeout'})

# ─────────────────────────────────────────────────────────────────────────────
# Colori
# ─────────────────────────────────────────────────────────────────────────────
color_palette = [
    '#1f77b4', '#ff7f0e', '#2ca02c', '#9467bd', '#8c564b',
    '#e377c2', '#7f7f7f', '#bcbd22', '#17becf', '#a6cee3',
    '#b2df8a', '#fb9a99', '#fdbf6f', '#cab2d6', '#ffff99',
    '#191970', '#006400', '#8B0000', '#4B0082', '#2F4F4F',
    '#8B4513', '#FF1493', '#696969', '#556B2F', '#008B8B',
    '#E9967A', '#90EE90', '#B0C4DE', '#D3D3D3'
]

# ─────────────────────────────────────────────────────────────────────────────
# Calcoli
# ─────────────────────────────────────────────────────────────────────────────
def calculate_rolling_information_ratio(asset_returns, benchmark_returns, window):
    active_return = asset_returns - benchmark_returns.to_numpy()
    rolling_mean  = active_return.rolling(window=window).mean()
    rolling_std   = active_return.rolling(window=window).std()
    return (rolling_mean / rolling_std) * np.sqrt(252)

def calculate_rolling_sharpe_ratio(returns, window):
    rolling_mean = returns.rolling(window=window).mean()
    rolling_std  = returns.rolling(window=window).std()
    return (rolling_mean / rolling_std) * np.sqrt(252)

def calculate_tracking_error_volatility(asset_returns, benchmark_returns, window):
    active_return = asset_returns - benchmark_returns
    return active_return.rolling(window=window).std() * np.sqrt(252)

def calculate_drawdown(returns_series):
    cumulative   = (1 + returns_series).cumprod()
    rolling_max  = cumulative.cummax()
    return (cumulative - rolling_max) / rolling_max

def calculate_historical_cvar(returns_series, window, tail_pct=0.05):
    # Rendimenti composti rolling a N giorni
    n_day_ret = (1 + returns_series).rolling(window, min_periods=window).apply(np.prod, raw=True) - 1
    # CVaR calcolato sulla distribuzione espansa di tutti i rendimenti N-giorni visti fino a t
    def _cvar(w):
        n_tail = max(1, int(len(w) * tail_pct))
        return np.partition(w, n_tail)[:n_tail].mean()
    return n_day_ret.expanding(min_periods=window + 1).apply(_cvar, raw=True)

def _rolling_volatility(returns_series, window):
    return returns_series.rolling(window, min_periods=window // 2).std() * np.sqrt(252)

def _thin(s, max_pts=500):
    """Downsample una serie/DataFrame a max_pts punti per il rendering."""
    if len(s) <= max_pts:
        return s
    step = max(1, len(s) // max_pts)
    return s.iloc[::step]

# ─────────────────────────────────────────────────────────────────────────────
# Cache DataFrame
# ─────────────────────────────────────────────────────────────────────────────
_DF_CACHE: dict = {}

def _df_key(json_str):
    return json_str[:4000]

def _get_df(json_str):
    if not json_str:
        return None
    key = _df_key(json_str)
    if key not in _DF_CACHE:
        df = pd.read_json(io.StringIO(json_str), orient='split')
        df.index = pd.to_datetime(df.index)
        _DF_CACHE[key] = df
        if len(_DF_CACHE) > 20:
            oldest = next(iter(_DF_CACHE))
            del _DF_CACHE[oldest]
    return _DF_CACHE[key].copy()

# ─────────────────────────────────────────────────────────────────────────────
# Cache giornaliera su disco
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# Dati di mercato — download + persistenza su disco
# ─────────────────────────────────────────────────────────────────────────────
DOWNLOAD_BATCH_SIZE = 10

_DL_STATE  = {'status': 'idle', 'current': 0, 'total': 0, 'errors': []}
_DL_BUFFER: dict = {}   # dati file attivo — salvati su disco, permanenti
_DL_LOCK   = threading.Lock()

_CL_BUFFERS: dict = {}   # {username: {dati cliente}} — per-utente, solo in memoria
_CL_STATES:  dict = {}   # {username: stato download cliente}
_CL_LOCK   = threading.Lock()

_active_file_store: dict = {'filename': 'ETF.xlsx'}  # file attivo corrente
_PENDING: dict = {}  # ticker in attesa di download da Gestisci — processati da start_refresh

_ISIN_LOCK  = threading.Lock()
_ISIN_STATE: dict = {
    'running': False, 'done': False, 'req_id': None,
    'progress': '', 'n_done': 0, 'n_total': 0,
    'result_bytes': None, 'excluded': [],
    'tickers': [], 'descr': [], 'valuta': [], 'pesi': [],
}

_ROOT_DIR = Path(os.path.dirname(os.path.abspath(__file__))).parent

def _cloud_push(path):
    """Replica il file su storage persistente (S3/R2) se configurato. Best-effort."""
    try:
        import cloud_storage
        cloud_storage.push(path)
    except Exception:
        pass


def _user_json_path(username=None):
    u = username or _get_username()
    d = _ROOT_DIR / 'sessions' / u
    d.mkdir(parents=True, exist_ok=True)
    return d / 'current.json'


def _personale_path(username=None):
    """File personale PERSISTENTE (copia di current.json) per tornarci dopo i default."""
    u = username or _get_username()
    d = _ROOT_DIR / 'sessions' / u
    d.mkdir(parents=True, exist_ok=True)
    return d / 'personale.json'


def _mark_personale(username=None):
    """Marca il file di lavoro come Personale. UN SOLO file: scrive il tipo dentro
    current.json (nessuna copia/altro file). Al rientro current.json viene
    ricaricato così com'è (è già il file personale)."""
    u = username or _get_username()
    _active_file_store['is_personale'] = True
    _active_file_store['filename'] = '__personale__'
    _set_tipo('personale', username=u)


# ─────────────────────────────────────────────────────────────────────────────
# TAPPA 1 — Scrittura unica, ATOMICA e COERENTE del file dati (current.json)
# Garanzia: il file non resta mai scritto a metà (temp+rename) e non viene
# committato se incoerente → se un processo non si chiude, i dati restano
# allineati all'ultimo stato buono.
# ─────────────────────────────────────────────────────────────────────────────
def _profile_consistency(data):
    """Verifica la coerenza del profilo. Ritorna (ok, lista_errori)."""
    errs = []
    if not isinstance(data, dict):
        return False, ['struttura non valida']
    if not data:
        return True, []  # vuoto = lecito (default non ancora caricato)
    for desc, v in data.items():
        if not isinstance(v, dict):
            continue  # chiavi meta (es. "_tipo": "personale") — non sono asset, si ignorano
        dates = v.get('dates'); rets = v.get('returns')
        if not dates or not rets:
            errs.append(f'{desc}: senza prezzi/rendimenti')
        elif len(dates) != len(rets):
            errs.append(f'{desc}: dates({len(dates)})≠returns({len(rets)})')
    return (len(errs) == 0), errs


_JSON_WRITE_LOCK = threading.Lock()


def _atomic_json_write(path, data, *, validate=True):
    """
    Scrive il JSON in modo atomico (temp+rename). Se validate=True e i dati sono
    incoerenti NON sovrascrive il file (resta l'ultimo stato buono) e ritorna False.

    Temp file con nome UNIVOCO (pid+thread+counter): con gunicorn multi-thread (e
    la Frontiera nello stesso processo) due scritture concorrenti NON devono mai
    condividere lo stesso .tmp, altrimenti si sovrappongono e corrompono il JSON
    (es. virgole doppie). Un lock serializza inoltre le scritture nel processo.
    """
    if validate:
        ok, errs = _profile_consistency(data)
        if not ok:
            print(f"⚠ profilo INCOERENTE — non salvato. Esempi: {errs[:3]}")
            return False
    tmp = Path(f'{path}.{os.getpid()}.{threading.get_ident()}.{uuid.uuid4().hex[:8]}.tmp')
    try:
        with _JSON_WRITE_LOCK:
            with open(tmp, 'w') as f:
                json.dump(data, f)
            os.replace(tmp, path)   # atomico sullo stesso filesystem
        _cloud_push(path)           # replica su storage persistente (S3/R2)
        return True
    except Exception as e:
        print(f'⚠ _atomic_json_write: {e}')
        try:
            if tmp.exists():
                tmp.unlink()
        except Exception:
            pass
        return False


def profile_report(username=None):
    """Report ispezionabile dello stato del file (per verifica oggettiva)."""
    try:
        data = json.load(open(_user_json_path(username)))
    except Exception:
        return {'asset': 0, 'tutti_con_prezzi': True, 'incoerenze': [], 'esiste': False}
    ok, errs = _profile_consistency(data)
    con_prezzi = sum(1 for v in data.values() if v.get('returns'))
    return {
        'esiste': True,
        'asset': len(data),
        'con_prezzi': con_prezzi,
        'tutti_con_prezzi': (con_prezzi == len(data)),
        'pesi_P1': sum(1 for v in data.values() if v.get('P1', 0)),
        'pesi_P2': sum(1 for v in data.values() if v.get('P2', 0)),
        'pesi_P3': sum(1 for v in data.values() if v.get('P3', 0)),
        'coerente': ok,
        'incoerenze': errs[:10],
    }


def _write_user_json(cr, op, tm, vm=None, username=None, reset_state=False, tipo=None):
    if cr is None or op is None:
        return
    vm = vm or {}
    path = _user_json_path(username)
    existing = {}
    try:
        existing = json.load(open(path))   # serve sempre per conservare _tipo
    except Exception:
        pass
    dates = [d.strftime('%Y-%m-%d') for d in op.index]
    result = {}
    for desc in cr.columns:
        if desc not in op.columns:
            continue
        ex = existing.get(desc, {}) if not reset_state else {}
        if not isinstance(ex, dict):
            ex = {}
        result[desc] = {
            'ticker':   tm.get(desc, desc),
            'currency': vm.get(desc, ex.get('currency', 'EUR')),
            'dates':    dates,
            'prices':   [round(float(v), 4) if pd.notna(v) else None for v in op[desc]],
            'returns':  [round(float(v), 6) if pd.notna(v) else None for v in cr[desc]],
            'checked':  False if reset_state else ex.get('checked', False),
            'P1':       0     if reset_state else ex.get('P1', 0),
            'P2':       0     if reset_state else ex.get('P2', 0),
            'P3':       0     if reset_state else ex.get('P3', 0),
        }
    # Tipo del file di lavoro (unica fonte: current.json). Se non passato, conserva
    # quello esistente.
    _tipo = tipo if tipo is not None else existing.get('_tipo')
    if _tipo:
        result['_tipo'] = _tipo
    _atomic_json_write(path, result)

def _update_user_json(checked=None, weights=None, username=None):
    path = _user_json_path(username)
    try:
        data = json.load(open(path))
    except Exception:
        return
    changed = False
    if checked is not None:
        s = set(checked)
        for desc in data:
            if not isinstance(data[desc], dict):
                continue
            v = desc in s
            if data[desc].get('checked') != v:
                data[desc]['checked'] = v
                changed = True
    if weights is not None:
        for desc in data:
            if not isinstance(data[desc], dict):
                continue
            for k in ('P1', 'P2', 'P3'):
                v = float(weights.get(k, {}).get(desc, 0) or 0)
                if data[desc].get(k) != v:
                    data[desc][k] = v
                    changed = True
    if changed:
        _atomic_json_write(path, data)

def _read_user_json(username=None):
    """Solo voci-asset di current.json (le chiavi meta tipo _tipo sono escluse)."""
    try:
        raw = json.load(open(_user_json_path(username)))
        return {k: v for k, v in raw.items() if isinstance(v, dict)}
    except Exception:
        return {}


def _read_tipo(username=None):
    """Tipo del file di lavoro: 'personale' oppure 'default:ETF.xlsx' ('' se assente)."""
    try:
        raw = json.load(open(_user_json_path(username)))
        return raw.get('_tipo', '') or ''
    except Exception:
        return ''


def _set_tipo(tipo, username=None):
    """Imposta il tipo dentro current.json (nessun file separato)."""
    path = _user_json_path(username)
    try:
        raw = json.load(open(path))
    except Exception:
        return
    if raw.get('_tipo') != tipo:
        raw['_tipo'] = tipo
        _atomic_json_write(path, raw)


def _get_username() -> str:
    """Username dalla sessione Flask; 'anon' come fallback."""
    try:
        from flask import session as _fs
        return _fs.get('username') or 'anon'
    except Exception:
        return 'anon'


def _cl_buf(username: str) -> dict:
    with _CL_LOCK:
        if username not in _CL_BUFFERS:
            _CL_BUFFERS[username] = {}
        return _CL_BUFFERS[username]


def _cl_state(username: str) -> dict:
    with _CL_LOCK:
        if username not in _CL_STATES:
            _CL_STATES[username] = {'status': 'idle', 'current': 0, 'total': 0, 'errors': []}
        return _CL_STATES[username]


def _cl_clear(username: str):
    with _CL_LOCK:
        _CL_BUFFERS[username] = {}
        _CL_STATES[username]  = {'status': 'idle', 'current': 0, 'total': 0, 'errors': []}


def _user_cache_path(username: str, filename: str = None) -> Path:
    fn   = filename or _active_file_store.get('filename', 'ETF.xlsx')
    stem = Path(fn).stem
    return SESSIONS_DIR / f'market_data_{stem}_user_{username}.pkl'


def _build_ticker_list(filename='ETF.xlsx'):
    df   = pd.read_excel(_xlsx_path(filename))
    cols = df.columns.tolist()
    return list(df[cols[0]]), list(df[cols[1]]), list(df[cols[2]])



def _clean_prices(df: pd.DataFrame) -> pd.DataFrame:
    """
    Corregge valori fuori scala ×100 o ÷100 (es. Yahoo Finance che alterna GBp↔GBP/EUR).
    Per ogni asset:
    - valori < 2% della mediana → troppo bassi (÷100) → ×100
    - valori > 50× la mediana  → troppo alti  (×100)  → ÷100
    Soglie molto conservative: non tocca nulla che non sia inequivocabilmente sbagliato.
    """
    df = df.copy()
    for col in df.columns:
        s = df[col].dropna()
        if len(s) < 10:
            continue
        med = s.median()
        if med <= 0:
            continue
        # valori troppo bassi (÷100)
        low_mask = (df[col] > 0) & (df[col] < med * 0.02)
        if low_mask.any():
            df.loc[low_mask, col] = df.loc[low_mask, col] * 100
            print(f"  ⚠ {col}: {low_mask.sum()} valori ×100 (erano < 2% mediana)")
        # valori troppo alti (×100) — ricalcola mediana dopo eventuale correzione
        med2 = df[col].dropna().median()
        high_mask = (df[col] > 0) & (df[col] > med2 * 50)
        if high_mask.any():
            df.loc[high_mask, col] = df.loc[high_mask, col] / 100
            print(f"  ⚠ {col}: {high_mask.sum()} valori ÷100 (erano > 50× mediana)")
    return df


def _atomic_pkl_write(path: Path, data: dict):
    """Scrive il pkl in modo atomico: prima su .tmp, poi os.replace."""
    tmp = path.with_suffix('.tmp')
    with open(tmp, 'wb') as f:
        pickle.dump(data, f)
    os.replace(tmp, path)
    _cloud_push(path)           # replica su storage persistente (S3/R2)


def _do_add_tickers(new_tickers, new_descr, new_valuta, start_date, cache_file):
    """
    Scarica solo i nuovi ticker (stessa procedura nightly) su file temp,
    fa il merge con il pkl esistente, scrive atomicamente, aggiorna il buffer.
    Lo status rimane 'running' fino a merge completato.
    """
    global _DL_STATE, _DL_BUFFER
    total = len(new_tickers)
    tmp_pkl = cache_file.parent / (cache_file.stem + '_adding_new.pkl')

    try:
        # Scarica i nuovi ticker su file temporaneo (NON tocca il pkl di default)
        _do_download(new_tickers, new_descr, new_valuta, start_date,
                     cache_file=tmp_pkl, update_buffer=False)
        # _do_download lascia _DL_STATE['status'] invariato (update_buffer=False)

        if not tmp_pkl.exists():
            with _DL_LOCK:
                _DL_STATE['status'] = 'error'
            return

        with open(tmp_pkl, 'rb') as f:
            new_data = pickle.load(f)

        # Merge: leggi il pkl esistente, aggiungi le nuove colonne
        merged = dict(new_data)
        if cache_file.exists():
            try:
                with open(cache_file, 'rb') as f:
                    ex = pickle.load(f)
                ex_op = ex.get('original_prices')
                ex_cr = ex.get('close_returns')
                ex_tm = dict(ex.get('ticker_map', {}))
                if ex_op is not None and ex_cr is not None:
                    for col in new_data['original_prices'].columns:
                        ex_op[col] = new_data['original_prices'][col].reindex(ex_op.index).ffill()
                        ex_cr[col] = new_data['close_returns'][col].reindex(ex_op.index)
                    ex_tm.update(new_data['ticker_map'])
                    merged['original_prices'] = ex_op
                    merged['close_returns']   = ex_cr
                    merged['ticker_map']      = ex_tm
                    print(f"✓ Merge — {len(ex_op.columns)} asset totali")
            except Exception as e:
                print(f"⚠ Merge fallito, uso solo nuovi: {e}")

        # Scrittura atomica — il pkl di default viene sostituito solo qui
        _atomic_pkl_write(cache_file, merged)

        with _DL_LOCK:
            _DL_BUFFER.update(merged)
            _DL_STATE['status']  = 'done'
            _DL_STATE['current'] = total

    except Exception as e:
        print(f"❌ _do_add_tickers: {e}")
        with _DL_LOCK:
            _DL_STATE['status'] = 'error'
    finally:
        if tmp_pkl.exists():
            tmp_pkl.unlink()


def _do_reload_from_disk(cache_file, active_file='ETF.xlsx'):
    """Ricarica dati + ARIMA dal disco senza scaricare nulla da internet."""
    global _DL_STATE, _DL_BUFFER
    with _DL_LOCK:
        _DL_STATE = {'status': 'running', 'current': 0, 'total': 1, 'errors': []}
    try:
        if not cache_file.exists():
            with _DL_LOCK:
                _DL_STATE['status'] = 'error'
            print(f"⚠ Reload: {cache_file.name} non trovato")
            return
        with open(cache_file, 'rb') as f:
            data = pickle.load(f)
        # Carica anche ARIMA dal file separato
        arima_pkl = _arima_cache_path(active_file)
        if arima_pkl.exists():
            try:
                with open(arima_pkl, 'rb') as f:
                    ad = pickle.load(f)
                if ad.get('arima'):
                    data['arima']             = ad['arima']
                    data['arima_computed_at'] = ad.get('arima_computed_at', '')
            except Exception:
                pass
        with _DL_LOCK:
            _DL_BUFFER.update(data)
            _DL_STATE['status']  = 'done'
            _DL_STATE['current'] = 1
        print(f"✓ Dati ricaricati da disco — {data.get('saved_at', '?')}")
    except Exception as e:
        print(f"⚠ Reload dal disco fallito: {e}")
        with _DL_LOCK:
            _DL_STATE['status'] = 'error'


def _do_full_update(tickers, descr, valuta, start_date, cache_file, incremental=False):
    """Download incrementale o completo. ARIMA non viene eseguito qui (gira solo a mezzanotte)."""
    if incremental:
        _do_add_tickers(tickers, descr, valuta, start_date, cache_file)
    else:
        _do_download(tickers, descr, valuta, start_date, cache_file=cache_file, update_buffer=True)


def _do_download(tickers, descrizione, valuta, start_date, cache_file=None, update_buffer=True, username=None, tipo=None):
    """Scarica da Yahoo Finance e salva nella cache; cache_file=None usa market_data.pkl.
    Se username è dato e update_buffer, scrive anche current.json (fonte unica).
    tipo: imposta '_tipo' in current.json (es. 'default:CRIPTO.xlsx')."""
    global _DL_STATE, _DL_BUFFER
    total = len(tickers)
    with _DL_LOCK:
        _DL_STATE = {'status': 'running', 'current': 0, 'total': total, 'errors': []}

    # Proxy e User-Agent da variabili d'ambiente (configura su Render se Yahoo blocca)
    _proxy = os.environ.get('HTTPS_PROXY') or os.environ.get('https_proxy') or None
    _ua    = (os.environ.get('YF_USER_AGENT') or
              'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
              'AppleWebKit/537.36 (KHTML, like Gecko) '
              'Chrome/124.0.0.0 Safari/537.36')

    _dl_kwargs = dict(auto_adjust=True, progress=False)
    if _proxy:
        # yfinance recente NON accetta proxy= in download() (TypeError). requests
        # usa già HTTPS_PROXY/https_proxy dall'ambiente; in più lo impostiamo via
        # set_config quando disponibile. NON passare proxy nei kwargs di download.
        try:
            yf.set_config(proxy=_proxy)
        except Exception:
            pass
        print(f"▶ Download via proxy: {_proxy.split('@')[-1]}")

    # Patch User-Agent sulla sessione requests usata da yfinance
    try:
        import yfinance.data as _yfd
        if hasattr(_yfd, 'YfData'):
            _yfd.YfData._headers = {'User-Agent': _ua}
    except Exception:
        pass

    # FX
    fx = None
    try:
        fx = yf.download(['EURUSD=X', 'EURGBP=X'], start=start_date,
                         group_by='ticker', **_dl_kwargs)
    except Exception as e:
        print(f"⚠ FX download fallito: {e}")

    def _fx(name):
        if fx is None or fx.empty:
            return None
        try:
            return fx[(name, 'Close')] if isinstance(fx.columns, pd.MultiIndex) else fx['Close']
        except Exception:
            return None

    eurusd, eurgbp = _fx('EURUSD=X'), _fx('EURGBP=X')
    all_prices = {}

    for i in range(0, total, DOWNLOAD_BATCH_SIZE):
        bt = tickers[i:i + DOWNLOAD_BATCH_SIZE]
        bd = descrizione[i:i + DOWNLOAD_BATCH_SIZE]
        bv = valuta[i:i + DOWNLOAD_BATCH_SIZE]
        try:
            raw = yf.download(bt, start=start_date, group_by='ticker', **_dl_kwargs)
            if raw.empty:
                raise ValueError("risposta vuota")
            for j, t in enumerate(bt):
                desc, curr = bd[j], bv[j]
                try:
                    px = (raw[(t, 'Close')].copy() if isinstance(raw.columns, pd.MultiIndex)
                          else raw['Close'].copy())
                    px = px.ffill()
                    if curr == 'USD' and eurusd is not None:
                        px = px / eurusd.reindex(px.index).ffill()
                    elif curr == 'GBP' and eurgbp is not None:
                        px = px / eurgbp.reindex(px.index).ffill()
                    all_prices[desc] = px
                except Exception as e2:
                    with _DL_LOCK:
                        _DL_STATE['errors'].append(f"{t}: {e2}")
        except Exception as e:
            with _DL_LOCK:
                _DL_STATE['errors'].append(f"Batch {i}: {e}")
        with _DL_LOCK:
            _DL_STATE['current'] = min(i + DOWNLOAD_BATCH_SIZE, total)
        time.sleep(0.3)

    if not all_prices:
        with _DL_LOCK:
            _DL_STATE['status'] = 'error'
        print("❌ Download fallito: nessun dato")
        return

    original_prices = pd.DataFrame(all_prices)
    original_prices.index = pd.to_datetime(original_prices.index)
    original_prices = original_prices.ffill()
    original_prices = _clean_prices(original_prices)
    close_returns   = original_prices.pct_change(fill_method=None)
    ticker_map      = {descrizione[i]: tickers[i] for i in range(len(tickers))}
    saved_at        = datetime.now().strftime('%d/%m/%Y %H:%M')

    data = {
        'date':            datetime.now().strftime('%Y-%m-%d'),
        'saved_at':        saved_at,
        'ticker_map':      ticker_map,
        'original_prices': original_prices,
        'close_returns':   close_returns,
    }
    target_pkl = cache_file or _MARKET_DATA_FILE
    try:
        _atomic_pkl_write(target_pkl, data)
        print(f"✓ {target_pkl.name} salvato — {len(all_prices)} asset — {saved_at}")
    except Exception as e:
        print(f"⚠ Salvataggio su disco fallito: {e}")

    with _DL_LOCK:
        if update_buffer:
            _DL_BUFFER.update(data)
            _DL_STATE['status'] = 'done'
        _DL_STATE['current'] = total
    # Fonte unica: scrivi current.json per l'utente (così il polling lo legge)
    if username and update_buffer:
        try:
            _write_user_json(close_returns, original_prices, ticker_map,
                             username=username, reset_state=True, tipo=tipo)
        except Exception as e:
            print(f"⚠ _do_download current.json fallito: {e}", flush=True)


def _do_download_client(tickers, descrizione, valuta, start_date, username='anon', pesi_p1=None):
    """Download per file cliente: dati isolati in _CL_BUFFERS[username]."""
    total = len(tickers)
    with _CL_LOCK:
        _CL_STATES[username]  = {'status': 'running', 'current': 0, 'total': total, 'errors': []}
        _CL_BUFFERS[username] = {}

    _proxy = os.environ.get('HTTPS_PROXY') or os.environ.get('https_proxy') or None
    _ua    = (os.environ.get('YF_USER_AGENT') or
              'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
              'AppleWebKit/537.36 (KHTML, like Gecko) '
              'Chrome/124.0.0.0 Safari/537.36')
    _dl_kwargs = dict(auto_adjust=True, progress=False)
    if _proxy:
        # yfinance recente non accetta proxy= in download(): requests usa già
        # HTTPS_PROXY dall'ambiente; lo impostiamo anche via set_config se c'è.
        try:
            yf.set_config(proxy=_proxy)
        except Exception:
            pass

    try:
        import yfinance.data as _yfd
        if hasattr(_yfd, 'YfData'):
            _yfd.YfData._headers = {'User-Agent': _ua}
    except Exception:
        pass

    fx = None
    try:
        fx = yf.download(['EURUSD=X', 'EURGBP=X'], start=start_date,
                         group_by='ticker', **_dl_kwargs)
    except Exception as e:
        print(f"⚠ FX download cliente fallito: {e}")

    def _fx(name):
        if fx is None or fx.empty:
            return None
        try:
            return fx[(name, 'Close')] if isinstance(fx.columns, pd.MultiIndex) else fx['Close']
        except Exception:
            return None

    eurusd, eurgbp = _fx('EURUSD=X'), _fx('EURGBP=X')
    all_prices = {}

    for i in range(0, total, DOWNLOAD_BATCH_SIZE):
        bt = tickers[i:i + DOWNLOAD_BATCH_SIZE]
        bd = descrizione[i:i + DOWNLOAD_BATCH_SIZE]
        bv = valuta[i:i + DOWNLOAD_BATCH_SIZE]
        raw_res = [None]
        def _dl_batch(syms=bt):
            try:
                raw_res[0] = yf.download(syms, start=start_date,
                                         group_by='ticker', **_dl_kwargs)
            except Exception:
                pass
        _bt = threading.Thread(target=_dl_batch, daemon=True)
        _bt.start()
        _bt.join(timeout=60)  # max 60s per batch — se si blocca salta
        if _bt.is_alive() or raw_res[0] is None or raw_res[0].empty:
            with _CL_LOCK:
                _CL_STATES[username]['errors'].append(
                    f"Batch {i//DOWNLOAD_BATCH_SIZE+1} ({','.join(bt)}): timeout o risposta vuota")
        else:
            raw = raw_res[0]
            for j, t in enumerate(bt):
                desc, curr = bd[j], bv[j]
                try:
                    px = (raw[(t, 'Close')].copy() if isinstance(raw.columns, pd.MultiIndex)
                          else raw['Close'].copy())
                    px = px.ffill()
                    if curr == 'USD' and eurusd is not None:
                        px = px / eurusd.reindex(px.index).ffill()
                    elif curr == 'GBP' and eurgbp is not None:
                        px = px / eurgbp.reindex(px.index).ffill()
                    all_prices[desc] = px
                except Exception as e2:
                    with _CL_LOCK:
                        _CL_STATES[username]['errors'].append(f"{t}: {e2}")
        with _CL_LOCK:
            _CL_STATES[username]['current'] = min(i + DOWNLOAD_BATCH_SIZE, total)
        time.sleep(0.3)

    if not all_prices:
        with _CL_LOCK:
            _CL_STATES[username]['status'] = 'error'
        print(f"❌ Download cliente [{username}] fallito: nessun dato")
        return

    original_prices = pd.DataFrame(all_prices)
    original_prices.index = pd.to_datetime(original_prices.index)
    original_prices = original_prices.ffill()
    original_prices = _clean_prices(original_prices)
    close_returns   = original_prices.pct_change(fill_method=None)
    ticker_map      = {descrizione[i]: tickers[i] for i in range(len(tickers))}
    saved_at        = datetime.now().strftime('%d/%m/%Y %H:%M')

    valuta_map = {descrizione[i]: valuta[i] for i in range(len(tickers))}
    with _CL_LOCK:
        _CL_BUFFERS[username].update({
            'date':            datetime.now().strftime('%Y-%m-%d'),
            'saved_at':        saved_at,
            'ticker_map':      ticker_map,
            'original_prices': original_prices,
            'close_returns':   close_returns,
            'valuta_map':      valuta_map,
        })
        _CL_STATES[username]['current'] = total
    # Scrivi JSON e pesi P1 PRIMA di segnalare 'done' al poll
    _write_user_json(close_returns, original_prices, ticker_map, valuta_map, username=username, reset_state=True)
    if pesi_p1:
        _update_user_json(weights={'P1': pesi_p1, 'P2': {}, 'P3': {}}, username=username)
        print(f"✓ Pesi P1 impostati: {len(pesi_p1)} asset")
    # Auto-save nella sessione di lavoro condivisa
    _sm.save_working(username, {
        'close_returns':   close_returns,
        'original_prices': original_prices,
        'ticker_map':      ticker_map,
        'valuta_map':      valuta_map,
        'saved_at':        saved_at,
        '_source':         'user_file',
    })
    with _CL_LOCK:
        _CL_STATES[username]['status'] = 'done'
    # Caricare un file ticker da zero → file di lavoro Personale (persistito).
    _mark_personale(username)
    print(f"✓ Download cliente [{username}]: {len(all_prices)} asset isolati")


def _do_gestisci_download(new_tickers, new_descr, new_valuta, start_date, username, filename, all_descr=None, persist_personale=False):
    """Scarica nuovi ticker da Gestisci, merge con i dati correnti (_DL_BUFFER), salva pkl utente.
    all_descr: lista completa delle descrizioni desiderate — filtra il merged result a questi soli."""
    import shutil as _shutil
    total    = len(new_tickers)
    user_pkl = SESSIONS_DIR / f'market_data_{Path(filename).stem}_user_{username}.pkl'

    with _DL_LOCK:
        _DL_STATE.update({'status': 'running', 'current': 0, 'total': total, 'errors': []})
    # Su file personale il polling legge il buffer CLIENT: tienilo allineato (running)
    with _CL_LOCK:
        _CL_STATES.setdefault(username, {}).update(
            {'status': 'running', 'current': 0, 'total': total, 'errors': []})

    # Non copiare il pkl esistente: partiamo sempre da zero per non includere ticker non voluti.
    # Se l'utente non ha un pkl, non creiamone uno adesso — lo creiamo dopo il merge.

    _proxy     = os.environ.get('HTTPS_PROXY') or os.environ.get('https_proxy') or None
    _ua        = (os.environ.get('YF_USER_AGENT') or
                  'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36')
    _dl_kwargs = dict(auto_adjust=True, progress=False)
    if _proxy:
        # yfinance recente non accetta proxy= in download(): requests usa già
        # HTTPS_PROXY dall'ambiente; lo impostiamo anche via set_config se c'è.
        try:
            yf.set_config(proxy=_proxy)
        except Exception:
            pass
    try:
        import yfinance.data as _yfd
        if hasattr(_yfd, 'YfData'):
            _yfd.YfData._headers = {'User-Agent': _ua}
    except Exception:
        pass

    _needs_fx = any(v.upper() not in ('EUR', '', 'NAN') for v in new_valuta if v)
    _fx_res   = [None]

    def _dl_fx():
        try:
            _fx_res[0] = yf.download(['EURUSD=X', 'EURGBP=X'], start=start_date,
                                     group_by='ticker', **_dl_kwargs)
        except Exception as e:
            print(f"⚠ FX gestisci fallito: {e}")

    # Avvia download FX in parallelo con i batch ticker
    _fx_thread = None
    if _needs_fx:
        _fx_thread = threading.Thread(target=_dl_fx, daemon=True)
        _fx_thread.start()

    # Fase 1: scarica i prezzi grezzi di tutti i batch (FX scarica in parallelo)
    all_prices = {}

    for i in range(0, total, DOWNLOAD_BATCH_SIZE):
        bt = new_tickers[i:i + DOWNLOAD_BATCH_SIZE]
        bd = new_descr[i:i + DOWNLOAD_BATCH_SIZE]
        bv = new_valuta[i:i + DOWNLOAD_BATCH_SIZE]
        try:
            raw = yf.download(bt, start=start_date, group_by='ticker', **_dl_kwargs)
            if raw.empty:
                raise ValueError("risposta vuota")
            # Al primo batch: FX thread dovrebbe essere finito in parallelo — join con timeout
            if i == 0 and _fx_thread is not None:
                _fx_thread.join(timeout=30)
            _eurusd = _fx_res[0]
            def _get_fx(name):
                if _eurusd is None or _eurusd.empty: return None
                try:
                    return _eurusd[(name, 'Close')] if isinstance(_eurusd.columns, pd.MultiIndex) else _eurusd['Close']
                except Exception: return None
            eurusd = _get_fx('EURUSD=X')
            eurgbp = _get_fx('EURGBP=X')
            for j, t in enumerate(bt):
                desc, curr = bd[j], bv[j]
                try:
                    px = (raw[(t, 'Close')].copy() if isinstance(raw.columns, pd.MultiIndex)
                          else raw['Close'].copy())
                    px = px.ffill()
                    if curr == 'USD' and eurusd is not None:
                        px = px / eurusd.reindex(px.index).ffill()
                    elif curr == 'GBP' and eurgbp is not None:
                        px = px / eurgbp.reindex(px.index).ffill()
                    all_prices[desc] = px
                except Exception as e2:
                    with _DL_LOCK:
                        _DL_STATE['errors'].append(f"{t}: {e2}")
        except Exception as e:
            with _DL_LOCK:
                _DL_STATE['errors'].append(f"Batch {i}: {e}")
        with _DL_LOCK:
            _DL_STATE['current'] = min(i + DOWNLOAD_BATCH_SIZE, total)
        time.sleep(0.3)

    if not all_prices:
        with _DL_LOCK:
            _DL_STATE['status'] = 'error'
        with _CL_LOCK:
            _CL_STATES.setdefault(username, {})['status'] = 'error'
        print(f"❌ Gestisci [{username}]: nessun dato scaricato")
        return

    new_prices = pd.DataFrame(all_prices)
    new_prices.index = pd.to_datetime(new_prices.index)
    new_prices = new_prices.ffill()
    new_prices = _clean_prices(new_prices)

    new_tm     = {new_descr[i]: new_tickers[i] for i in range(len(new_tickers))}
    merged_op  = new_prices.copy()
    merged_tm  = dict(new_tm)

    # Merge con i dati esistenti. FONTE PRIMARIA: current.json (sempre la lista
    # corrente completa) → l'accodo non perde mai gli asset esistenti.
    # Fallback: pkl utente, poi _DL_BUFFER.
    ex_op, ex_tm = None, {}
    try:
        _cur = json.load(open(_user_json_path(username)))
        _cols = {}
        for _a, _v in _cur.items():
            if isinstance(_v, dict) and _v.get('prices') and _v.get('dates'):
                _cols[_a] = pd.Series(_v['prices'], index=pd.to_datetime(_v['dates']))
                ex_tm[_a] = _v.get('ticker') or _a
        if _cols:
            ex_op = pd.DataFrame(_cols).sort_index()
            print(f"✓ Merge gestisci da current.json — {len(ex_op.columns)} asset esistenti")
    except Exception as _e:
        ex_op, ex_tm = None, {}

    if (ex_op is None or ex_op.empty) and user_pkl.exists():
        try:
            with open(user_pkl, 'rb') as f:
                ex = pickle.load(f)
            ex_op = ex.get('original_prices')
            ex_tm = dict(ex.get('ticker_map', {}))
            print(f"✓ Merge gestisci da pkl utente")
        except Exception as e:
            print(f"⚠ Merge gestisci pkl fallito: {e}")
            ex_op, ex_tm = None, {}

    if ex_op is None or ex_op.empty:
        with _DL_LOCK:
            ex_op = _DL_BUFFER.get('original_prices')
            ex_tm = dict(_DL_BUFFER.get('ticker_map', {}))
        if ex_op is not None and not ex_op.empty:
            print(f"✓ Merge gestisci da _DL_BUFFER — {len(ex_op.columns)} asset esistenti")

    if ex_op is not None and not ex_op.empty:
        ex_op = ex_op.copy()
        for col in new_prices.columns:
            ex_op[col] = new_prices[col].reindex(ex_op.index).ffill()
        merged_op = ex_op
        ex_tm.update(new_tm)
        merged_tm = ex_tm
        print(f"✓ Merge gestisci — {len(merged_op.columns)} asset pre-filtro")

    # Filtra: tieni solo le descrizioni che l'utente vuole (all_descr)
    if all_descr is not None:
        want = set(all_descr)
        keep = [c for c in merged_op.columns if c in want]
        merged_op = merged_op[keep]
        merged_tm = {k: v for k, v in merged_tm.items() if k in want}
        print(f"✓ Filtro gestisci — {len(merged_op.columns)} asset finali")

    merged_cr   = merged_op.pct_change(fill_method=None)
    saved_at    = datetime.now().strftime('%d/%m/%Y %H:%M')
    merged_data = {
        'date':            datetime.now().strftime('%Y-%m-%d'),
        'saved_at':        saved_at,
        'ticker_map':      merged_tm,
        'original_prices': merged_op,
        'close_returns':   merged_cr,
    }
    _atomic_pkl_write(user_pkl, merged_data)
    with _DL_LOCK:
        _DL_BUFFER.update(merged_data)
        _DL_STATE['status']  = 'done'
        _DL_STATE['current'] = total
    _write_user_json(merged_cr, merged_op, merged_tm, username=username)
    # Stato client 'done': il polling (su file personale) legge current.json aggiornato
    with _CL_LOCK:
        _CL_STATES.setdefault(username, {}).update({'status': 'done', 'current': total})
    print(f"✓ Gestisci [{username}]: {user_pkl.name} — {len(merged_op.columns)} asset")

    # L'accodo modifica il file → tipo "personale" dentro current.json (un solo
    # file, nessuna copia). Così al rientro current.json viene ricaricato com'è.
    if persist_personale:
        _set_tipo('personale', username=username)


# ─────────────────────────────────────────────────────────────────────────────
# ISIN → Ticker conversion helpers
# ─────────────────────────────────────────────────────────────────────────────
import re as _re_isin

_ISIN_PATTERN = _re_isin.compile(r'^[A-Z]{2}[A-Z0-9]{9}[0-9]$')

_DEFAULT_PORTFOLIO_ROWS = [
    ('ISAC.L',    'Az. ACWI',                    'USD'),
    ('SWDA.MI',   'Az. World',                    'EUR'),
    ('DBMFE.PA',  'Alt. MF',                      'EUR'),
    ('UIQ4.DE',   'Alt. Eu.Def',                  'EUR'),
    ('JEGA.MI',   'Alt. Pr.Inc.Ac',               'EUR'),
    ('INFL.PA',   'Alt. Inf.Br.',                 'EUR'),
    ('IWVU.L',    'Az. World Value Factor',        'USD'),
    ('QDVE.DE',   'AZ. USA Info Tech SP500',       'EUR'),
    ('CNDX.L',    'Az.USA Nasdaq100',              'USD'),
    ('CSSPX.MI',  'Az. USA SP500',                'EUR'),
    ('IUSG',      'Az. USA Growth SP500',          'USD'),
    ('VUG',       'Az. USA Growth',               'USD'),
    ('IVE',       'Az. USA Value SP500',           'USD'),
    ('ZPRV.DE',   'Az USA Value Small Cap',        'EUR'),
    ('R2US.MI',   'Az USA Small Cap Russell2000',  'EUR'),
    ('ZPRX.DE',   'Az EUROPA Value Small Cap',     'EUR'),
    ('XSX6.MI',   'Az. Europe Stoxx 600',          'EUR'),
    ('EIMI.MI',   'Az. Emerging Market',           'EUR'),
]


def _is_isin(s):
    return bool(_ISIN_PATTERN.match(str(s).strip().upper()))


def _parse_weight_col(series):
    """Normalizza una colonna pesi a float 0-100. Ritorna Series o None."""
    try:
        cleaned = series.astype(str).str.strip().str.replace('%', '', regex=False)
        nums = pd.to_numeric(cleaned, errors='coerce').dropna()
        if len(nums) == 0 or nums.min() < 0:
            return None
        if nums.max() <= 1.01:
            return (nums * 100).round(4)
        if nums.max() <= 100.01:
            return nums.round(4)
    except Exception:
        pass
    return None


def _detect_ticker_and_weight_cols(df):
    """Restituisce (ticker_col, weight_col, weight_series)."""
    _TICKER_H = {'ticker', 'isin', 'simbolo', 'symbol', 'codice', 'code', 'titolo', 'cusip'}
    _WEIGHT_H = {'peso', 'weight', 'allocation', '%', 'pct', 'percentual',
                 'quota', 'alloc', 'perc', 'ponder', 'peso %', 'weight %'}

    ticker_col = None
    for col in df.columns:
        if str(col).strip().lower() in _TICKER_H:
            ticker_col = col
            break
    if ticker_col is None:
        ticker_col = df.columns[0]

    weight_col, w_vals = None, None
    for col in df.columns:
        if col == ticker_col:
            continue
        if str(col).strip().lower() in _WEIGHT_H:
            w_vals = _parse_weight_col(df[col])
            if w_vals is not None:
                weight_col = col
                break
    if weight_col is None:
        for col in df.columns:
            if col == ticker_col:
                continue
            w_vals = _parse_weight_col(df[col])
            if w_vals is not None:
                weight_col = col
                break

    return ticker_col, weight_col, w_vals


_OPENFIGI_EXCH = {
    'LN': '.L',   'IM': '.MI',  'GY': '.DE',  'FP': '.PA',
    'NA': '.AS',  'BB': '.BR',  'SM': '.MC',  'VX': '.SW',
    'SW': '.SW',  'AV': '.VI',  'SS': '.ST',  'HE': '.HE',
    'DC': '.CO',  'OS': '.OL',  'LI': '.LI',
}


def _yahoo_search_isin(isin, ua):
    """Cerca ISIN: yf.Search con retry, poi OpenFIGI come fallback."""
    results = []

    # 1. yf.Search con retry su rate-limit
    for attempt in range(3):
        try:
            s = yf.Search(isin, max_results=10, news_count=0)
            for q in (s.quotes or []):
                if q.get('quoteType') not in ('EQUITY', 'ETF', 'FUND', 'MUTUALFUND',
                                              'Equity', 'ETF', 'Fund', 'Mutualfund'):
                    continue
                sym  = q.get('symbol', '').strip()
                name = q.get('shortname') or q.get('longname') or sym
                curr = q.get('currency', 'EUR')
                if sym:
                    results.append((sym, name, curr))
            break  # successo, esci dal retry
        except Exception:
            if attempt < 2:
                time.sleep(1.0)

    if results:
        return results

    # 2. OpenFIGI come fallback (Bloomberg ISIN→ticker, gratuito)
    try:
        import requests as _rq
        r = _rq.post(
            'https://api.openfigi.com/v3/mapping',
            json=[{'idType': 'ID_ISIN', 'idValue': isin}],
            headers={'Content-Type': 'application/json'},
            timeout=10,
        )
        for entry in (r.json()[0].get('data') or []):
            ticker_base = entry.get('ticker', '').strip()
            exch        = entry.get('exchCode', '')
            suffix      = _OPENFIGI_EXCH.get(exch, '')
            sym         = ticker_base + suffix if ticker_base else ''
            name        = entry.get('name', sym)
            if sym:
                results.append((sym, name, 'EUR'))
    except Exception:
        pass

    return results


def _best_ticker_from_candidates(candidates, start_date, dl_kwargs):
    """Tra i candidati sceglie quello con più dati storici.
    Con 1 solo candidato lo accetta direttamente (yf.Search già restituisce il migliore).
    Con più candidati fa un batch download per confrontare."""
    if not candidates:
        return None
    # Candidato unico: fidarsi di yf.Search senza download aggiuntivo
    if len(candidates) == 1:
        return candidates[0]
    # Più candidati: confronta con batch download
    symbols = [c[0] for c in candidates[:6]]
    try:
        raw = yf.download(symbols, start=start_date, group_by='ticker',
                          progress=False, **dl_kwargs)
        if raw.empty:
            return candidates[0]
        best_sym, best_n = None, 0
        for sym in symbols:
            try:
                col = (raw[(sym, 'Close')] if isinstance(raw.columns, pd.MultiIndex)
                       else raw['Close'])
                n = int(col.dropna().shape[0])
                if n > best_n:
                    best_n, best_sym = n, sym
            except Exception:
                pass
        if best_sym:
            return next((c for c in candidates if c[0] == best_sym), candidates[0])
    except Exception:
        pass
    return candidates[0]


def _run_isin_conversion(file_bytes, username, req_id):
    """Thread principale: ISIN→ticker, costruisce Excel, avvia download dati."""
    import io as _io

    ua = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
          'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36')
    dl_kwargs  = dict(auto_adjust=True, progress=False)
    start_date = (pd.Timestamp.today() - pd.DateOffset(years=10)).strftime('%Y-%m-%d')

    def _upd(**kw):
        with _ISIN_LOCK:
            if _ISIN_STATE.get('req_id') != req_id:
                return False
            _ISIN_STATE.update(kw)
            return True

    try:
        df = _read_asset_excel(file_bytes)
    except Exception as e:
        _upd(running=False, done=True, progress=f'⚠ Errore lettura file: {e}')
        return

    if df.empty:
        _upd(running=False, done=True, progress='⚠ File vuoto.')
        return

    ticker_col, weight_col, w_vals = _detect_ticker_and_weight_cols(df)

    # Rileva colonna descrizione e valuta per il modo TICKER
    col_names = df.columns.tolist()
    desc_col  = col_names[1] if len(col_names) > 1 else None
    curr_col  = col_names[2] if len(col_names) > 2 else None

    rows = df[ticker_col].dropna().astype(str).str.strip().tolist()
    rows = [r for r in rows if r and r.lower() not in ('nan', '')]

    # Rileva modalità: ISIN o TICKER
    n_isins   = sum(1 for r in rows if _is_isin(r.strip().upper()))
    mode_isin = n_isins > 0   # True = file con ISIN, False = file con ticker
    _upd(n_total=len(rows), n_done=0,
         progress=f'Modalità: {"conversione ISIN" if mode_isin else "validazione ticker"} — {len(rows)} righe')

    converted, excluded = [], []

    for idx, raw_val in enumerate(rows):
        if not _upd(progress=f'Analisi {idx+1}/{len(rows)}: {raw_val}…', n_done=idx):
            return

        val_up = raw_val.strip().upper()

        if _is_isin(val_up):
            # ── Modalità ISIN: cerca ticker su Yahoo ─────────────────────────
            candidates = _yahoo_search_isin(val_up, ua)
            time.sleep(0.4)
            if not candidates:
                excluded.append((raw_val, 'ISIN non trovato su Yahoo Finance'))
                continue
            ticker, name, currency = candidates[0]
        else:
            # ── Modalità TICKER: leggi da file ───────────────────────────────
            ticker = val_up
            try:
                row_df = df[df[ticker_col].astype(str).str.strip().str.upper() == val_up]
                name   = str(row_df[desc_col].iloc[0]).strip() if desc_col and not row_df.empty else ticker
                currency = str(row_df[curr_col].iloc[0]).strip().upper() if curr_col and not row_df.empty else 'EUR'
                if currency.lower() in ('nan', ''):
                    currency = 'EUR'
            except Exception:
                name, currency = ticker, 'EUR'

        # Peso
        peso = None
        if weight_col is not None and w_vals is not None:
            try:
                mask = df[ticker_col].astype(str).str.strip() == raw_val
                idx_match = df[mask].index
                if not idx_match.empty and idx_match[0] in w_vals.index:
                    peso = float(w_vals.loc[idx_match[0]])
            except Exception:
                pass

        converted.append((ticker, name, currency, peso))

    # ── Validazione / correzione ticker (solo modalità TICKER) ───────────────
    # OpenFIGI (Bloomberg) + yf.Search per trovare il simbolo Yahoo corretto
    # con l'estensione di borsa giusta (.MI, .L, .DE…) basata sulla valuta.
    if converted and not mode_isin:
        import requests as _rq

        # Mappa codice exchange OpenFIGI → suffisso Yahoo + valuta tipica
        _EXCH = {
            'LN': ('.L',  'GBP'), 'IM': ('.MI', 'EUR'), 'GY': ('.DE', 'EUR'),
            'FP': ('.PA', 'EUR'), 'NA': ('.AS', 'EUR'), 'BB': ('.BR', 'EUR'),
            'SM': ('.MC', 'EUR'), 'AV': ('.VI', 'EUR'), 'SS': ('.ST', 'SEK'),
            'HE': ('.HE', 'EUR'), 'DC': ('.CO', 'DKK'), 'OS': ('.OL', 'NOK'),
            'SW': ('.SW', 'CHF'), 'US': ('',    'USD'), 'UW': ('',    'USD'),
            'UN': ('',    'USD'), 'UA': ('',    'USD'),
        }
        # Preferenza borsa per valuta
        _PREF = {
            'EUR': ['IM','GY','FP','NA','BB','SM','AV','HE'],
            'GBP': ['LN'],
            'USD': ['UW','UN','US','UA'],
            'CHF': ['SW'], 'SEK': ['SS'], 'DKK': ['DC'], 'NOK': ['OS'],
        }

        def _resolve_ticker(ticker, currency):
            """Trova il simbolo Yahoo corretto tramite OpenFIGI, fallback yf.Search."""
            curr_up = (currency or 'EUR').strip().upper()

            # 1. OpenFIGI: cerca per ticker su tutti gli exchange
            try:
                r = _rq.post(
                    'https://api.openfigi.com/v3/mapping',
                    json=[{'idType': 'TICKER', 'idValue': ticker}],
                    headers={'Content-Type': 'application/json'},
                    timeout=8,
                )
                entries = r.json()[0].get('data') or []
                candidates = []
                for e in entries:
                    t_raw = e.get('ticker', '').strip()
                    exch  = e.get('exchCode', '')
                    if not t_raw or exch not in _EXCH:
                        continue
                    suffix, _ = _EXCH[exch]
                    candidates.append((t_raw + suffix, exch))

                # Scegli la borsa che corrisponde alla valuta del file
                prefs = _PREF.get(curr_up, [])
                for pref_exch in prefs:
                    for sym, exch in candidates:
                        if exch == pref_exch:
                            return sym, currency
                # Fallback: primo candidato OpenFIGI
                if candidates:
                    return candidates[0][0], currency
            except Exception:
                pass

            # 2. yf.Search come fallback
            try:
                s = yf.Search(ticker, max_results=10, news_count=0)
                quotes = [q for q in (s.quotes or [])
                          if q.get('quoteType') in
                          ('EQUITY','ETF','FUND','Mutualfund','Equity','Fund')]
                # Preferisci il quote con la valuta giusta
                for q in quotes:
                    if q.get('currency','').upper() == curr_up:
                        return q['symbol'], q.get('currency', currency)
                if quotes:
                    return quotes[0]['symbol'], quotes[0].get('currency', currency)
            except Exception:
                pass

            return ticker, currency  # nessuna correzione trovata

        ok = []
        n_tot = len(converted)
        for idx_v, (t, d, v, p) in enumerate(converted):
            if not _upd(progress=f'Verifica {idx_v+1}/{n_tot}: {t}…'):
                return

            new_sym, new_curr = _resolve_ticker(t, v)
            time.sleep(0.25)

            # Validazione rapida del simbolo trovato
            valid = [False]
            def _chk(sym=new_sym):
                try:
                    fi = yf.Ticker(sym).fast_info
                    valid[0] = bool(getattr(fi, 'currency', None))
                except Exception:
                    pass
            th = threading.Thread(target=_chk, daemon=True)
            th.start()
            th.join(timeout=6)

            if valid[0]:
                ok.append((new_sym, d, new_curr or v, p))
            else:
                excluded.append((t, f'Ticker non trovato su Yahoo Finance (provato: {new_sym})'))

        converted = ok

    if not _upd(progress='Creazione file Excel…', n_done=len(rows)):
        return

    # ── Excel output ──────────────────────────────────────────────────────────
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Portafoglio'

    hdr_fill = PatternFill('solid', fgColor='1A3A5C')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    hdr_aln  = Alignment(horizontal='center', vertical='center')
    has_pesi = any(c[3] is not None for c in converted)
    headers  = ['TICKER', 'DESCRIZIONE', 'VALUTA'] + (['Peso %'] if has_pesi else [])

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill, c.font, c.alignment = hdr_fill, hdr_font, hdr_aln

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 8
    if has_pesi:
        ws.column_dimensions['D'].width = 10

    alt = PatternFill('solid', fgColor='EEF4FF')
    whi = PatternFill('solid', fgColor='FFFFFF')

    def _wr(r, t, d, v, p):
        f = alt if r % 2 == 0 else whi
        for ci, val in enumerate([t, d, v] + ([round(p, 2) if p is not None else ''] if has_pesi else []), 1):
            ws.cell(r, ci, val).fill = f

    row_n = 2
    # Modalità ISIN: aggiungi default come riferimento
    # Modalità TICKER: solo i ticker dell'utente (già validati)
    if mode_isin:
        for t, d, v in _DEFAULT_PORTFOLIO_ROWS:
            _wr(row_n, t, d, v, None)
            row_n += 1
    for t, d, v, p in converted:
        _wr(row_n, t, d, v, p)
        row_n += 1

    if excluded:
        ws2 = wb.create_sheet('Titoli non trovati')
        ws2.cell(1, 1, 'Valore originale').font = Font(bold=True)
        ws2.cell(1, 2, 'Motivo esclusione').font = Font(bold=True)
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 48
        for i, (v, r) in enumerate(excluded, 2):
            ws2.cell(i, 1, v)
            ws2.cell(i, 2, r)

    buf = _io.BytesIO()
    wb.save(buf)

    tickers = [c[0] for c in converted]
    descrs  = [c[1] for c in converted]
    valutas = [c[2] for c in converted]
    pesi    = [c[3] for c in converted]

    modo_txt = 'convertiti da ISIN' if mode_isin else 'validati'
    _upd(
        running=False, done=True,
        progress=f'✓ {len(converted)} titoli {modo_txt}, {len(excluded)} esclusi.',
        result_bytes=buf.getvalue(),
        tickers=tickers, descr=descrs, valuta=valutas, pesi=pesi,
    )

    if tickers:
        def _download_and_set_weights():
            # Assicura che _DL_BUFFER sia popolato prima del merge
            with _DL_LOCK:
                buf_op = _DL_BUFFER.get('original_prices')
            if buf_op is None and _MARKET_DATA_FILE.exists():
                try:
                    with open(_MARKET_DATA_FILE, 'rb') as _f:
                        _d = pickle.load(_f)
                    with _DL_LOCK:
                        if not _DL_BUFFER.get('original_prices'):
                            _DL_BUFFER.update(_d)
                    buf_op = _d.get('original_prices')
                except Exception:
                    pass
            # all_descr = asset esistenti + nuovi da ISIN (non filtrare via il portafoglio)
            existing_descr = list(buf_op.columns) if buf_op is not None else []
            all_descr = existing_descr + [d for d in descrs if d not in existing_descr]
            _do_gestisci_download(tickers, descrs, valutas, start_date,
                                  username, 'ETF.xlsx', all_descr)
            if any(p is not None for p in pesi):
                w_map = {descrs[i]: pesi[i] for i in range(len(descrs))
                         if pesi[i] is not None}
                _update_user_json(weights={'P1': w_map, 'P2': {}, 'P3': {}},
                                  username=username)
        threading.Thread(target=_download_and_set_weights, daemon=True).start()


# ─────────────────────────────────────────────────────────────────────────────
# Carica solo nomi (avvio rapido)
# ─────────────────────────────────────────────────────────────────────────────
def load_ticker_names_only(filename='ETF.xlsx'):
    try:
        df        = pd.read_excel(_xlsx_path(filename))
        col_names = df.columns.tolist()
        if len(col_names) < 2:
            return [], {}
        tickers     = list(df[col_names[0]])
        descrizione = list(df[col_names[1]]) if len(col_names) > 1 else [str(t) for t in tickers]
        ticker_map  = {descrizione[i]: tickers[i] for i in range(len(tickers))}
        options     = [{'label': d, 'value': d} for d in descrizione]
        print(f"✓ Nomi caricati: {len(options)} asset da {filename}")
        return options, ticker_map
    except Exception as e:
        print(f"Errore lettura nomi: {e}")
        return [], {}


# ─────────────────────────────────────────────────────────────────────────────
# Gestione sessioni
# ─────────────────────────────────────────────────────────────────────────────
SESSIONS_DIR = Path(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sessions'))
SESSIONS_DIR.mkdir(exist_ok=True)
_INDEX_FILE          = SESSIONS_DIR / 'index.json'
_MARKET_DATA_FILE    = SESSIONS_DIR / 'market_data.pkl'
_ACTIVE_TICKERS_FILE = SESSIONS_DIR / 'active_tickers.xlsx'

def _file_cache_path(filename='ETF.xlsx'):
    """Percorso del pkl di cache per un dato file (ETF usa market_data.pkl per compat)."""
    stem = Path(filename).stem
    if stem == 'ETF':
        return _MARKET_DATA_FILE
    return SESSIONS_DIR / f'market_data_{stem}.pkl'


def _arima_cache_path(filename='ETF.xlsx'):
    """Percorso del pkl ARIMA separato dal pkl principale."""
    stem = Path(filename).stem
    if stem == 'ETF':
        return SESSIONS_DIR / 'market_data_arima.pkl'
    return SESSIONS_DIR / f'market_data_{stem}_arima.pkl'


def _load_xlsx_rows(filename):
    """Carica un file xlsx da Files/ come lista di dict per DataTable."""
    try:
        df = pd.read_excel(_xlsx_path(filename))
        cols = df.columns.tolist()
        rows = []
        for _, row in df.iterrows():
            rows.append({
                'ticker':      str(row[cols[0]]),
                'descrizione': str(row[cols[1]]) if len(cols) > 1 else str(row[cols[0]]),
                'valuta':      str(row[cols[2]]) if len(cols) > 2 else 'EUR',
            })
        return rows
    except Exception as e:
        print(f"⚠ _load_xlsx_rows({filename}): {e}")
        return []


def _save_xlsx_rows(filename, rows):
    """Scrive i dati del DataTable nel file Files/{filename} preservando i nomi colonna."""
    try:
        try:
            existing = pd.read_excel(_xlsx_path(filename))
            cols = existing.columns.tolist()
        except Exception:
            cols = ['Ticker', 'Descrizione', 'Valuta']
        c0 = cols[0] if cols else 'Ticker'
        c1 = cols[1] if len(cols) > 1 else 'Descrizione'
        c2 = cols[2] if len(cols) > 2 else 'Valuta'
        df = pd.DataFrame({
            c0: [r.get('ticker', '') for r in rows],
            c1: [r.get('descrizione', '') for r in rows],
            c2: [r.get('valuta', 'EUR') for r in rows],
        })
        target = _FILES_DIR / filename
        df.to_excel(str(target), index=False)
        return True
    except Exception as e:
        print(f"⚠ _save_xlsx_rows({filename}): {e}")
        return False


CLIENT_SESSION_STORES = [
    "weights-store-P1",
    "weights-store-P2",
    "weights-store-P3",
    "global-assets-selected",
    "stock-data",
    "original-prices-data",
    "asset-checklist",
    "ticker-map-store",
    "insufficient-data-store",
]


def _read_index():
    if not _INDEX_FILE.exists():
        return []
    try:
        return json.loads(_INDEX_FILE.read_text(encoding='utf-8'))
    except Exception:
        return []

def _write_index(records):
    _INDEX_FILE.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding='utf-8')

def list_sessions():
    records = _read_index()
    return sorted(records, key=lambda r: r.get('updated_at', ''), reverse=True)

def save_session(name, description, store_data):
    sid  = str(uuid.uuid4())[:8]
    now  = datetime.now().isoformat(timespec='seconds')
    data = json.dumps(store_data, ensure_ascii=False)
    size_kb = round(len(data.encode()) / 1024, 1)
    (SESSIONS_DIR / f'{sid}.json').write_text(data, encoding='utf-8')
    rec = {'id': sid, 'name': name or sid, 'description': description,
           'size_kb': size_kb, 'created_at': now, 'updated_at': now}
    records = _read_index()
    records.append(rec)
    _write_index(records)
    return rec

def load_session(session_id):
    f = SESSIONS_DIR / f'{session_id}.json'
    if not f.exists():
        return {}
    try:
        return json.loads(f.read_text(encoding='utf-8'))
    except Exception:
        return {}

def delete_session(session_id):
    f = SESSIONS_DIR / f'{session_id}.json'
    if f.exists():
        f.unlink()
    records = [r for r in _read_index() if r['id'] != session_id]
    _write_index(records)


def _format_ts(iso_str):
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime('%d/%m/%Y %H:%M')
    except Exception:
        return iso_str[:16]


def _build_session_row(record):
    sid  = record['id']
    name = record.get('name', sid[:8])
    desc = record.get('description', '')
    ts   = _format_ts(record.get('updated_at', record.get('created_at', '')))
    size = record.get('size_kb', '?')
    btn_s = {'border': 'none', 'border-radius': '3px', 'cursor': 'pointer',
             'font-size': '10px', 'padding': '3px 8px', 'font-weight': 'bold'}
    return html.Div([
        html.Div([
            html.Span(name, style={'font-weight': 'bold', 'font-size': '11px',
                                   'color': '#1a3a5c', 'margin-right': '8px'}),
            html.Span(f'({size} KB)', style={'font-size': '9px', 'color': '#999'}),
            html.Br(),
            html.Span(desc[:60] + ('…' if len(desc) > 60 else ''),
                      style={'font-size': '10px', 'color': '#555'}),
            html.Br(),
            html.Span(f'🕐 {ts}', style={'font-size': '9px', 'color': '#888'}),
        ], style={'flex': '1', 'min-width': '0'}),
        html.Div([
            html.Button('📂 Carica',
                        id={'type': 'session-load-btn',   'index': sid},
                        n_clicks=0,
                        style={**btn_s, 'background-color': '#1a3a5c',
                               'color': 'white', 'margin-right': '4px'}),
            html.Button('🗑 Elimina',
                        id={'type': 'session-delete-btn', 'index': sid},
                        n_clicks=0,
                        style={**btn_s, 'background-color': '#c0392b', 'color': 'white'}),
        ], style={'display': 'flex', 'align-items': 'center',
                  'flex-shrink': '0', 'margin-left': '10px'}),
    ], style={'display': 'flex', 'align-items': 'center', 'padding': '6px 8px',
              'border-bottom': '1px solid #eee', 'background': 'white',
              'border-radius': '3px', 'margin-bottom': '3px'})


def get_session_panel_layout():
    btn_base = {'border': 'none', 'border-radius': '4px', 'cursor': 'pointer',
                'font-size': '11px', 'padding': '4px 10px', 'font-weight': 'bold'}
    return html.Div([
        html.Div([
            html.Button('💾 Sessioni', id='session-toggle-btn', n_clicks=0,
                        style={**btn_base, 'background-color': '#1a3a5c',
                               'color': 'white', 'margin-left': '12px',
                               'padding': '6px 14px', 'font-size': '12px'}),
            html.Button('🔄 Default', id='reset-default-btn', n_clicks=0,
                        title='Ricarica asset dai file di default',
                        style={**btn_base, 'background-color': '#5a3a1a',
                               'color': 'white', 'padding': '6px 12px', 'font-size': '12px'}),
        ], style={'display': 'inline-flex', 'align-items': 'center', 'gap': '4px'}),
        html.Div(id='session-panel', style={'display': 'none'}, children=[
            html.Div([
                # Colonna sinistra: salva
                html.Div([
                    html.B('💾 Salva sessione corrente',
                           style={'font-size': '12px', 'color': '#1a3a5c',
                                  'display': 'block', 'margin-bottom': '8px'}),
                    dcc.Input(id='session-name-input', type='text',
                              placeholder='Nome sessione…', debounce=False,
                              style={'width': '100%', 'margin-bottom': '6px',
                                     'padding': '5px 8px', 'border': '1px solid #aaa',
                                     'border-radius': '4px', 'font-size': '11px'}),
                    dcc.Textarea(id='session-desc-input',
                                 placeholder='Note / descrizione (opzionale)…', rows=2,
                                 style={'width': '100%', 'margin-bottom': '6px',
                                        'padding': '5px 8px', 'border': '1px solid #aaa',
                                        'border-radius': '4px', 'font-size': '11px',
                                        'resize': 'none'}),
                    html.Button('💾 Salva', id='session-save-btn', n_clicks=0,
                                style={**btn_base, 'background-color': '#1b7a34',
                                       'color': 'white', 'width': '100%'}),
                    html.Div(id='session-save-status',
                             style={'font-size': '10px', 'margin-top': '5px',
                                    'color': '#555', 'min-height': '16px'}),
                ], style={'width': '30%', 'padding-right': '20px',
                          'border-right': '1px solid #ddd'}),
                # Colonna destra: elenco
                html.Div([
                    html.Div([
                        html.B('📂 Sessioni salvate',
                               style={'font-size': '12px', 'color': '#1a3a5c'}),
                        html.Button('🔄', id='session-refresh-btn', n_clicks=0,
                                    title='Aggiorna elenco',
                                    style={**btn_base, 'background-color': '#e8e8e8',
                                           'color': '#333', 'margin-left': '8px',
                                           'padding': '3px 8px'}),
                    ], style={'display': 'flex', 'align-items': 'center',
                              'margin-bottom': '8px'}),
                    html.Div(id='session-list-container',
                             style={'max-height': '240px', 'overflow-y': 'auto'}),
                    dcc.Store(id='session-load-trigger',   data=None),
                    dcc.Store(id='session-delete-trigger', data=None),
                    dcc.Store(id='session-selected-id',    data=None),
                ], style={'flex': '1', 'padding-left': '20px'}),
            ], style={'display': 'flex'}),
        ]),
    ], style={'display': 'inline-block'})


# ─────────────────────────────────────────────────────────────────────────────
# File panel — selettore default + file personali + salva con nome
# ─────────────────────────────────────────────────────────────────────────────
def get_file_panel_layout():
    _bb = {'border': 'none', 'border-radius': '4px', 'cursor': 'pointer',
           'font-size': '11px', 'padding': '4px 10px', 'font-weight': 'bold'}
    return html.Div([
        html.Button('📁 File', id='file-panel-btn', n_clicks=0,
                    style={**_bb, 'background-color': '#5a1a6a', 'color': 'white',
                           'padding': '6px 14px', 'font-size': '12px'}),
        html.Div(id='file-panel', style={'display': 'none'}, children=[
            html.Div([
                # ── Colonna sinistra: Salva (i default sono nel selettore esterno) ──
                html.Div([
                    html.B('💾 Salva sessione come…',
                           style={'font-size': '11px', 'color': '#1a3a5c',
                                  'display': 'block', 'margin-bottom': '8px'}),
                    dcc.Input(id='fp-save-name', type='text',
                              placeholder='Es. Mario_Cliente_ABC…',
                              style={'width': '100%', 'padding': '5px 8px',
                                     'border': '1px solid #aaa', 'border-radius': '4px',
                                     'font-size': '11px', 'margin-bottom': '6px'}),
                    html.Button('💾 Salva', id='fp-save-btn', n_clicks=0,
                                style={**_bb, 'background': '#1b7a34',
                                       'color': 'white', 'width': '100%'}),
                    html.Div(id='fp-save-status',
                             style={'font-size': '10px', 'margin-top': '5px',
                                    'color': '#555', 'min-height': '16px'}),
                ], style={'width': '220px', 'padding-right': '20px',
                          'border-right': '1px solid #ddd'}),
                # ── Colonna destra: file personali ────────────────────────
                html.Div([
                    html.Div([
                        html.B('📁 I miei file',
                               style={'font-size': '11px', 'color': '#1a3a5c'}),
                        html.Button('🔄', id='fp-refresh-btn', n_clicks=0,
                                    style={**_bb, 'background': '#e8e8e8',
                                           'color': '#333', 'margin-left': '8px',
                                           'padding': '3px 8px'}),
                    ], style={'display': 'flex', 'align-items': 'center',
                              'margin-bottom': '8px'}),
                    html.Div(id='fp-file-list',
                             style={'max-height': '280px', 'overflow-y': 'auto'}),
                ], style={'flex': '1', 'padding-left': '20px'}),
            ], style={'display': 'flex'}),
            # Footer: pulsante Chiudi in basso a destra
            html.Div([
                html.Button('Chiudi', id='fp-close-btn', n_clicks=0,
                            style={**_bb, 'background-color': '#5a1a6a', 'color': 'white',
                                   'padding': '7px 20px'}),
            ], style={'display': 'flex', 'justify-content': 'flex-end', 'margin-top': '14px',
                      'border-top': '1px solid #eee', 'padding-top': '10px'}),
        ]),
    ], style={'display': 'inline-block', 'position': 'relative'})


# ─────────────────────────────────────────────────────────────────────────────
# Date range bar
# ─────────────────────────────────────────────────────────────────────────────
def get_date_range_bar(suffix):
    _today         = pd.Timestamp.today().normalize()
    _ten_years_ago = (_today - pd.DateOffset(years=10)).strftime('%Y-%m-%d')
    _today_str     = _today.strftime('%Y-%m-%d')
    return html.Div([
        html.Div([
            html.Div('📅 Range temporale:',
                     style={'font-size': '11px', 'font-weight': 'bold',
                            'color': '#1a3a5c', 'white-space': 'nowrap',
                            'margin-right': '10px', 'align-self': 'center'}),
            html.Div('da:', style={'font-size': '11px', 'color': '#555',
                                   'align-self': 'center', 'white-space': 'nowrap',
                                   'margin-right': '4px'}),
            dcc.DatePickerSingle(id=f'dr-start-{suffix}', display_format='DD/MM/YYYY',
                                 first_day_of_week=1, date=_ten_years_ago,
                                 clearable=False, style={'width': '135px'}),
            html.Div('a:', style={'font-size': '11px', 'color': '#555',
                                  'align-self': 'center', 'white-space': 'nowrap',
                                  'margin': '0 4px'}),
            dcc.DatePickerSingle(id=f'dr-end-{suffix}', display_format='DD/MM/YYYY',
                                 first_day_of_week=1, date=_today_str,
                                 clearable=False, style={'width': '135px'}),
            html.Div(id=f'dr-label-{suffix}',
                     style={'font-size': '10px', 'color': '#888', 'white-space': 'nowrap',
                            'align-self': 'center', 'margin-left': '10px',
                            'min-width': '200px'}),
        ], style={'display': 'flex', 'align-items': 'center', 'padding': '6px 12px',
                  'background': '#f0f4fa', 'border': '1px solid #d0d8e8',
                  'border-radius': '6px', 'margin-bottom': '8px', 'gap': '4px'}),
    ])


# ─────────────────────────────────────────────────────────────────────────────
# Layout Tab — Matrice di Correlazione (portata da ir_fe_14.py, adattata)
# ─────────────────────────────────────────────────────────────────────────────
def get_correlation_matrix_tab(options_tickers=None):
    """Heatmap di correlazione (rendimenti settimanali) + correlazione rolling."""
    return html.Div([
        # ── Riga controlli ───────────────────────────────────────────────────
        html.Div([
            get_date_range_bar('corr'),
            html.Div([
                html.Label("Finestra (giorni):", style={'margin-right': '6px', 'font-size': '12px', 'white-space': 'nowrap'}),
                dcc.Input(id='correlation-window-input', type='number', value=252, min=10,
                          placeholder='Giorni', style={'width': '80px', 'margin-right': '14px'}),
                html.Label("Benchmark:", style={'margin-right': '6px', 'font-size': '12px', 'white-space': 'nowrap'}),
                dcc.Dropdown(id='benchmark-selector-corr', options=[], value=None,
                             placeholder='Seleziona benchmark…', clearable=False,
                             style={'width': '200px', 'font-size': '11px', 'margin-right': '14px'}),
                html.Button('Update', id='update-correlation-button', n_clicks=0,
                            style={'background-color': '#28a745', 'color': 'white', 'border': 'none',
                                   'padding': '7px 18px', 'border-radius': '4px', 'cursor': 'pointer',
                                   'font-weight': 'bold', 'font-size': '12px'}),
                html.Div(id='corr-filter-info',
                         style={'font-size': '11px', 'color': '#0066cc', 'margin-left': '16px', 'align-self': 'center'}),
            ], style={'display': 'flex', 'align-items': 'center', 'margin-bottom': '6px', 'flex-wrap': 'wrap'}),
            html.Div(id='date-values-correlation',
                     style={'font-size': '11px', 'color': '#666', 'margin-bottom': '4px'}),
        ], style={'padding': '8px 12px 0 12px'}),

        # ── Corpo a due colonne ──────────────────────────────────────────────
        html.Div([
            # Colonna sinistra: griglia asset con spunte MT (matrice) e ML (rolling)
            html.Div([
                html.Div('Asset da confrontare',
                         style={'font-weight': 'bold', 'font-size': '11px', 'color': '#1a3a5c',
                                'margin-bottom': '2px', 'text-align': 'center'}),
                html.Div('MT = matrice · ML = rolling',
                         style={'font-size': '9px', 'color': '#888', 'margin-bottom': '8px', 'text-align': 'center'}),
                # Intestazione colonne con pulsanti seleziona/deseleziona tutti
                html.Div([
                    html.Div('', style={'flex': '1'}),
                    html.Div([
                        html.Div('MT', style={'font-size': '10px', 'font-weight': 'bold', 'color': '#1a3a5c'}),
                        html.Button('☑', id='corr-mt-all', n_clicks=0,
                                    title='Seleziona/deseleziona tutti — Matrice',
                                    style={'font-size': '11px', 'padding': '0 3px', 'border': '1px solid #ccc',
                                           'background': '#eef4ff', 'borderRadius': '3px', 'cursor': 'pointer',
                                           'lineHeight': '14px'}),
                    ], style={'width': '34px', 'display': 'flex', 'flexDirection': 'column',
                              'alignItems': 'center', 'gap': '2px'}),
                    html.Div([
                        html.Div('ML', style={'font-size': '10px', 'font-weight': 'bold', 'color': '#e6550d'}),
                        html.Button('☑', id='corr-ml-all', n_clicks=0,
                                    title='Seleziona/deseleziona tutti — Rolling',
                                    style={'font-size': '11px', 'padding': '0 3px', 'border': '1px solid #ccc',
                                           'background': '#fff3e8', 'borderRadius': '3px', 'cursor': 'pointer',
                                           'lineHeight': '14px'}),
                    ], style={'width': '34px', 'display': 'flex', 'flexDirection': 'column',
                              'alignItems': 'center', 'gap': '2px'}),
                ], style={'display': 'flex', 'align-items': 'flex-end', 'gap': '4px',
                          'padding': '0 2px 6px 2px', 'borderBottom': '1px solid #e0e0e0',
                          'margin-bottom': '4px'}),
                html.Div(id='corr-asset-grid', children=[
                    html.Div('(carica i dati per popolare)',
                             style={'font-size': '10px', 'color': '#aaa', 'font-style': 'italic',
                                    'text-align': 'center', 'padding': '10px 4px'})
                ]),
            ], style={'width': '24%', 'min-width': '180px', 'padding': '10px 8px 10px 12px',
                      'border-right': '1px solid #e0e0e0', 'overflow-y': 'auto',
                      'max-height': '88vh', 'box-sizing': 'border-box'}),

            # Colonna destra: heatmap + legenda + rolling chart
            html.Div([
                dcc.Loading(id='loading-correlation', type='circle', children=[
                    dcc.Graph(id='correlation-heatmap', style={'width': '100%', 'height': '48vh'},
                              config={'responsive': True})]),
                html.Div([
                    html.Div('Legenda: ', style={'font-weight': 'bold', 'margin-right': '12px', 'font-size': '11px'}),
                    html.Div('Verde ≥0.70', style={'background': '#2ca02c', 'color': 'white', 'padding': '3px 8px',
                                                   'margin': '2px', 'border-radius': '3px', 'font-size': '10px'}),
                    html.Div('Bianco [-0.50, 0.70)', style={'background': '#e0e0e0', 'color': 'black', 'padding': '3px 8px',
                                                            'margin': '2px', 'border-radius': '3px', 'font-size': '10px'}),
                    html.Div('Rosso ≤-0.50', style={'background': '#d62728', 'color': 'white', 'padding': '3px 8px',
                                                    'margin': '2px', 'border-radius': '3px', 'font-size': '10px'}),
                ], style={'display': 'flex', 'align-items': 'center', 'margin': '4px 0 6px 0', 'flex-wrap': 'wrap'}),
                html.Hr(style={'margin': '6px 0'}),
                html.Div('Correlazione Rolling',
                         style={'font-weight': 'bold', 'font-size': '12px', 'color': '#1a3a5c', 'margin-bottom': '4px'}),
                dcc.Loading(id='loading-rolling-corr', type='circle', children=[
                    dcc.Graph(id='rolling-corr-chart', style={'width': '100%', 'height': '35vh'},
                              config={'responsive': True})]),
            ], style={'width': '76%', 'padding': '4px 8px', 'box-sizing': 'border-box'}),
        ], style={'display': 'flex', 'align-items': 'flex-start'}),
    ])


# ─────────────────────────────────────────────────────────────────────────────
# Layout Tab 1
# ─────────────────────────────────────────────────────────────────────────────
def get_portfolio_analysis_tab(options_tickers):
    return html.Div([
        html.Div([
            # ── Riga controlli ────────────────────────────────────────────
            html.Div([
                html.Span('Parametri indicatori di rischio:',
                          style={'margin-right': '10px', 'white-space': 'nowrap',
                                 'font-size': '12px', 'font-weight': '700',
                                 'color': '#1a3a5c'}),
                html.Div([
                    html.Div([
                        html.Label('Benchmark:',
                                   style={'margin-right': '4px', 'white-space': 'nowrap',
                                          'font-size': '10px'}),
                        dcc.Dropdown(
                            options=options_tickers or [],
                            value=None,
                            id='benchmark-selector',
                            placeholder='Seleziona…',
                            style={'width': '180px', 'font-size': '10px', 'min-width': '180px'}
                        ),
                    ], style={'margin-right': '10px', 'display': 'flex', 'align-items': 'center'}),
                    html.Div([
                        html.Label('AK-SH-TV W:',
                                   style={'margin-right': '4px', 'white-space': 'nowrap',
                                          'font-size': '10px'}),
                        dcc.Input(id='ir-window-input', type='number', value=30, min=1,
                                  placeholder='30', style={'width': '45px', 'font-size': '10px'}),
                    ], style={'display': 'flex', 'align-items': 'center', 'margin-right': '8px'}),
                    html.Div([
                        html.Label('MA:',
                                   style={'margin-right': '4px', 'white-space': 'nowrap',
                                          'font-size': '10px'}),
                        dcc.Input(id='ak-ma-input', type='number', value=1, min=1,
                                  placeholder='1', style={'width': '40px', 'font-size': '10px'}),
                    ], style={'display': 'flex', 'align-items': 'center', 'margin-right': '8px'}),
                    html.Div([
                        html.Label('VOL-cVAR W:',
                                   style={'margin-right': '4px', 'white-space': 'nowrap',
                                          'font-size': '10px'}),
                        dcc.Input(id='vol-window-input', type='number', value=30, min=1,
                                  placeholder='30', style={'width': '45px', 'font-size': '10px'}),
                    ], style={'display': 'flex', 'align-items': 'center', 'margin-right': '8px'}),
                    html.Button('Update', id='update-portfolio-button', n_clicks=0,
                                style={'background-color': '#c0392b', 'color': 'white',
                                       'border': 'none', 'padding': '4px 10px',
                                       'border-radius': '4px', 'cursor': 'pointer',
                                       'font-weight': 'bold', 'font-size': '10px',
                                       'box-shadow': '0 2px 6px rgba(192,57,43,0.4)'}),
                    html.Div([
                        html.Label('Filtro AKR:',
                                   style={'font-size': '9px', 'font-weight': 'bold',
                                          'color': '#1a3a5c', 'margin-right': '4px',
                                          'white-space': 'nowrap'}),
                        dcc.RadioItems(
                            id='ir-filter-radio',
                            options=[
                                {'label': 'Tutti', 'value': 'all'},
                                {'label': '>−1',   'value': 'gt_minus1'},
                                {'label': '>0',    'value': 'gt_0'},
                            ],
                            value='all', inline=True,
                            inputStyle={'margin-right': '2px'},
                            labelStyle={'margin-right': '6px', 'font-size': '9px',
                                        'cursor': 'pointer'},
                        ),
                    ], style={'display': 'flex', 'align-items': 'center',
                              'margin-left': '8px', 'padding': '2px 6px',
                              'background': '#f0f4fa', 'border': '1px solid #d0d8e8',
                              'border-radius': '5px'}),
                ], style={'display': 'flex', 'align-items': 'center', 'flex-wrap': 'wrap', 'gap': '0px'}),
            ], style={'margin-bottom': '8px', 'display': 'flex', 'align-items': 'center',
                      'flex-wrap': 'wrap', 'padding': '5px 8px',
                      'background': '#f8fafc', 'border': '1px solid #e2e8f0',
                      'border-radius': '6px'}),

            html.Hr(),

            # ── Intestazione full-width: etichette colonne (sx) + range date (dx) ──
            html.Div([
                # Sx 35%: etichette colonne con pulsanti ☑ integrati
                html.Div([
                    html.Div('Asset', **{'data-tooltip': 'Nome dell\'asset'}, style={
                        'width': '19%', 'fontWeight': 'bold', 'fontSize': '6px',
                        'paddingLeft': '4px', 'color': '#1a3a5c',
                        'display': 'flex', 'alignItems': 'center',
                        'overflow': 'hidden',
                    }),
                    *[html.Div([
                        html.Span(lbl, style={
                            'fontWeight': 'bold', 'fontSize': '7px', 'color': col,
                            'lineHeight': '1', 'whiteSpace': 'nowrap',
                        }),
                        html.Button('☑', id=btn_id, n_clicks=0, title=tip,
                            style={'fontSize': '7px', 'border': 'none', 'background': 'none',
                                   'cursor': 'pointer', 'color': col,
                                   'padding': '0', 'margin': '0', 'lineHeight': '1'}),
                    ], className='col-header-cell', style={
                        'width': w, 'display': 'flex', 'flexDirection': 'column',
                        'alignItems': 'center', 'justifyContent': 'center',
                        'overflow': 'hidden', 'gap': '0px',
                    }) for w, lbl, col, tip, btn_id in [
                        ('3%',   'CH',   '#1a3a5c', 'Deseleziona grafici',     'deselect-all-tickers'),
                        ('12%',  'P1',   '#e6194b', 'Azzera pesi P1',          'reset-p1-tab1'),
                        ('12%',  'P2',   '#3cb44b', 'Azzera pesi P2',          'reset-p2-tab1'),
                        ('12%',  'P3',   '#4363d8', 'Azzera pesi P3',          'reset-p3-tab1'),
                        ('5%',   'AKR',  '#1a3a5c', 'Deseleziona AKRatio',     'deselect-all-ir'),
                        ('5%',   'SH',   '#1a3a5c', 'Deseleziona Sharpe',      'deselect-all-sharpe'),
                        ('5%',   'TV',   '#1a3a5c', 'Deseleziona TEV',         'deselect-all-tev'),
                        ('5%',   'DD',   '#1a3a5c', 'Deseleziona DrawDown',    'deselect-all-dd'),
                        ('5%',   'VOL',  '#1a3a5c', 'Deseleziona Volatilità',  'deselect-all-vol'),
                        ('5%',   'VA90', '#1a3a5c', 'Deseleziona VaR 90%',     'deselect-all-var90'),
                        ('5%',   'VA95', '#1a3a5c', 'Deseleziona VaR 95%',     'deselect-all-var95'),
                    ]],
                ], style={
                    'width': '35%', 'display': 'flex', 'alignItems': 'center',
                    'minHeight': '18px', 'padding': '0', 'overflow': 'hidden',
                }),
                # Dx 65%: range temporale — allineato a destra
                html.Div(
                    get_date_range_bar('tab1'),
                    style={'width': '65%', 'display': 'flex', 'align-items': 'center',
                           'justify-content': 'flex-end'},
                ),
            ], style={
                'display': 'flex', 'background': '#eaf4fb',
                'border-top': '2px solid #2e6da4', 'border-bottom': '1px solid #aed6f1',
            }),

            # ── Riga dati: griglia asset (sx 35%) + grafico (dx 65%) ─────
            html.Div([
                # Colonna sinistra: pulsanti Des + righe asset (generati dal callback)
                html.Div([
                    html.Div(id='asset-count-display',
                             style={'font-size': '10px', 'color': '#555',
                                    'padding': '3px 5px 5px 5px', 'margin-bottom': '2px'}),
                    html.Div(
                        '▶ Dati caricati — clicca UPDATE per aggiornare i grafici',
                        id='update-hint',
                        style={'display': 'none', 'font-size': '9px', 'color': '#c0392b',
                               'font-weight': '600', 'padding': '2px 5px 4px 5px',
                               'background': '#fdf2f0', 'border-left': '3px solid #c0392b',
                               'margin-bottom': '4px', 'border-radius': '0 4px 4px 0'}
                    ),
                    html.Div(id='weights-grid-container', style={'display': 'block'}),
                    html.Div([
                        dcc.Input(id='inline-ticker-input', placeholder='Ticker (es. AAPL)',
                                  style={'width': '90px', 'fontSize': '9px', 'padding': '2px 4px',
                                         'border': '1px solid #ccc', 'borderRadius': '3px'}),
                        dcc.Input(id='inline-desc-input', placeholder='Descrizione',
                                  style={'width': '140px', 'fontSize': '9px', 'padding': '2px 4px',
                                         'border': '1px solid #ccc', 'borderRadius': '3px'}),
                        dcc.Dropdown(id='inline-valuta-dropdown',
                                     options=[{'label': 'EUR', 'value': 'EUR'},
                                              {'label': 'USD', 'value': 'USD'},
                                              {'label': 'GBP', 'value': 'GBP'}],
                                     value='EUR', clearable=False,
                                     style={'width': '65px', 'fontSize': '9px',
                                            'display': 'inline-block'}),
                        html.Button('➕', id='inline-add-btn', n_clicks=0,
                                    style={'fontSize': '10px', 'padding': '2px 7px',
                                           'background': '#e8f5e9', 'border': '1px solid #a5d6a7',
                                           'color': '#1b5e20', 'borderRadius': '3px',
                                           'cursor': 'pointer'}),
                        html.Span(id='inline-add-status',
                                  style={'fontSize': '9px', 'color': '#555', 'marginLeft': '4px'}),
                    ], style={'display': 'flex', 'alignItems': 'center', 'gap': '3px',
                              'padding': '4px 2px', 'marginTop': '4px'}),
                    html.Div([
                        html.Hr(style={'margin': '10px 0'}),
                        html.Div([
                            html.Div('Totale Pesi:',
                                     style={'width': '22%', 'font-weight': 'bold',
                                            'padding-left': '4px', 'font-size': '9px'}),
                            html.Div(id='sum-p1-display', children='0%',
                                     style={'width': '12%', 'text-align': 'right',
                                            'padding-right': '5px', 'font-weight': 'bold',
                                            'color': '#d62728', 'font-size': '10px',
                                            'box-sizing': 'border-box'}),
                            html.Div(id='sum-p2-display', children='0%',
                                     style={'width': '12%', 'text-align': 'right',
                                            'padding-right': '5px', 'font-weight': 'bold',
                                            'color': '#d62728', 'font-size': '10px',
                                            'box-sizing': 'border-box'}),
                            html.Div(id='sum-p3-display', children='0%',
                                     style={'width': '12%', 'text-align': 'right',
                                            'padding-right': '5px', 'font-weight': 'bold',
                                            'color': '#d62728', 'font-size': '10px',
                                            'box-sizing': 'border-box'}),
                            html.Div('', style={'width': '42%'}),
                        ], style={'display': 'flex', 'align-items': 'center'}),
                    ], style={'margin-top': '10px'}),
                ], style={'width': '35%', 'vertical-align': 'top'}),

                # Colonna destra: grafico
                html.Div([
                    dcc.Loading(
                        id='graph-loading',
                        target_components={'controls-and-graph': 'figure'},
                        overlay_style={'visibility': 'visible',
                                       'background': 'rgba(248,250,255,0.85)'},
                        custom_spinner=html.Div([
                            html.Div(style={
                                'width': '44px', 'height': '44px',
                                'border': '4px solid #d0ddf0',
                                'borderTop': '4px solid #1a3a6b',
                                'borderRadius': '50%',
                                'animation': 'spin 0.8s linear infinite',
                            }),
                            html.Div('Calcolo in corso…', style={
                                'marginTop': '14px', 'color': '#1a3a6b',
                                'fontWeight': '700', 'fontSize': '13px',
                                'fontFamily': 'Inter, sans-serif',
                                'letterSpacing': '0.02em',
                            }),
                        ], style={'display': 'flex', 'flexDirection': 'column',
                                  'alignItems': 'center', 'padding': '60px'}),
                        children=html.Div(
                            dcc.Graph(id='controls-and-graph',
                                      style={'width': '100%', 'height': '1900px',
                                             'margin': '0', 'padding': '0'},
                                      config={'responsive': True}),
                            style={'overflow-y': 'auto', 'max-height': '82vh',
                                   'width': '100%', 'margin-bottom': '-10px'},
                        ),
                    ),
                    html.Div(id='output', style={'display': 'none'}),
                    html.Div(id='date-values', style={'display': 'none'}),
                    dcc.Textarea(id='selected-column',           value='', style={'display': 'none'}),
                    dcc.Textarea(id='insufficient-data-tickers', value='', style={'display': 'none'}),
                ], style={'width': '65%', 'vertical-align': 'top'}),
            ], style={'display': 'flex', 'border-bottom': '2px solid #2e6da4'}),
        ]),
    ])


# ─────────────────────────────────────────────────────────────────────────────
# Stili modale progresso
# ─────────────────────────────────────────────────────────────────────────────
_MODAL_HIDDEN = {
    'display': 'none', 'position': 'fixed', 'top': '0', 'left': '0',
    'width': '100%', 'height': '100%', 'background': 'rgba(26,58,92,0.45)',
    'zIndex': '2000', 'justifyContent': 'center', 'alignItems': 'center',
}
_MODAL_SHOWN = {**_MODAL_HIDDEN, 'display': 'flex'}

_EDITOR_HIDDEN = {'display': 'none'}
_EDITOR_SHOWN  = {
    'display': 'flex', 'position': 'fixed', 'top': '0', 'left': '0',
    'width': '100%', 'height': '100%', 'background': 'rgba(0,0,0,0.5)',
    'zIndex': '3000', 'justifyContent': 'center', 'alignItems': 'center',
}

_FILL_LOADING = {
    'height': '100%', 'width': '0%',
    'background': 'linear-gradient(90deg,#007755,#00aa77)',
    'borderRadius': '8px', 'transition': 'width 0.5s ease',
}
_STATUS_GREY  = {'fontSize': '0.78rem', 'color': '#6b7a99',
                 'fontFamily': 'Inter, sans-serif', 'textAlign': 'center',
                 'minHeight': '20px', 'marginTop': '6px'}
_STATUS_GREEN = {**_STATUS_GREY, 'color': '#007755', 'fontWeight': '600'}
_STATUS_RED   = {**_STATUS_GREY, 'color': '#c0392b', 'fontWeight': '600'}

# Etichetta gruppo + separatore verticale per la barra comandi
_GRP_LABEL = {'font-size': '9px', 'font-weight': '700', 'color': '#8a96a8',
              'letter-spacing': '0.05em', 'text-transform': 'uppercase',
              'margin': '0 6px 0 2px'}
_GRP_SEP   = {'width': '1px', 'height': '26px', 'background': '#d0d8e4',
              'margin': '0 10px'}


# ─────────────────────────────────────────────────────────────────────────────
# Layout principale
# ─────────────────────────────────────────────────────────────────────────────
def _navbar():
    from navbar import make_navbar
    return make_navbar(current='portafoglio')
app.layout = html.Div([
    # ── Navbar ───────────────────────────────────────────────────────────────
    _navbar(),

    # ── Contenuto (margine top per navbar fissa 64px) ────────────────────────
    html.Div([

    # ── Intestazione pagina ───────────────────────────────────────────────────
    html.Div([
        html.H1([
            'Analisi Rischi di Portafoglio',
            html.Span(' - ', style={'color': '#9baabf'}),
            html.Span('Analisi delle metriche dei rischi Finanziari', style={
                'font-size': '1.1rem', 'font-weight': '400', 'color': '#4a5d7a',
            }),
        ], style={
            'margin': '0',
            'font-size': '1.6rem',
            'font-weight': '700',
            'color': '#1a3a6b',
            'font-family': "'Playfair Display', serif",
            'letter-spacing': '0.02em',
        }),
    ], style={
        'padding': '14px 20px 12px',
        'border-bottom': '2px solid #e2e8f0',
        'background': 'linear-gradient(90deg, #f0f4fb 0%, #ffffff 100%)',
        'margin-bottom': '10px',
    }),

    # ── Barra comandi ─────────────────────────────────────────────────────────
    html.Div(id='cmd-bar', children=[
        # ── Controlli legacy nascosti (callback ancora attivi) ────────────────
        html.Div([
            dcc.Loading(type='circle', color='#007755', children=[
                html.Button('⟳ Aggiorna', id='refresh-data-btn', n_clicks=0,
                            title='Forza nuovo download da Yahoo Finance',
                            style={'background-color': '#007755', 'color': 'white',
                                   'border': 'none', 'padding': '7px 16px',
                                   'border-radius': '4px', 'cursor': 'pointer',
                                   'font-weight': 'bold', 'font-size': '12px'}),
            ]),
            html.Span(id='data-last-updated'),
            html.Button('✏️ Gestisci', id='gestisci-btn', n_clicks=0),
            get_session_panel_layout(),
            html.Button(id='delete-column-button', n_clicks=0),
            html.Div(id='upload-status'),
            # Importa Frontiera: legacy, sostituito da Importa/Esporta Portafoglio
            html.Button(id='import-frontier-btn', n_clicks=0),
            html.Div(id='import-frontier-msg'),
            dcc.ConfirmDialog(id='import-frontier-confirm',
                              message='Sovrascrivere i pesi P1/P2/P3 con quelli della Frontiera?'),
        ], style={'display': 'none'}),

        # ── Selettore file ESTERNO: default (ETF/Cripto/Commodities) o file personale ──
        html.Div([
            dcc.Dropdown(id='file-selector', options=_list_files(),
                         value='ETF.xlsx', clearable=False,
                         style={'width': '200px', 'font-size': '11px'}, optionHeight=28),
        ], style={'display': 'flex', 'align-items': 'center', 'margin-right': '12px'}),

        # ── Pulsanti (etichette/separatori rimossi; ordine da definire) ───────
        get_file_panel_layout(),

        dcc.Upload(
            id='upload-data',
            children=html.Div(['⬆ Carica file']),
            style={'height': '30px', 'lineHeight': '30px', 'padding': '0 12px',
                   'borderWidth': '1px', 'borderStyle': 'dashed', 'borderColor': '#9bb0cc',
                   'borderRadius': '4px', 'textAlign': 'center', 'fontSize': '11px',
                   'color': '#1a3a5c', 'background': '#f5f8fc', 'cursor': 'pointer',
                   'margin-right': '4px', 'display': 'inline-block'},
            multiple=False,
        ),
        html.Button('🔀 Converti ISIN', id='isin-open-btn', n_clicks=0,
                    title='Carica un file con ISIN e converti in ticker Yahoo Finance',
                    style={'font-size': '11px', 'padding': '5px 12px', 'border-radius': '4px',
                           'cursor': 'pointer', 'background': '#fff3e0',
                           'border': '1px solid #ffb74d', 'color': '#e65100',
                           'font-weight': 'bold', 'margin-right': '4px'}),
        html.Button('📋 Template', id='btn-download-template', n_clicks=0,
                    title='Scarica il file Excel template da compilare con i tuoi titoli',
                    style={'font-size': '11px', 'padding': '5px 12px',
                           'border-radius': '4px', 'cursor': 'pointer',
                           'background': '#e8f5e9', 'border': '1px solid #a5d6a7',
                           'color': '#1b5e20'}),

        html.Button('🔄 Importa/Esporta Portafoglio', id='port-io-btn', n_clicks=0,
                    title='Salva i portafogli P1/P2/P3 in un profilo o importane di salvati',
                    style={'font-size': '11px', 'padding': '5px 12px',
                           'border-radius': '4px', 'cursor': 'pointer',
                           'background': '#eafaf1', 'border': '1px solid #1a7a4a',
                           'color': '#1a7a4a', 'font-weight': 'bold', 'margin-right': '4px'}),
        html.Button('📥 Esporta prezzi', id='save-data-button', n_clicks=0,
                    title='Scarica i prezzi correnti come file Excel (date + prezzi per asset)',
                    style={'font-size': '11px', 'padding': '5px 12px',
                           'border-radius': '4px', 'cursor': 'pointer',
                           'background': '#f0f4fb', 'border': '1px solid #c0d0e8',
                           'color': '#1a3a5c', 'margin-right': '4px'}),
        html.Div(id='download-status', style={'font-size': '11px', 'margin-right': '8px'}),
    ], style={'display': 'flex', 'align-items': 'center',
              'font-size': '10px', 'position': 'relative',
              'padding': '6px 0', 'flex-wrap': 'wrap', 'gap': '2px'}),

    # ── Modal Importa/Esporta Portafoglio ─────────────────────────────────────
    html.Div(id='pio-overlay',
             style={'display': 'none', 'position': 'fixed', 'top': '0', 'left': '0',
                    'width': '100%', 'height': '100%', 'z-index': '9000',
                    'background': 'rgba(0,0,0,0.45)', 'align-items': 'center',
                    'justify-content': 'center'},
             children=[
        html.Div([
            html.Div([
                html.Span('🔄 Importa / Esporta Portafoglio',
                          style={'font-weight': '700', 'font-size': '14px', 'color': '#1a3a5c'}),
                html.Button('✕', id='pio-close', n_clicks=0,
                            style={'background': 'none', 'border': 'none', 'font-size': '18px',
                                   'cursor': 'pointer', 'color': '#666', 'float': 'right'}),
            ], style={'display': 'flex', 'justify-content': 'space-between',
                      'align-items': 'center', 'margin-bottom': '12px'}),

            dcc.RadioItems(id='pio-mode',
                           options=[{'label': ' 📤 Esporta', 'value': 'export'},
                                    {'label': ' 📥 Importa', 'value': 'import'}],
                           value='export', inline=True,
                           inputStyle={'margin-right': '4px'},
                           labelStyle={'margin-right': '18px', 'font-size': '12px',
                                       'font-weight': '600'},
                           style={'margin-bottom': '14px', 'padding-bottom': '10px',
                                  'border-bottom': '1px solid #eee'}),

            # ── ESPORTA ───────────────────────────────────────────────────────
            html.Div(id='pio-export-view', children=[
                html.Div('1. Colonna da esportare:',
                         style={'font-size': '11px', 'font-weight': '600',
                                'color': '#1a3a5c', 'margin-bottom': '6px'}),
                dcc.RadioItems(id='pio-exp-col',
                               options=[{'label': ' P1', 'value': 'P1'},
                                        {'label': ' P2', 'value': 'P2'},
                                        {'label': ' P3', 'value': 'P3'}],
                               value='P1', inline=True,
                               inputStyle={'margin-right': '4px'},
                               labelStyle={'margin-right': '16px', 'font-size': '12px',
                                           'font-weight': '600'},
                               style={'margin-bottom': '12px'}),
                html.Div('2. Salva come Analisi (nuova o sovrascrivi esistente):',
                         style={'font-size': '11px', 'font-weight': '600',
                                'color': '#1a3a5c', 'margin-bottom': '6px'}),
                html.Div([
                    dcc.Dropdown(id='pio-exp-profile', placeholder='Sovrascrivi un\'analisi esistente…',
                                 style={'font-size': '11px', 'flex': '1'}),
                    html.Button('🗑', id='pio-exp-del-btn', n_clicks=0,
                                title="Cancella l'analisi selezionata",
                                style={'background': '#c0392b', 'color': 'white', 'border': 'none',
                                       'padding': '6px 10px', 'border-radius': '4px',
                                       'cursor': 'pointer', 'font-size': '12px', 'margin-left': '6px'}),
                ], style={'display': 'flex', 'align-items': 'center', 'margin-bottom': '6px'}),
                dcc.Input(id='pio-exp-new', placeholder='…oppure scrivi una nuova analisi',
                          style={'width': '100%', 'font-size': '11px', 'margin-bottom': '12px',
                                 'padding': '5px 8px', 'border': '1px solid #aaa',
                                 'border-radius': '4px'}),
                html.Button('📤 Esporta', id='pio-exp-btn', n_clicks=0,
                            style={'background': '#1b7a34', 'color': 'white', 'border': 'none',
                                   'padding': '7px 16px', 'border-radius': '4px',
                                   'cursor': 'pointer', 'font-size': '12px', 'font-weight': 'bold'}),
                html.Div(id='pio-exp-status',
                         style={'font-size': '11px', 'margin-top': '8px', 'min-height': '16px'}),
            ]),

            # ── IMPORTA ────────────────────────────────────────────────────────
            html.Div(id='pio-import-view', style={'display': 'none'}, children=[
                html.Div('1. Analisi da importare:',
                         style={'font-size': '11px', 'font-weight': '600',
                                'color': '#1a3a5c', 'margin-bottom': '6px'}),
                dcc.Dropdown(id='pio-imp-profile', placeholder="Scegli un'analisi…",
                             style={'font-size': '11px', 'margin-bottom': '12px'}),
                html.Div('2. Metti nella colonna:',
                         style={'font-size': '11px', 'font-weight': '600',
                                'color': '#1a3a5c', 'margin-bottom': '6px'}),
                dcc.RadioItems(id='pio-imp-target',
                               options=[{'label': ' P1', 'value': 'P1'},
                                        {'label': ' P2', 'value': 'P2'},
                                        {'label': ' P3', 'value': 'P3'}],
                               value='P1', inline=True,
                               inputStyle={'margin-right': '4px'},
                               labelStyle={'margin-right': '16px', 'font-size': '12px',
                                           'font-weight': '600'},
                               style={'margin-bottom': '12px'}),
                html.Button('📥 Importa', id='pio-imp-btn', n_clicks=0,
                            style={'background': '#1a3a5c', 'color': 'white', 'border': 'none',
                                   'padding': '7px 16px', 'border-radius': '4px',
                                   'cursor': 'pointer', 'font-size': '12px', 'font-weight': 'bold'}),
                html.Div(id='pio-imp-status',
                         style={'font-size': '11px', 'margin-top': '8px', 'min-height': '16px'}),

                # ── Gestione dell'analisi selezionata sopra ──────────────────
                html.Hr(style={'margin': '14px 0 10px'}),
                html.Div('Gestisci l\'analisi selezionata sopra:',
                         style={'font-size': '11px', 'font-weight': '600',
                                'color': '#1a3a5c', 'margin-bottom': '6px'}),
                html.Div([
                    dcc.Input(id='pio-rename-input', placeholder='Nuovo nome…',
                              style={'flex': '1', 'font-size': '11px', 'padding': '5px 8px',
                                     'border': '1px solid #aaa', 'border-radius': '4px',
                                     'margin-right': '4px'}),
                    html.Button('✏️ Rinomina', id='pio-rename-btn', n_clicks=0,
                                style={'background': '#2e6da4', 'color': 'white', 'border': 'none',
                                       'padding': '6px 10px', 'border-radius': '4px',
                                       'cursor': 'pointer', 'font-size': '11px', 'margin-right': '4px'}),
                    html.Button('🗑 Cancella', id='pio-del-btn', n_clicks=0,
                                style={'background': '#c0392b', 'color': 'white', 'border': 'none',
                                       'padding': '6px 10px', 'border-radius': '4px',
                                       'cursor': 'pointer', 'font-size': '11px'}),
                ], style={'display': 'flex', 'align-items': 'center', 'margin-bottom': '10px'}),
                html.Button('🧹 Azzera P1/P2/P3 (svuota i portafogli)',
                            id='pio-reset-cols-btn', n_clicks=0,
                            style={'background': '#fff3e0', 'color': '#e65100',
                                   'border': '1px solid #ffb74d', 'padding': '6px 12px',
                                   'border-radius': '4px', 'cursor': 'pointer', 'font-size': '11px',
                                   'font-weight': 'bold'}),
                html.Div(id='pio-manage-status',
                         style={'font-size': '11px', 'margin-top': '8px', 'min-height': '16px',
                                'color': '#555'}),
            ]),

        ], style={'background': 'white', 'border-radius': '10px', 'padding': '20px 24px',
                  'width': '460px', 'box-shadow': '0 4px 24px rgba(0,0,0,0.18)',
                  'position': 'relative'}),
    ]),

    # ── Modal Gestione Lista ──────────────────────────────────────────────────
    html.Div(id='file-editor-overlay', style=_EDITOR_HIDDEN, children=[
        html.Div(style={
            'background': 'white', 'borderRadius': '8px', 'padding': '20px',
            'width': '700px', 'maxHeight': '82vh', 'overflowY': 'auto',
            'boxShadow': '0 4px 24px rgba(0,0,0,0.35)',
        }, children=[
            html.Div([
                html.H4(id='file-editor-title',
                        style={'margin': 0, 'fontSize': '14px', 'fontWeight': '700'}),
                html.Button('✕', id='file-editor-close', n_clicks=0,
                            style={'background': 'none', 'border': 'none', 'fontSize': '20px',
                                   'cursor': 'pointer', 'color': '#666', 'lineHeight': 1}),
            ], style={'display': 'flex', 'justifyContent': 'space-between',
                      'alignItems': 'center', 'marginBottom': '14px'}),
            dash_table.DataTable(
                id='file-editor-table',
                columns=[
                    {'name': 'Ticker',      'id': 'ticker',      'editable': True},
                    {'name': 'Descrizione', 'id': 'descrizione', 'editable': True},
                    {'name': 'Valuta',      'id': 'valuta',      'editable': True},
                ],
                data=[],
                row_deletable=True,
                style_table={'maxHeight': '280px', 'overflowY': 'auto'},
                style_header={'backgroundColor': '#f5f7fa', 'fontWeight': '700',
                              'fontSize': '12px', 'padding': '6px 8px'},
                style_cell={'fontSize': '12px', 'padding': '5px 8px', 'textAlign': 'left',
                            'border': '1px solid #e0e4ec'},
                style_data_conditional=[{'if': {'row_index': 'odd'},
                                         'backgroundColor': '#fafbfd'}],
            ),
            html.Div([
                dcc.Input(id='new-ticker-input', placeholder='Ticker (es. AAPL)',
                          debounce=False,
                          style={'width': '120px', 'fontSize': '12px', 'padding': '5px 8px',
                                 'border': '1px solid #ccc', 'borderRadius': '4px'}),
                dcc.Input(id='new-desc-input', placeholder='Descrizione',
                          debounce=False,
                          style={'width': '210px', 'fontSize': '12px', 'padding': '5px 8px',
                                 'border': '1px solid #ccc', 'borderRadius': '4px'}),
                dcc.Dropdown(
                    id='new-valuta-dropdown',
                    options=[{'label': 'EUR', 'value': 'EUR'},
                             {'label': 'USD', 'value': 'USD'},
                             {'label': 'GBP', 'value': 'GBP'}],
                    value='EUR', clearable=False,
                    style={'width': '85px', 'fontSize': '12px', 'display': 'inline-block'},
                ),
                html.Button('➕ Aggiungi', id='add-ticker-btn', n_clicks=0,
                            style={'fontSize': '12px', 'padding': '5px 14px',
                                   'background': '#e8f5e9', 'border': '1px solid #a5d6a7',
                                   'color': '#1b5e20', 'borderRadius': '4px',
                                   'cursor': 'pointer', 'fontWeight': '600'}),
            ], style={'display': 'flex', 'alignItems': 'center', 'gap': '8px',
                      'marginTop': '12px', 'flexWrap': 'wrap'}),
            html.Div([
                html.Div(id='file-editor-status',
                         style={'fontSize': '12px', 'color': '#555', 'flex': 1}),
                html.Button('💾 Salva e Aggiorna', id='save-file-editor-btn', n_clicks=0,
                            style={'fontSize': '12px', 'padding': '7px 18px',
                                   'background': '#1a3a5c', 'color': 'white',
                                   'border': 'none', 'borderRadius': '4px',
                                   'cursor': 'pointer', 'fontWeight': '600'}),
            ], style={'display': 'flex', 'justifyContent': 'space-between',
                      'alignItems': 'center', 'marginTop': '14px'}),
        ]),
    ]),

    # ── Stores ───────────────────────────────────────────────────────────────
    dcc.Interval(id='refresh-poll-interval', interval=1000, n_intervals=0, disabled=True),
    dcc.Store(id='active-xlsx-file',        data='ETF.xlsx'),
    dcc.Store(id='asset-checklist',         data=[]),
    dcc.Store(id='stock-data',              data=None),
    dcc.Store(id='original-prices-data',    data=None),
    dcc.Store(id='ticker-map-store',        data={}),
    dcc.Store(id='insufficient-data-store', data=[]),
    dcc.Store(id='weights-store-P1',        data={}),
    dcc.Store(id='weights-store-P2',        data={}),
    dcc.Store(id='weights-store-P3',        data={}),
    dcc.Store(id='global-assets-selected',  data=[]),
    dcc.Store(id='tab1-slider-store',       data=None),
    dcc.Store(id='custom-tickers-store',    data=None),
    dcc.Store(id='upload-done-ts',          data=None),
    dcc.Store(id='nav-reload',              storage_type='memory', data=None),
    dcc.Download(id='download-data'),
    dcc.Download(id='download-template'),
    dcc.Download(id='isin-download-data'),
    dcc.Interval(id='isin-poll', interval=600, n_intervals=0, disabled=True),
    dcc.Store(id='isin-req-id', data=None),
    dcc.Store(id='style-analysis-store', data=None),
    dcc.Store(id='pending-upload-store', data=None),
    # Modal stilizzato per la conferma di sovrascrittura all'upload di un file
    html.Div(id='overwrite-modal-overlay', style=_EDITOR_HIDDEN, children=[
        html.Div(style={'background': 'white', 'borderRadius': '8px', 'padding': '24px 26px',
                        'width': '420px', 'maxWidth': '90vw',
                        'boxShadow': '0 4px 24px rgba(0,0,0,0.35)'}, children=[
            html.Div([
                html.H4('Sovrascrivere il lavoro in corso?',
                        style={'margin': 0, 'fontSize': '14px', 'fontWeight': '700', 'color': '#1a3a5c'}),
                html.Button('✕', id='overwrite-x', n_clicks=0,
                            style={'background': 'none', 'border': 'none', 'fontSize': '20px',
                                   'cursor': 'pointer', 'color': '#666', 'lineHeight': 1}),
            ], style={'display': 'flex', 'justifyContent': 'space-between',
                      'alignItems': 'center', 'marginBottom': '14px'}),
            html.Div("Hai una sessione non salvata: caricare il nuovo file sostituirà "
                     "l'analisi in corso. Continuare?",
                     style={'fontSize': '12px', 'color': '#444', 'marginBottom': '18px',
                            'lineHeight': '1.4'}),
            html.Div([
                html.Button('Annulla', id='overwrite-no', n_clicks=0,
                            style={'background': '#e8e8e8', 'color': '#333', 'border': 'none',
                                   'borderRadius': '4px', 'padding': '8px 16px', 'cursor': 'pointer',
                                   'fontSize': '12px', 'fontWeight': 'bold', 'marginRight': '8px'}),
                html.Button('Sì, carica', id='overwrite-yes', n_clicks=0,
                            style={'background': '#1a3a5c', 'color': 'white', 'border': 'none',
                                   'borderRadius': '4px', 'padding': '8px 16px', 'cursor': 'pointer',
                                   'fontSize': '12px', 'fontWeight': 'bold'}),
            ], style={'display': 'flex', 'justifyContent': 'flex-end'}),
        ]),
    ]),
    dcc.Store(id='pending-fileload-store', data=None),
    dcc.Store(id='fp-delete-trigger', data=None),
    dcc.Store(id='sm-dirty-sink', data=None),
    # Modal stilizzato (coerente col sito) per la conferma di sovrascrittura
    html.Div(id='fileload-modal-overlay', style=_EDITOR_HIDDEN, children=[
        html.Div(style={'background': 'white', 'borderRadius': '8px', 'padding': '24px 26px',
                        'width': '420px', 'maxWidth': '90vw',
                        'boxShadow': '0 4px 24px rgba(0,0,0,0.35)'}, children=[
            html.Div([
                html.H4('Sovrascrivere il lavoro in corso?',
                        style={'margin': 0, 'fontSize': '14px', 'fontWeight': '700', 'color': '#1a3a5c'}),
                html.Button('✕', id='fileload-x', n_clicks=0,
                            style={'background': 'none', 'border': 'none', 'fontSize': '20px',
                                   'cursor': 'pointer', 'color': '#666', 'lineHeight': 1}),
            ], style={'display': 'flex', 'justifyContent': 'space-between',
                      'alignItems': 'center', 'marginBottom': '14px'}),
            html.Div("Hai una sessione non salvata: caricare questo file sostituirà "
                     "l'analisi in corso. Continuare?",
                     style={'fontSize': '12px', 'color': '#444', 'marginBottom': '18px',
                            'lineHeight': '1.4'}),
            html.Div([
                html.Button('Annulla', id='fileload-no', n_clicks=0,
                            style={'background': '#e8e8e8', 'color': '#333', 'border': 'none',
                                   'borderRadius': '4px', 'padding': '8px 16px', 'cursor': 'pointer',
                                   'fontSize': '12px', 'fontWeight': 'bold', 'marginRight': '8px'}),
                html.Button('Sì, carica', id='fileload-yes', n_clicks=0,
                            style={'background': '#1a3a5c', 'color': 'white', 'border': 'none',
                                   'borderRadius': '4px', 'padding': '8px 16px', 'cursor': 'pointer',
                                   'fontSize': '12px', 'fontWeight': 'bold'}),
            ], style={'display': 'flex', 'justifyContent': 'flex-end'}),
        ]),
    ]),

    # ── Modale ISIN conversion ────────────────────────────────────────────────
    html.Div(id='isin-modal-overlay',
             style={'display': 'none', 'position': 'fixed', 'top': '0', 'left': '0',
                    'width': '100%', 'height': '100%', 'z-index': '9000',
                    'background': 'rgba(0,0,0,0.45)', 'align-items': 'center',
                    'justify-content': 'center'},
             children=[
        html.Div([
            html.Div([
                html.Span('🔀 Converti ISIN in Ticker',
                          style={'font-weight': '700', 'font-size': '14px', 'color': '#1a3a5c'}),
                html.Button('✕', id='isin-close-btn', n_clicks=0,
                            style={'background': 'none', 'border': 'none', 'font-size': '18px',
                                   'cursor': 'pointer', 'color': '#666', 'float': 'right'}),
            ], style={'display': 'flex', 'justify-content': 'space-between',
                      'align-items': 'center', 'margin-bottom': '12px'}),

            html.P('Carica un file Excel con una colonna di ISIN (o ticker). '
                   'Il sistema trova il ticker Yahoo con più dati storici, '
                   'costruisce il file portafoglio e scarica i prezzi.',
                   style={'font-size': '11px', 'color': '#555', 'margin-bottom': '12px'}),

            dcc.Upload(id='isin-upload',
                children=html.Div([
                    html.I(className='fa-solid fa-file-excel',
                           style={'font-size': '28px', 'color': '#e65100', 'margin-bottom': '8px'}),
                    html.Div('Trascina il file Excel qui', style={'font-weight': '600', 'font-size': '12px'}),
                    html.Div('oppure clicca per selezionare', style={'font-size': '10px', 'color': '#888'}),
                ], style={'textAlign': 'center', 'padding': '20px'}),
                style={'border': '2px dashed #ffb74d', 'border-radius': '8px',
                       'background': '#fffde7', 'cursor': 'pointer', 'margin-bottom': '12px'},
                multiple=False,
            ),

            html.Div(id='isin-progress-text',
                     style={'font-size': '11px', 'color': '#555', 'min-height': '20px',
                            'margin-bottom': '8px'}),

            html.Div([
                html.Button('📥 Scarica file portafoglio', id='isin-dl-btn', n_clicks=0,
                            style={'display': 'none', 'font-size': '11px', 'padding': '6px 14px',
                                   'background': '#1b7a34', 'color': 'white', 'border': 'none',
                                   'border-radius': '4px', 'cursor': 'pointer',
                                   'font-weight': 'bold', 'margin-right': '8px'}),
                html.Span(id='isin-load-status',
                          style={'font-size': '11px', 'color': '#1a7a4a', 'font-weight': '600'}),
            ]),

        ], style={
            'background': 'white', 'border-radius': '10px', 'padding': '20px 24px',
            'width': '520px', 'box-shadow': '0 4px 24px rgba(0,0,0,0.18)',
            'position': 'relative',
        }),
    ]),

    # ── Modale progresso aggiornamento ───────────────────────────────────────
    html.Div([
        html.Div([
            # Intestazione
            html.Div([
                html.Div([
                    html.I(className='fas fa-chart-line',
                           style={'marginRight': '8px', 'color': '#007755'}),
                    html.Span('Aggiornamento Dati di Mercato', style={
                        'fontFamily': "'Playfair Display', serif",
                        'fontSize': '1.05rem', 'fontWeight': '700',
                        'color': '#1a3a5c',
                    }),
                ], style={'display': 'flex', 'alignItems': 'center'}),
                html.Button('✕', id='progress-modal-close', n_clicks=0, style={
                    'background': 'none', 'border': 'none', 'cursor': 'pointer',
                    'fontSize': '18px', 'color': '#6b7a99', 'padding': '0 4px',
                    'lineHeight': '1', 'marginLeft': '16px',
                }),
            ], style={
                'display': 'flex', 'justifyContent': 'space-between',
                'alignItems': 'center', 'marginBottom': '22px',
                'paddingBottom': '14px', 'borderBottom': '1px solid #e2e8f0',
            }),
            # Barra progresso
            html.Div(
                html.Div(id='modal-progress-fill', style=_FILL_LOADING),
                style={
                    'width': '100%', 'height': '12px', 'background': '#e2e8f0',
                    'borderRadius': '8px', 'overflow': 'hidden', 'marginBottom': '12px',
                },
            ),
            # Testo percentuale
            html.Div(id='modal-pct-text', style={
                'fontSize': '0.9rem', 'color': '#1a3a5c', 'fontWeight': '700',
                'fontFamily': 'Inter, sans-serif', 'textAlign': 'center',
                'marginBottom': '6px',
            }),
            # Messaggio stato
            html.Div(id='modal-status-text', style=_STATUS_GREY),
        ], style={
            'background': '#ffffff', 'borderRadius': '14px',
            'padding': '30px 36px', 'width': '440px', 'maxWidth': '90vw',
            'boxShadow': '0 24px 64px rgba(26,58,92,0.22)',
        }),
    ], id='progress-modal-overlay', style=_MODAL_HIDDEN),

    # ── Overlay scarica dati ──────────────────────────────────────────────────
    html.Div([
        html.Div([
            html.Div(style={
                'width': '64px', 'height': '64px',
                'border': '6px solid #e2e8f0',
                'borderTop': '6px solid #1a3a5c',
                'borderRadius': '50%',
                'animation': 'spin 0.9s linear infinite',
                'marginBottom': '20px',
            }),
            html.Div('Scarica dati', style={
                'fontFamily': "'Inter', sans-serif",
                'fontSize': '1.05rem', 'fontWeight': '600',
                'color': '#1a3a5c', 'letterSpacing': '0.02em',
            }),
        ], style={
            'display': 'flex', 'flexDirection': 'column',
            'alignItems': 'center', 'justifyContent': 'center',
            'background': '#ffffff', 'borderRadius': '16px',
            'padding': '40px 56px',
            'boxShadow': '0 24px 64px rgba(26,58,92,0.22)',
        }),
    ], id='download-overlay', style=_MODAL_HIDDEN),

    # ── Tab navigation ────────────────────────────────────────────────────────
    dcc.Tabs(id='main-tabs', value='tab-portfolio',
             style={'margin': '0 0 0 0'},
             colors={'border': '#dee2e6', 'primary': '#1a3a5c', 'background': '#f0f4fa'},
             children=[
        dcc.Tab(label='📊 Analisi Portafoglio', value='tab-portfolio',
                style={'font-size': '12px', 'padding': '8px 18px'},
                selected_style={'font-size': '12px', 'padding': '8px 18px',
                                'font-weight': 'bold', 'border-top': '3px solid #1a3a5c'}),
        dcc.Tab(label='📈 Frontiera Efficiente', value='tab-frontiera',
                style={'font-size': '12px', 'padding': '8px 18px'},
                selected_style={'font-size': '12px', 'padding': '8px 18px',
                                'font-weight': 'bold', 'border-top': '3px solid #1a3a5c'}),
        dcc.Tab(label='📅 Rendimenti Storici', value='tab-rendimenti',
                style={'font-size': '12px', 'padding': '8px 18px'},
                selected_style={'font-size': '12px', 'padding': '8px 18px',
                                'font-weight': 'bold', 'border-top': '3px solid #1a3a5c'}),
        dcc.Tab(label='📐 Style Analysis', value='tab-sa',
                style={'font-size': '12px', 'padding': '8px 18px'},
                selected_style={'font-size': '12px', 'padding': '8px 18px',
                                'font-weight': 'bold', 'border-top': '3px solid #1a3a5c'}),
        dcc.Tab(label='🔗 Matrice Correlazioni', value='tab-correlazioni',
                style={'font-size': '12px', 'padding': '8px 18px'},
                selected_style={'font-size': '12px', 'padding': '8px 18px',
                                'font-weight': 'bold', 'border-top': '3px solid #1a3a5c'}),
    ]),
    html.Div(id='tab1-content'),
    # SA layout sempre presente nel DOM (nascosto finché non si clicca il tab)
    # Garantisce che i callback sa_populate_x/y trovino i componenti al caricamento
    html.Div(id='tab-sa-content',
             children=get_style_analysis_tab([]),
             style={'display': 'none'}),
    # Matrice Correlazioni: layout sempre presente nel DOM (nascosto), così i
    # callback trovano sempre i componenti (stesso pattern di Style Analysis).
    html.Div(id='tab-corr-content',
             children=get_correlation_matrix_tab([]),
             style={'display': 'none'}),
    dcc.Store(id='corr-calculated', data=False),
    # Frontiera e Rendimenti incorporate via iframe (app standalone, navbar nascosta).
    # src impostato dinamicamente al click del tab con timestamp anti-cache,
    # così l'iframe si ricarica fresco e i callback colpiscono il server corrente.
    html.Div(id='tab-frontiera-content', style={'display': 'none'}, children=[
        html.Iframe(id='iframe-frontiera', src='',
                    style={'width': '100%', 'height': 'calc(100vh - 150px)',
                           'border': 'none'}),
    ]),
    html.Div(id='tab-rendimenti-content', style={'display': 'none'}, children=[
        html.Iframe(id='iframe-rendimenti', src='',
                    style={'width': '100%', 'height': 'calc(100vh - 150px)',
                           'border': 'none'}),
    ]),

], style={'marginTop': '106px', 'padding': '0 1%'}),
])


def _reconstruct_from_json(ns):
    first = next(iter(ns.values()))
    dates = pd.to_datetime(first['dates'])
    pr, ret, tm, vm = {}, {}, {}, {}
    for desc, v in ns.items():
        p = v.get('prices') or []
        r = v.get('returns') or []
        if p:  pr[desc]  = [float(x) if x is not None else float('nan') for x in p]
        if r:  ret[desc] = [float(x) if x is not None else float('nan') for x in r]
        tm[desc] = v.get('ticker', desc)
        vm[desc] = v.get('currency', 'EUR')
    op = pd.DataFrame(pr,  index=dates) if pr  else None
    cr = pd.DataFrame(ret, index=dates) if ret else None
    return cr, op, tm, vm


app.clientside_callback(
    """
    function(_) {
        var t = 0;
        try {
            if (window.performance && window.performance.getEntriesByType) {
                var e = window.performance.getEntriesByType('navigation');
                if (e.length > 0) { t = e[0].type === 'reload' ? 1 : 0; }
            } else if (window.performance && window.performance.navigation) {
                t = window.performance.navigation.type;
            }
        } catch(ex) {}
        return t;
    }
    """,
    Output('nav-reload', 'data'),
    Input('nav-reload',  'data'),
    prevent_initial_call=False,
)


# ─────────────────────────────────────────────────────────────────────────────
# Callback: intercetta upload → warning se sessione non salvata
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('pending-upload-store', 'data'),
    Output('overwrite-modal-overlay', 'style'),
    Output('upload-data',          'contents', allow_duplicate=True),
    Input('upload-data',           'contents'),
    State('upload-data',           'filename'),
    prevent_initial_call=True,
)
def _stage_upload(contents, filename):
    if not contents:
        raise PreventUpdate
    _u = _get_username()
    needs = _sm.has_unsaved_changes(_u)
    payload = {'contents': contents, 'filename': filename or '', 'await': needs}
    if needs:
        # Mostra warning — resetta contents per permettere re-upload stesso file
        return payload, _EDITOR_SHOWN, None
    # Nessuna sessione a rischio — procedi direttamente (contents rimane)
    return payload, _EDITOR_HIDDEN, no_update


@app.callback(
    Output('pending-upload-store', 'data', allow_duplicate=True),
    Output('overwrite-modal-overlay', 'style', allow_duplicate=True),
    Input('overwrite-no',          'n_clicks'),
    Input('overwrite-x',           'n_clicks'),
    prevent_initial_call=True,
)
def _cancel_upload(n_no, n_x):
    if not (n_no or n_x):
        raise PreventUpdate
    return None, _EDITOR_HIDDEN   # annulla il pending + chiudi modal


# Chiudi il modal di sovrascrittura upload al click di "Sì, carica"
@app.callback(
    Output('overwrite-modal-overlay', 'style', allow_duplicate=True),
    Input('overwrite-yes', 'n_clicks'),
    prevent_initial_call=True,
)
def _close_overwrite_modal(n):
    if not n:
        raise PreventUpdate
    return _EDITOR_HIDDEN


# ─────────────────────────────────────────────────────────────────────────────
# Callback: inizializzazione + upload
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('upload-status',           'children'),
    Output('asset-checklist',         'data'),
    Output('stock-data',              'data'),
    Output('original-prices-data',    'data'),
    Output('insufficient-data-store', 'data'),
    Output('ticker-map-store',        'data'),
    Output('data-last-updated',       'children'),
    Output('custom-tickers-store',    'data'),
    Output('refresh-poll-interval',   'disabled',    allow_duplicate=True),
    Output('refresh-poll-interval',   'n_intervals', allow_duplicate=True),
    Output('refresh-data-btn',        'disabled',    allow_duplicate=True),
    Output('progress-modal-overlay',  'style',       allow_duplicate=True),
    Output('modal-progress-fill',     'style',       allow_duplicate=True),
    Output('modal-pct-text',          'children',    allow_duplicate=True),
    Output('modal-status-text',       'children',    allow_duplicate=True),
    Output('modal-status-text',       'style',       allow_duplicate=True),
    Output('upload-done-ts',          'data',        allow_duplicate=True),
    Input('nav-reload',              'data'),
    Input('pending-upload-store',    'data'),     # sostituisce upload-data diretta
    Input('overwrite-yes',           'n_clicks'),
    State('upload-data',             'filename'),
    prevent_initial_call=True,
)
def update_output(nav_reload, pending_upload, _confirm_n, filename):
    import time as _time
    ctx = callback_context
    triggered_id = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''

    # Risolvi contents e filename da pending-upload-store
    contents = None
    if triggered_id in ('pending-upload-store', 'overwrite-yes') and pending_upload:
        # Serve conferma e non è ancora arrivato il "Sì" → attendi (non caricare)
        if triggered_id == 'pending-upload-store' and pending_upload.get('await'):
            raise PreventUpdate
        contents = pending_upload.get('contents')
        filename = pending_upload.get('filename', filename)
        # Se triggered dal "Sì" ma pending è None → l'utente ha annullato
        if triggered_id == 'overwrite-yes' and not contents:
            raise PreventUpdate
        triggered_id = 'upload-data'   # trattalo come upload normale

    _noup = (no_update,) * 8

    if triggered_id == 'nav-reload':
        _cl_clear(_get_username())
        _PENDING.clear()
        _u = _get_username()

        # ── FILE UNICO (Opzione A) ─────────────────────────────────────────────
        # All'avvio carica SEMPRE current.json (l'unico file di lavoro): NON
        # resetta mai. Così Portafoglio, Correlazioni, Frontiera e Rendimenti
        # leggono tutti lo stesso file → niente disallineamenti. Il default ETF
        # si carica solo come SEED iniziale (file vuoto) o scegliendolo dal menu.
        cr, op, tm = None, None, {}
        try:
            cr, op, tm = dc.build_dataset(_u)
        except Exception:
            cr = None
        if cr is not None and not cr.empty:
            _is_pers = (_read_tipo(_u) == 'personale')
            _active_file_store['filename'] = '__personale__' if _is_pers else 'ETF.xlsx'
            _active_file_store['is_personale'] = _is_pers
            with _DL_LOCK:
                _DL_BUFFER.clear()
                _DL_BUFFER.update({'close_returns': cr, 'original_prices': op, 'ticker_map': tm})
            options = [{'label': c, 'value': c} for c in cr.columns]
            _lbl = 'File personale' if _is_pers else 'File di lavoro'
            return (
                html.Div(f'✓ {len(options)} asset — {_lbl}',
                         style={'color': '#007755', 'font-size': '11px'}),
                options,
                cr.to_json(date_format='iso', orient='split'),
                op.to_json(date_format='iso', orient='split'),
                [], tm, _lbl, None,
                *_noup, no_update,
            )

        # ── SEED INIZIALE: current.json vuoto/assente → default ETF fresco ─────
        _active_file_store['filename'] = 'ETF.xlsx'
        _active_file_store['is_personale'] = False
        saved_at = ''
        _data = _sm.load_default('ETF')
        if _data is not None:
            cr       = _data.get('close_returns')
            op       = _data.get('original_prices')
            tm       = _data.get('ticker_map', {})
            saved_at = _data.get('saved_at', '')
        if cr is None:
            with _DL_LOCK:
                cr       = _DL_BUFFER.get('close_returns')
                op       = _DL_BUFFER.get('original_prices')
                tm       = _DL_BUFFER.get('ticker_map', {})
                saved_at = _DL_BUFFER.get('saved_at', '')
        if cr is not None:
            with _DL_LOCK:
                _DL_BUFFER.update({'close_returns': cr, 'original_prices': op,
                                   'ticker_map': tm, 'saved_at': saved_at})
            _write_user_json(cr, op, tm, reset_state=True, tipo='default:ETF.xlsx')
            options  = [{'label': col, 'value': col} for col in cr.columns]
            return (
                html.Div(f'✓ {len(options)} asset — ETF (default)',
                         style={'color': '#007755', 'font-size': '11px'}),
                options,
                cr.to_json(date_format='iso', orient='split'),
                op.to_json(date_format='iso', orient='split'),
                [], tm, f"Aggiornati: {saved_at}" if saved_at else '', None,
                *_noup, no_update,
            )
        options, ticker_map = load_ticker_names_only()
        return (
            html.Div('⏳ Download dati in corso…',
                     style={'color': '#e67e22', 'font-size': '11px'}),
            options, None, None, [], ticker_map, '', None,
            *_noup, no_update,
        )

    elif triggered_id == 'upload-data' and contents is not None:
        try:
            _, content_string = contents.split(',')
            decoded  = base64.b64decode(content_string)
            done_ts  = _time.time()

            df        = _read_asset_excel(decoded)
            col_names = df.columns.tolist()

            is_price_file = False
            try:
                pd.to_datetime(df[col_names[0]], errors='raise')
                num_cols = df.drop(columns=[col_names[0]]).select_dtypes(include='number').columns.tolist()
                is_price_file = len(num_cols) >= 1
            except Exception:
                pass

            if is_price_file:
                df_prices = df.set_index(col_names[0])
                df_prices.index = pd.to_datetime(df_prices.index)
                df_prices = df_prices.select_dtypes(include='number').ffill().dropna(how='all')
                df_prices = _clean_prices(df_prices)
                close_returns = df_prices.pct_change(fill_method=None)
                options    = [{'label': c, 'value': c} for c in df_prices.columns]
                ticker_map = {c: c for c in df_prices.columns}
                saved_at   = datetime.now().strftime('%d/%m/%Y %H:%M')
                with _DL_LOCK:
                    _DL_BUFFER.update({
                        'date': datetime.now().strftime('%Y-%m-%d'),
                        'saved_at': saved_at,
                        'close_returns': close_returns,
                        'original_prices': df_prices,
                        'ticker_map': ticker_map,
                    })
                _write_user_json(close_returns, df_prices, ticker_map, reset_state=True)
                # Caricare un file da zero → file di lavoro Personale (persistito).
                _mark_personale()
                return (
                    html.Div(f'✓ {len(options)} asset — prezzi caricati dal file',
                             style={'color': '#007755', 'font-size': '11px'}),
                    options,
                    close_returns.to_json(date_format='iso', orient='split'),
                    df_prices.to_json(date_format='iso', orient='split'),
                    [], ticker_map, f"Caricati: {saved_at}", None,
                    *_noup, done_ts,
                )

            else:
                try:
                    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    _arch = _ROOT_DIR / 'archive' / _get_username()
                    _arch.mkdir(parents=True, exist_ok=True)
                    stem = Path(filename or 'upload').stem
                    copy_path = _arch / f'{stem}_{ts}.xlsx'
                    with open(copy_path, 'wb') as fout:
                        fout.write(decoded)
                    print(f"✓ File cliente archiviato: {copy_path}")
                except Exception as e:
                    print(f"⚠ Archiviazione file cliente fallita: {e}")

                tickers     = list(df[col_names[0]])
                descrizione = (list(df[col_names[1]]) if len(col_names) > 1
                               else [str(t) for t in tickers])
                valuta      = (list(df[col_names[2]]) if len(col_names) > 2
                               else ['EUR'] * len(tickers))
                ticker_map  = {descrizione[i]: tickers[i] for i in range(len(tickers))}
                options     = [{'label': d, 'value': d} for d in descrizione]
                custom      = {'tickers': tickers, 'descr': descrizione, 'valuta': valuta}

                # Leggi colonna pesi se presente (col 3 o colonna con header peso/weight/%)
                pesi_map = {}
                _, peso_col, peso_vals = _detect_ticker_and_weight_cols(df)
                if peso_col is not None and peso_vals is not None:
                    for i, desc in enumerate(descrizione):
                        try:
                            row_mask = df[col_names[0]].astype(str).str.strip() == str(tickers[i]).strip()
                            idx_match = df[row_mask].index
                            if not idx_match.empty and idx_match[0] in peso_vals.index:
                                v = float(peso_vals.loc[idx_match[0]])
                                if v > 0:
                                    pesi_map[desc] = v
                        except Exception:
                            pass

                start_date = (pd.Timestamp.today() - pd.DateOffset(years=10)).strftime('%Y-%m-%d')
                _username  = _get_username()
                print(f"▶ Upload file ticker [{_username}]: {len(tickers)} asset da scaricare da {start_date}"
                      + (f', pesi P1 trovati: {len(pesi_map)}' if pesi_map else ''))
                threading.Thread(
                    target=_do_download_client,
                    args=(tickers, descrizione, valuta, start_date),
                    kwargs={'username': _username, 'pesi_p1': pesi_map},
                    daemon=True,
                ).start()

                return (
                    html.Div(f'⏳ Download avviato — {len(options)} asset da Yahoo Finance…',
                             style={'color': '#e67e22', 'font-size': '11px'}),
                    options, None, None, [], ticker_map, '', custom,
                    False, 0, True, _MODAL_SHOWN, _FILL_LOADING,
                    f'Avvio — {len(tickers)} asset…', '', _STATUS_GREY,
                    done_ts,
                )

        except Exception as e:
            return (
                html.Div(f'Errore: {e}'), [], None, None, [], {}, '', None,
                *_noup, no_update,
            )

    raise PreventUpdate


# Clientside: appena upload-done-ts cambia (ogni upload completato),
# azzera upload-data.contents nel browser → il prossimo upload dello
# stesso file viene rilevato come nuovo (None → base64).
app.clientside_callback(
    """
    function(ts) {
        if (!ts) return window.dash_clientside.no_update;
        return null;
    }
    """,
    Output('upload-data', 'contents', allow_duplicate=True),
    Input('upload-done-ts', 'data'),
    prevent_initial_call=True,
)


# ─────────────────────────────────────────────────────────────────────────────
# Callback: renderizza contenuto Tab1 + mostra/nasconde tab SA
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('tab1-content',           'children'),
    Output('tab1-content',           'style'),
    Output('tab-sa-content',         'style'),
    Output('tab-frontiera-content',  'style'),
    Output('tab-rendimenti-content', 'style'),
    Output('tab-corr-content',       'style'),
    Input('main-tabs',       'value'),
    Input('asset-checklist', 'data'),
)
def render_tab1(active_tab, options_tickers):
    show = {'display': 'block'}
    hide = {'display': 'none'}
    if active_tab == 'tab-sa':
        return no_update, hide, show, hide, hide, hide
    if active_tab == 'tab-frontiera':
        return no_update, hide, hide, show, hide, hide
    if active_tab == 'tab-rendimenti':
        return no_update, hide, hide, hide, show, hide
    if active_tab == 'tab-correlazioni':
        return no_update, hide, hide, hide, hide, show
    return get_portfolio_analysis_tab(options_tickers), show, hide, hide, hide, hide


# Carica/ricarica l'iframe (fresco, anti-cache) quando si apre il relativo tab
@app.callback(
    Output('iframe-frontiera',  'src'),
    Output('iframe-rendimenti', 'src'),
    Input('main-tabs', 'value'),
    prevent_initial_call=True,
)
def _load_tab_iframe(active_tab):
    ts = int(time.time() * 1000)
    if active_tab == 'tab-frontiera':
        return f'/frontiera/?embed=1&t={ts}', no_update
    if active_tab == 'tab-rendimenti':
        return no_update, f'/rendimenti/?embed=1&t={ts}'
    raise PreventUpdate


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Matrice di Correlazione (heatmap rendimenti settimanali)
# Portato da ir_fe_14.py, adattato ai date-picker e store del portafoglio.
# ─────────────────────────────────────────────────────────────────────────────
_CORR_COLORSCALE = [
    [0.0, '#8B0000'], [0.25, '#d62728'], [0.375, '#e0e0e0'], [0.5, 'white'],
    [0.625, '#f0f0f0'], [0.85, '#ccffcc'], [0.9, '#2ca02c'], [1.0, '#006400'],
]


def _corr_heatmap_figure(corr_matrix, title_text):
    n = len(corr_matrix.columns)
    max_lbl = max((len(str(c)) for c in corr_matrix.columns), default=6)
    tick_size = max(7, min(12, int(130 / max(n, 1))))
    l_margin = max(80, max_lbl * tick_size * 0.65)
    b_margin = max(80, max_lbl * tick_size * 0.65)
    fig = go.Figure(data=go.Heatmap(
        z=corr_matrix.values, x=list(corr_matrix.columns), y=list(corr_matrix.index),
        colorscale=_CORR_COLORSCALE, zmid=0, zmin=-1, zmax=1,
        text=np.round(corr_matrix.values, 2), texttemplate='%{text}',
        textfont={'size': max(8, min(16, int(180 / max(n, 1))))},
        colorbar=dict(title='Correlazione', tickvals=[-1, -0.5, 0, 0.7, 1],
                      ticktext=['-1.0', '-0.5', '0.0', '0.7', '1.0']),
        hovertemplate='%{y} vs %{x}<br>Correlazione: %{z:.3f}<extra></extra>',
    ))
    fig.update_layout(
        title=dict(text=title_text, font=dict(size=13), x=0.5),
        xaxis=dict(side='bottom', tickangle=-45, tickfont=dict(size=tick_size), automargin=True),
        yaxis=dict(autorange='reversed', tickfont=dict(size=tick_size), automargin=True),
        margin=dict(l=int(l_margin), b=int(b_margin), t=40, r=20), autosize=True,
        paper_bgcolor='white',
    )
    return fig


# Costruisce la griglia asset (colonne MT/ML) e popola il benchmark, quando
# cambiano i dati caricati. MT pre-spuntati = SELEZIONE CONDIVISA (campo `checked`
# di current.json) → coerente con tutte le altre tab. ML inizialmente vuoti.
@app.callback(
    Output('corr-asset-grid',         'children'),
    Output('benchmark-selector-corr', 'options'),
    Input('stock-data',               'data'),
    prevent_initial_call=True,
)
def build_corr_grid(stock_data):
    df = _get_df(stock_data)
    if df is None or df.empty:
        return ([html.Div('(carica i dati per popolare)',
                          style={'font-size': '10px', 'color': '#aaa', 'font-style': 'italic',
                                 'text-align': 'center', 'padding': '10px 4px'})], [])
    assets = list(df.columns)
    ns = _read_user_json(_get_username())
    mt_default = {a for a in assets if ns.get(a, {}).get('checked')}

    def _chk(kind, asset, checked, color):
        return html.Div(
            dcc.Checklist(id={'type': kind, 'index': asset},
                          options=[{'label': '', 'value': asset}],
                          value=[asset] if checked else [],
                          inputStyle={'cursor': 'pointer', 'accentColor': color, 'margin': '0'},
                          style={'display': 'flex', 'justifyContent': 'center'}),
            style={'width': '34px', 'display': 'flex', 'justifyContent': 'center', 'alignItems': 'center'})

    rows = []
    for i, a in enumerate(assets):
        rows.append(html.Div([
            html.Div(html.Span(a, **{'data-tooltip': a},
                               style={'overflow': 'hidden', 'whiteSpace': 'nowrap',
                                      'textOverflow': 'ellipsis', 'maxWidth': '100%',
                                      'fontSize': '10px', 'color': '#1a3a5c'}),
                     style={'flex': '1', 'overflow': 'hidden', 'paddingRight': '4px', 'minWidth': '0'}),
            _chk('corr-mt', a, a in mt_default, '#1a3a5c'),
            _chk('corr-ml', a, False, '#e6550d'),
        ], style={'display': 'flex', 'align-items': 'center', 'gap': '4px',
                  'padding': '2px', 'borderBottom': '1px dotted #eee',
                  'background': 'white' if i % 2 == 0 else '#fafcff'}))
    return rows, [{'label': a, 'value': a} for a in assets]


# Pulsanti seleziona/deseleziona tutti (toggle) per colonna MT e ML
@app.callback(
    Output({'type': 'corr-mt', 'index': ALL}, 'value'),
    Input('corr-mt-all', 'n_clicks'),
    State({'type': 'corr-mt', 'index': ALL}, 'value'),
    State({'type': 'corr-mt', 'index': ALL}, 'id'),
    prevent_initial_call=True,
)
def corr_mt_all(n, vals, ids):
    if not n:
        raise PreventUpdate
    any_unchecked = any(not v for v in (vals or []))
    return [[i['index']] if any_unchecked else [] for i in ids]


@app.callback(
    Output({'type': 'corr-ml', 'index': ALL}, 'value'),
    Input('corr-ml-all', 'n_clicks'),
    State({'type': 'corr-ml', 'index': ALL}, 'value'),
    State({'type': 'corr-ml', 'index': ALL}, 'id'),
    prevent_initial_call=True,
)
def corr_ml_all(n, vals, ids):
    if not n:
        raise PreventUpdate
    any_unchecked = any(not v for v in (vals or []))
    return [[i['index']] if any_unchecked else [] for i in ids]


# La spunta MT è la SELEZIONE CONDIVISA: scrivila nel campo `checked` di
# current.json (fonte unica) + aggiorna lo store globale, così è coerente con
# tutte le altre tab e persiste nella dashboard temporanea / file personale.
@app.callback(
    Output('global-assets-selected', 'data', allow_duplicate=True),
    Input({'type': 'corr-mt', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def corr_mt_to_checked(mt_vals):
    sel = [v[0] for v in (mt_vals or []) if v]
    _update_user_json(checked=sel)
    return sel


# All'apertura del tab Correlazioni allinea le spunte MT alla selezione condivisa
# (campo `checked`), SENZA ricostruire la griglia → le spunte ML restano intatte.
@app.callback(
    Output({'type': 'corr-mt', 'index': ALL}, 'value', allow_duplicate=True),
    Input('main-tabs', 'value'),
    State({'type': 'corr-mt', 'index': ALL}, 'id'),
    prevent_initial_call=True,
)
def sync_mt_on_open(active_tab, ids):
    if active_tab != 'tab-correlazioni' or not ids:
        raise PreventUpdate
    ns = _read_user_json(_get_username())
    return [[i['index']] if ns.get(i['index'], {}).get('checked') else [] for i in ids]


@app.callback(
    Output('correlation-heatmap',     'figure'),
    Output('date-values-correlation',  'children'),
    Output('corr-filter-info',         'children'),
    Output('corr-calculated',          'data'),
    Input({'type': 'corr-mt', 'index': ALL}, 'value'),
    Input('benchmark-selector-corr',   'value'),
    Input('update-correlation-button', 'n_clicks'),
    State('stock-data',                'data'),
    State('dr-start-corr',             'date'),
    State('dr-end-corr',               'date'),
    State('correlation-window-input',  'value'),
    prevent_initial_call=True,
)
def update_correlation_matrix(mt_vals, benchmark_sel, n_clicks, stock_data,
                              date_start, date_end, corr_window):
    def _empty(msg):
        f = go.Figure().add_annotation(text=msg, xref='paper', yref='paper',
                                       x=0.5, y=0.5, showarrow=False, font=dict(size=16, color='#888'))
        f.update_layout(paper_bgcolor='white', margin=dict(l=20, r=20, t=20, b=20))
        return f

    close_returns = _get_df(stock_data)
    if close_returns is None or close_returns.empty:
        return _empty('Nessun dato disponibile'), 'Nessun dato', '', no_update

    # Asset spuntati nella colonna MT (la spunta determina il calcolo)
    mt_sel = [v[0] for v in (mt_vals or []) if v]
    cols   = [c for c in mt_sel if c in close_returns.columns]
    if len(cols) < 2:
        return _empty('Spunta almeno 2 asset nella colonna MT'), '', 'Seleziona ≥2 asset (MT)', no_update

    # Filtro date (DatePickerSingle → stringhe 'YYYY-MM-DD')
    try:
        filtered_df = close_returns.loc[pd.to_datetime(date_start):pd.to_datetime(date_end)] \
                      if (date_start and date_end) else close_returns
    except Exception:
        filtered_df = close_returns
    if filtered_df.empty:
        filtered_df = close_returns
    start_date = filtered_df.index[0]  if not filtered_df.empty else pd.Timestamp.now()
    end_date   = filtered_df.index[-1] if not filtered_df.empty else pd.Timestamp.now()

    filtered_df = filtered_df[cols]
    filter_info = f'{len(cols)} asset selezionati (MT)'

    if corr_window and corr_window > 0:
        filtered_df = filtered_df.tail(corr_window)

    if filtered_df.empty or filtered_df.shape[1] < 2:
        return _empty('Dati insufficienti nel range selezionato'), 'Dati insufficienti', filter_info, no_update

    # Rendimenti settimanali: neutralizza lo sfasamento degli orari di chiusura
    weekly_df   = filtered_df.resample('W').apply(lambda x: (1 + x).prod() - 1).dropna()
    corr_matrix = weekly_df.corr()
    n_weeks     = len(weekly_df)

    title_text = 'Matrice di Correlazione – rendimenti settimanali'
    if corr_window and corr_window > 0:
        title_text += f' (ultimi {min(corr_window, len(filtered_df))} gg → {n_weeks} settimane)'
    else:
        title_text += f' ({n_weeks} settimane)'

    if benchmark_sel and benchmark_sel in corr_matrix.columns:
        order = corr_matrix[benchmark_sel].abs().sort_values(ascending=False).index.tolist()
        corr_matrix = corr_matrix.loc[order, order]
        title_text += f'  |  ordinata per correlazione con {benchmark_sel}'

    fig = _corr_heatmap_figure(corr_matrix, title_text)
    date_message = (f"Range: {start_date.strftime('%d-%m-%Y')} — "
                    f"{end_date.strftime('%d-%m-%Y')} ({len(filtered_df)} giorni)")
    return fig, date_message, filter_info, True


@app.callback(
    Output('rolling-corr-chart',      'figure'),
    Input({'type': 'corr-ml', 'index': ALL}, 'value'),
    Input('benchmark-selector-corr',  'value'),
    State('stock-data',               'data'),
    State('dr-start-corr',            'date'),
    State('dr-end-corr',              'date'),
    State('correlation-window-input', 'value'),
    prevent_initial_call=True,
)
def update_rolling_corr(ml_vals, benchmark, stock_data, date_start, date_end, window):
    def _empty(msg=''):
        f = go.Figure()
        if msg:
            f.add_annotation(text=msg, xref='paper', yref='paper', x=0.5, y=0.5,
                             showarrow=False, font=dict(size=13, color='#888'))
        f.update_layout(paper_bgcolor='#fafafa', plot_bgcolor='#fafafa',
                        margin=dict(l=40, r=20, t=30, b=40))
        return f

    # Asset spuntati nella colonna ML (la spunta determina il rolling)
    selected_assets = [v[0] for v in (ml_vals or []) if v]
    if not selected_assets or not benchmark:
        return _empty('Spunta gli asset (colonna ML) e scegli un benchmark per la correlazione rolling')
    df = _get_df(stock_data)
    if df is None or df.empty:
        return _empty('Nessun dato disponibile')
    try:
        if date_start and date_end:
            df = df.loc[pd.to_datetime(date_start):pd.to_datetime(date_end)]
    except Exception:
        pass
    if benchmark not in df.columns:
        return _empty(f'Benchmark "{benchmark}" non trovato nel dataset')
    assets_to_plot = [a for a in selected_assets if a in df.columns and a != benchmark]
    if not assets_to_plot or df.empty:
        return _empty('Nessun asset valido selezionato')

    roll_window_days  = max(int(window or 60), 10)
    roll_window_weeks = max(roll_window_days // 5, 4)
    all_cols  = list(set(assets_to_plot + [benchmark]))
    df_weekly = (df[all_cols].dropna().resample('W')
                 .apply(lambda x: (1 + x).prod() - 1).dropna())
    if df_weekly.empty or len(df_weekly) < roll_window_weeks:
        return _empty('Dati insufficienti per la finestra rolling selezionata')

    fig = go.Figure()
    palette = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
               '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
    all_y = []
    for i, asset in enumerate(assets_to_plot):
        rolling_corr = (df_weekly[asset].rolling(roll_window_weeks)
                        .corr(df_weekly[benchmark]).dropna())
        all_y.extend(rolling_corr.values.tolist())
        fig.add_trace(go.Scatter(
            x=rolling_corr.index, y=rolling_corr.values, mode='lines',
            name=f'{asset} vs {benchmark}',
            line=dict(color=palette[i % len(palette)], width=1.8),
            hovertemplate=f'{asset} vs {benchmark}: %{{y:.3f}}<extra></extra>'))

    if all_y:
        y_min, y_max = min(all_y), max(all_y)
        margin = max((y_max - y_min) * 0.08, 0.05)
        y_range = [y_min - margin, y_max + margin]
    else:
        y_min, y_max, y_range = -1.0, 1.0, [-1.05, 1.05]
    fig.add_hline(y=0, line_dash='dash', line_color='#999', line_width=1)
    if y_max >= 0.7:
        fig.add_hline(y=0.7, line_dash='dot', line_color='#2ca02c', line_width=1,
                      annotation_text='0.70', annotation_position='right')
    if y_min <= -0.5:
        fig.add_hline(y=-0.5, line_dash='dot', line_color='#d62728', line_width=1,
                      annotation_text='-0.50', annotation_position='right')
    fig.update_layout(
        title=dict(text=f'Correlazione Rolling vs {benchmark} '
                        f'({roll_window_weeks} settimane – rend. settimanali)', font=dict(size=13), x=0.5),
        xaxis_title='Data', yaxis_title='Correlazione', yaxis=dict(range=y_range, zeroline=False),
        legend=dict(orientation='v', yanchor='top', y=1, xanchor='left', x=1.01, font=dict(size=10)),
        margin=dict(l=50, r=190, t=50, b=40), hovermode='x unified',
        paper_bgcolor='white', plot_bgcolor='#f9f9f9')
    return fig


# ─────────────────────────────────────────────────────────────────────────────
# Callback: cambio file dataset
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('active-xlsx-file',          'data'),
    Output('stock-data',                'data',     allow_duplicate=True),
    Output('original-prices-data',      'data',     allow_duplicate=True),
    Output('asset-checklist',           'data',     allow_duplicate=True),
    Output('ticker-map-store',          'data',     allow_duplicate=True),
    Output('data-last-updated',         'children', allow_duplicate=True),
    Output('upload-status',             'children', allow_duplicate=True),
    Output('weights-store-P1',          'data',     allow_duplicate=True),
    Output('weights-store-P2',          'data',     allow_duplicate=True),
    Output('weights-store-P3',          'data',     allow_duplicate=True),
    Output('update-portfolio-button',   'n_clicks', allow_duplicate=True),
    Input('file-selector',              'value'),
    State('update-portfolio-button',    'n_clicks'),
    prevent_initial_call=True,
)
def on_file_selected(filename, cur_clicks):
    if not filename:
        raise PreventUpdate
    _u = _get_username()

    # UN SOLO file: "👤 Personale" è solo l'ETICHETTA dello stato corrente —
    # current.json È GIÀ il file personale, non c'è nulla da ricaricare.
    # (Per questo l'accodo che porta il selettore a Personale non cambia i dati.)
    if filename == '__personale__':
        raise PreventUpdate

    # Selezione di un DEFAULT (ETF/CRIPTO/Commodities): sovrascrive il file di
    # lavoro col default fresco e riporta lo stato a "default" (non personale).
    _active_file_store['filename'] = filename
    _active_file_store['is_personale'] = False
    _cl_clear(_u)
    _tipo_def = f'default:{filename}'

    _noup4 = (no_update,) * 4

    cache = _file_cache_path(filename)
    if cache.exists():
        try:
            with open(cache, 'rb') as f:
                data = pickle.load(f)
            cr       = data.get('close_returns')
            op       = data.get('original_prices')
            tm       = data.get('ticker_map', {})
            vm       = data.get('valuta_map', {})
            saved_at = data.get('saved_at', '')
            if cr is not None:
                with _DL_LOCK:
                    _DL_BUFFER.clear()
                    _DL_BUFFER.update(data)
                _write_user_json(cr, op, tm, vm, reset_state=True, tipo=_tipo_def)
                options = [{'label': c, 'value': c} for c in cr.columns]
                return (filename,
                        cr.to_json(date_format='iso', orient='split'),
                        op.to_json(date_format='iso', orient='split'),
                        options, tm,
                        f'Aggiornati: {saved_at}',
                        html.Div(f'✓ {len(options)} asset — {Path(filename).stem}',
                                 style={'color': '#007755', 'font-size': '11px'}),
                        {}, {}, {}, (cur_clicks or 0) + 1)
        except Exception:
            pass

    # Cache non trovata → scarica da Yahoo Finance
    try:
        tickers, descr, valuta = _build_ticker_list(filename)
    except Exception:
        raise PreventUpdate
    options  = [{'label': d, 'value': d} for d in descr]
    tm       = {descr[i]: tickers[i] for i in range(len(tickers))}
    start    = (pd.Timestamp.today() - pd.DateOffset(years=10)).strftime('%Y-%m-%d')
    with _DL_LOCK:
        _DL_BUFFER.clear()
        _DL_STATE.update({'status': 'running', 'current': 0, 'total': len(tickers), 'errors': []})
    threading.Thread(target=_do_download,
                     args=(tickers, descr, valuta, start),
                     kwargs={'cache_file': cache, 'username': _get_username(),
                             'tipo': _tipo_def},
                     daemon=True).start()
    print(f"▶ Download {filename}: {len(tickers)} ticker")
    return (filename, None, None, options, tm,
            f'Download {Path(filename).stem}…',
            html.Div(f'⏳ Download {Path(filename).stem} — {len(tickers)} asset…',
                     style={'color': '#e67e22', 'font-size': '11px'}),
            {}, {}, {}, (cur_clicks or 0) + 1)


# ─────────────────────────────────────────────────────────────────────────────
# Callback: apri/chiudi modal editor lista asset
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('file-editor-overlay', 'style'),
    Output('file-editor-table',   'data'),
    Output('file-editor-title',   'children'),
    Output('file-editor-status',  'children', allow_duplicate=True),
    Input('gestisci-btn',         'n_clicks'),
    Input('file-editor-close',    'n_clicks'),
    State('file-selector',        'value'),
    prevent_initial_call=True,
)
def toggle_file_editor(open_n, close_n, filename):
    triggered = callback_context.triggered_id
    if triggered == 'gestisci-btn' and open_n:
        fn   = filename or 'ETF.xlsx'
        _u   = _get_username()
        rows = None

        # 1. Buffer cliente: file caricato via drag-and-drop nella sessione corrente (priorità max)
        with _CL_LOCK:
            cl_status = _CL_STATES.get(_u, {}).get('status', 'idle')
            cl_tm     = dict(_CL_BUFFERS.get(_u, {}).get('ticker_map', {}))
        if cl_status == 'done' and cl_tm:
            rows = [{'ticker': v, 'descrizione': k, 'valuta': 'EUR'}
                    for k, v in cl_tm.items()]

        # 2. JSON utente (source of truth — sempre coerente con i dati caricati)
        if not rows:
            ns = _read_user_json(_u)
            if ns:
                rows = [{'ticker': v['ticker'], 'descrizione': k, 'valuta': v.get('currency', 'EUR')}
                        for k, v in ns.items()]

        # 3. Xlsx del file selezionato (fallback diretto)
        if not rows:
            rows = _load_xlsx_rows(fn)

        stem = Path(fn).stem
        return _EDITOR_SHOWN, rows, f'Gestisci lista: {stem}', ''
    return _EDITOR_HIDDEN, no_update, no_update, no_update


# ─────────────────────────────────────────────────────────────────────────────
# Callback: aggiungi riga al DataTable editor
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('file-editor-table', 'data',         allow_duplicate=True),
    Output('new-ticker-input',  'value'),
    Output('new-desc-input',    'value'),
    Input('add-ticker-btn',     'n_clicks'),
    State('new-ticker-input',   'value'),
    State('new-desc-input',     'value'),
    State('new-valuta-dropdown','value'),
    State('file-editor-table',  'data'),
    prevent_initial_call=True,
)
def add_ticker_row(n, ticker, desc, valuta, rows):
    if not ticker or not ticker.strip():
        raise PreventUpdate
    new_row = {
        'ticker':      ticker.strip().upper(),
        'descrizione': (desc.strip() if desc and desc.strip() else ticker.strip().upper()),
        'valuta':      valuta or 'EUR',
    }
    return (rows or []) + [new_row], '', ''


# ─────────────────────────────────────────────────────────────────────────────
# Callback: salva xlsx e avvia download in background
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('file-editor-status',       'children',    allow_duplicate=True),
    Output('file-editor-overlay',      'style',       allow_duplicate=True),
    Output('file-selector',            'options',     allow_duplicate=True),
    Output('refresh-poll-interval',    'disabled',    allow_duplicate=True),
    Output('refresh-poll-interval',    'n_intervals', allow_duplicate=True),
    Output('refresh-data-btn',         'disabled',    allow_duplicate=True),
    Output('asset-checklist',          'data',        allow_duplicate=True),
    Input('save-file-editor-btn',      'n_clicks'),
    State('file-editor-table',         'data'),
    State('file-selector',             'value'),
    prevent_initial_call=True,
)
def save_file_editor(n, rows, filename):
    if not rows:
        return '⚠ La lista è vuota, nessun file salvato.', no_update, no_update, no_update, no_update, no_update, no_update
    filename = filename or 'ETF.xlsx'
    _u       = _get_username()
    _active_file_store['filename'] = filename

    try:
        tickers     = [r['ticker']                         for r in rows if str(r.get('ticker', '')).strip()]
        descr       = [r.get('descrizione') or r['ticker'] for r in rows if str(r.get('ticker', '')).strip()]
        valuta_list = [r.get('valuta', 'EUR')              for r in rows if str(r.get('ticker', '')).strip()]

        user_pkl = _user_cache_path(_u, filename)
        new_tm   = {descr[i]: tickers[i] for i in range(len(tickers))}
        want     = set(descr)

        # ── Determina la sorgente dati già scaricati ──────────────────────────
        # Priorità: buffer cliente (file caricato via drag-and-drop) >
        #           pkl utente (Gestisci precedente) > cache di default
        src_data = None
        with _CL_LOCK:
            cl_status = _CL_STATES.get(_u, {}).get('status', 'idle')
            cl_buf    = dict(_CL_BUFFERS.get(_u, {}))

        if cl_status == 'done' and cl_buf.get('original_prices') is not None:
            src_data = cl_buf
            # Salva subito il file personale come pkl utente così i prossimi
            # accessi lo trovano anche dopo che CL_BUFFERS sarà resettato
            _atomic_pkl_write(user_pkl, cl_buf)

        if src_data is None and user_pkl.exists():
            try:
                with open(user_pkl, 'rb') as f:
                    src_data = pickle.load(f)
            except Exception:
                pass

        if src_data is None:
            default_pkl = _file_cache_path(filename)
            if default_pkl.exists():
                try:
                    with open(default_pkl, 'rb') as f:
                        src_data = pickle.load(f)
                except Exception:
                    pass

        existing_descr = set()
        if src_data is not None:
            existing_descr = set(src_data.get('ticker_map', {}).keys())
            if not existing_descr and src_data.get('original_prices') is not None:
                existing_descr = set(src_data['original_prices'].columns)

        # Resetta il buffer cliente — ora il pkl utente è la fonte di verità
        _cl_clear(_u)

        # ── Ticker da scaricare (quelli non già presenti nella sorgente) ──────
        new_idx    = [i for i, d in enumerate(descr) if d not in existing_descr]
        dl_tickers = [tickers[i] for i in new_idx]
        dl_descr   = [descr[i]   for i in new_idx]
        dl_valuta  = [valuta_list[i] for i in new_idx]
        start      = (pd.Timestamp.today() - pd.DateOffset(years=10)).strftime('%Y-%m-%d')

        if not dl_tickers:
            # Nessun ticker nuovo — applica solo aggiunte/rimozioni alla sorgente
            if src_data is not None and src_data.get('original_prices') is not None:
                src_op    = src_data['original_prices']
                keep_cols = [c for c in src_op.columns if c in want]
                sliced    = {}
                for key, val in src_data.items():
                    if isinstance(val, pd.DataFrame) and not val.empty:
                        sliced[key] = val[[c for c in keep_cols if c in val.columns]]
                    else:
                        sliced[key] = val
                sliced['ticker_map']    = new_tm
                sliced['close_returns'] = sliced['original_prices'].pct_change(fill_method=None)
                sliced['saved_at']      = datetime.now().strftime('%d/%m/%Y %H:%M')
                _atomic_pkl_write(user_pkl, sliced)
                with _DL_LOCK:
                    _DL_BUFFER.update(sliced)
                    _DL_STATE.update({'status': 'done', 'current': len(keep_cols),
                                      'total': len(keep_cols), 'errors': []})
                _write_user_json(sliced['close_returns'], sliced['original_prices'], new_tm)
            return ('✓ Lista aggiornata — nessun nuovo asset da scaricare.',
                    no_update, _list_files(),
                    False, 0, True, no_update)

        # ── Avvia download in background, editor rimane aperto ───────────────
        print(f"▶ Download gestisci: {len(dl_tickers)} nuovi ticker per {_u}")
        t = threading.Thread(
            target=_do_gestisci_download,
            args=(dl_tickers, dl_descr, dl_valuta, start, _u, filename, descr),
            daemon=True,
        )
        t.start()

    except Exception as e:
        print(f"⚠ Errore save_file_editor: {e}")
        return f'⚠ {e}', _EDITOR_HIDDEN, _list_files(), no_update, no_update, no_update, no_update

    return (f'⏳ Download {len(dl_tickers)} asset in corso…',
            no_update, _list_files(),
            False, 0, True, no_update)


# ─────────────────────────────────────────────────────────────────────────────
# Callback: avvia aggiornamento manuale
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('refresh-poll-interval',   'disabled',    allow_duplicate=True),
    Output('refresh-poll-interval',   'n_intervals', allow_duplicate=True),
    Output('refresh-data-btn',        'disabled',    allow_duplicate=True),
    Output('progress-modal-overlay',  'style',       allow_duplicate=True),
    Output('modal-progress-fill',     'style',       allow_duplicate=True),
    Output('modal-pct-text',          'children',    allow_duplicate=True),
    Output('modal-status-text',       'children',    allow_duplicate=True),
    Output('modal-status-text',       'style',       allow_duplicate=True),
    Input('refresh-data-btn', 'n_clicks'),
    State('dr-start-tab1',    'date'),
    prevent_initial_call=True,
)
def start_refresh(n_clicks, start_date_picker):
    if not n_clicks:
        raise PreventUpdate

    with _DL_LOCK:
        if _DL_STATE.get('status') == 'running':
            current = _DL_STATE.get('current', 0)
            total   = _DL_STATE.get('total', 1) or 1
            pct     = int(current / total * 100)
            return (False, 0, True, _MODAL_SHOWN,
                    {**_FILL_LOADING, 'width': f'{pct}%'},
                    f'Aggiornamento in corso: {current}/{total}…',
                    'Un aggiornamento è già in corso…', _STATUS_GREY)

        # Prendi i ticker pendenti da Gestisci (se presenti)
        pending = dict(_PENDING)
        _PENDING.clear()

    active_file = _active_file_store.get('filename', 'ETF.xlsx')
    cache = _file_cache_path(active_file)
    _cl_clear(_get_username())

    if pending:
        # Nuovi ticker da Gestisci: download solo quelli + merge nel pkl
        dl_tickers  = pending['tickers']
        dl_descr    = pending['descr']
        dl_valuta   = pending['valuta']
        incremental = pending.get('incremental', True)
        start_date  = pending['start']
        with _DL_LOCK:
            _DL_STATE.update({'status': 'running', 'current': 0,
                              'total': len(dl_tickers), 'errors': []})
        threading.Thread(
            target=_do_full_update,
            args=(dl_tickers, dl_descr, dl_valuta, start_date, cache),
            kwargs={'incremental': incremental},
            daemon=True,
        ).start()
        print(f"▶ Download {len(dl_tickers)} nuovi ticker per {active_file}")
        label = f'{len(dl_tickers)} nuovi asset'
    else:
        # Nessun ticker pendente: ricarica il file di default dal disco (nessun download)
        threading.Thread(
            target=_do_reload_from_disk,
            args=(cache, active_file),
            daemon=True,
        ).start()
        print(f"▶ Ricaricamento {active_file} da disco")
        label = f'Caricamento {Path(active_file).stem}'

    return (False, 0, True, _MODAL_SHOWN, _FILL_LOADING,
            f'{label}…', '', _STATUS_GREY)


# ─────────────────────────────────────────────────────────────────────────────
# Callback: polling aggiornamento
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('stock-data',              'data',     allow_duplicate=True),
    Output('original-prices-data',    'data',     allow_duplicate=True),
    Output('asset-checklist',         'data',     allow_duplicate=True),
    Output('ticker-map-store',        'data',     allow_duplicate=True),
    Output('data-last-updated',       'children', allow_duplicate=True),
    Output('refresh-poll-interval',   'disabled', allow_duplicate=True),
    Output('refresh-data-btn',        'disabled', allow_duplicate=True),
    Output('modal-progress-fill',     'style',    allow_duplicate=True),
    Output('modal-pct-text',          'children', allow_duplicate=True),
    Output('modal-status-text',       'children', allow_duplicate=True),
    Output('modal-status-text',       'style',    allow_duplicate=True),
    Output('progress-modal-overlay',  'style',    allow_duplicate=True),
    Output('file-editor-overlay',     'style',    allow_duplicate=True),
    Output('file-selector',           'options',  allow_duplicate=True),
    Output('file-selector',           'value',    allow_duplicate=True),
    Output('update-portfolio-button', 'n_clicks', allow_duplicate=True),
    Output('inline-add-status',       'children', allow_duplicate=True),
    Output('weights-store-P1',        'data',     allow_duplicate=True),
    Output('weights-store-P2',        'data',     allow_duplicate=True),
    Output('weights-store-P3',        'data',     allow_duplicate=True),
    Input('refresh-poll-interval', 'n_intervals'),
    State('update-portfolio-button',  'n_clicks'),
    prevent_initial_call=True,
)
def poll_refresh_progress(n, n_btn):
    _u = _get_username()
    cl_state  = dict(_cl_state(_u))
    cl_buffer = dict(_cl_buf(_u))
    with _DL_LOCK:
        dl_state  = dict(_DL_STATE)
        dl_buffer = dict(_DL_BUFFER)

    client_active = cl_state.get('status') in ('running', 'done') and cl_state.get('status') != 'idle'
    state  = cl_state  if client_active else dl_state
    buffer = cl_buffer if client_active else dl_buffer

    def _ncol(b):
        cr = b.get('close_returns')
        return len(cr.columns) if cr is not None else 0
    print(f"[POLL] client_active={client_active} cl_status={cl_state.get('status')} "
          f"dl_status={dl_state.get('status')} cl_cols={_ncol(cl_buffer)} dl_cols={_ncol(dl_buffer)}", flush=True)

    status  = state.get('status', 'idle')
    current = state.get('current', 0)
    total   = state.get('total', 1) or 1
    pct     = int(current / total * 100)
    modal_fill = {**_FILL_LOADING, 'width': f'{pct}%'}

    if status == 'idle':
        raise PreventUpdate

    _NU3 = (no_update, no_update, no_update)
    if status == 'running':
        return (no_update, no_update, no_update, no_update, no_update,
                False, True,
                modal_fill, f'{current} / {total}  ({pct}%)',
                'Download in corso…', _STATUS_GREY, no_update, no_update,
                no_update, no_update, no_update, no_update, *_NU3)

    if status == 'error':
        err_fill = {**_FILL_LOADING, 'width': '100%', 'background': '#c0392b'}
        return (no_update, no_update, no_update, no_update, no_update,
                True, False,
                err_fill, '❌ Download fallito',
                'Si è verificato un errore.', _STATUS_RED, no_update, _EDITOR_HIDDEN,
                no_update, no_update, no_update, '❌ Errore download', *_NU3)

    # FONTE UNICA: leggi il dataset da current.json (così non dipende da quale
    # buffer ha scritto il download). Fallback al buffer se current.json è vuoto.
    close_returns, original_prices, ticker_map = dc.build_dataset(_u)
    if close_returns is None or close_returns.empty:
        close_returns   = buffer.get('close_returns')
        original_prices = buffer.get('original_prices')
        ticker_map      = buffer.get('ticker_map', {})
    if close_returns is None or close_returns.empty:
        err_fill = {**_FILL_LOADING, 'width': '100%', 'background': '#c0392b'}
        return (no_update, no_update, no_update, no_update, no_update,
                True, False,
                err_fill, '❌ Nessun dato ricevuto',
                'Il download è terminato senza dati.', _STATUS_RED, no_update, _EDITOR_HIDDEN,
                no_update, no_update, no_update, '❌ Nessun dato', *_NU3)

    options      = [{'label': col, 'value': col} for col in close_returns.columns]
    saved_at     = buffer.get('saved_at', '')
    returns_json = close_returns.to_json(date_format='iso', orient='split')
    prices_json  = original_prices.to_json(date_format='iso', orient='split')
    ok_fill      = {**_FILL_LOADING, 'width': '100%'}

    errors    = state.get('errors', [])
    n_ok      = len(options)
    n_err     = len(errors)
    err_note  = f' — ⚠ {n_err} non trovati su Yahoo' if n_err else ''
    status_msg = (f'Dati pronti — {n_ok} asset scaricati{err_note}.\n'
                  + ('\n'.join(errors[:5]) if errors else ''))

    # Ripristina i pesi P1/P2/P3 DAI dati correnti (current.json), così aggiungere
    # un asset NON cancella i portafogli creati. (Azzerare le store non bastava:
    # se c'erano asset selezionati, la griglia non li rileggeva da current.json.)
    _nsw = _read_user_json(_u)
    p1w = {d: v.get('P1', 0) for d, v in _nsw.items() if v.get('P1', 0)}
    p2w = {d: v.get('P2', 0) for d, v in _nsw.items() if v.get('P2', 0)}
    p3w = {d: v.get('P3', 0) for d, v in _nsw.items() if v.get('P3', 0)}

    return (
        returns_json, prices_json, options, ticker_map,
        f"Aggiornati: {saved_at}",
        True, False,
        ok_fill, f'✓ {n_ok} asset{err_note}',
        status_msg, _STATUS_GREEN if not n_err else {**_STATUS_GREEN, 'color': '#b8860b'},
        _MODAL_HIDDEN, _EDITOR_HIDDEN,
        no_update, no_update,
        (n_btn or 0) + 1, f'✓ {n_ok} asset caricati',
        p1w, p2w, p3w,   # pesi ripristinati da current.json (no wipe dei portafogli)
    )


# ─────────────────────────────────────────────────────────────────────────────
# Callback: elimina asset dalla lista inline
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('stock-data',              'data',     allow_duplicate=True),
    Output('asset-checklist',         'data',     allow_duplicate=True),
    Output('update-portfolio-button', 'n_clicks', allow_duplicate=True),
    Output('file-selector',           'options',  allow_duplicate=True),
    Output('file-selector',           'value',    allow_duplicate=True),
    Input({'type': 'delete-asset-btn', 'index': ALL}, 'n_clicks'),
    State('asset-checklist',          'data'),
    State('update-portfolio-button',  'n_clicks'),
    State('file-selector',            'value'),
    prevent_initial_call=True,
)
def delete_asset_inline(clicks, options, current_clicks, filename):
    triggered = callback_context.triggered_id
    if not triggered or not any(c for c in (clicks or [])):
        raise PreventUpdate
    asset = triggered['index']
    _u = _get_username()
    fn = filename or 'ETF.xlsx'

    with _DL_LOCK:
        cr = _DL_BUFFER.get('close_returns')
        op = _DL_BUFFER.get('original_prices')
        tm = dict(_DL_BUFFER.get('ticker_map', {}))
        if cr is not None and asset in cr.columns:
            _DL_BUFFER['close_returns'] = cr.drop(columns=[asset])
        if op is not None and asset in op.columns:
            _DL_BUFFER['original_prices'] = op.drop(columns=[asset])
        tm.pop(asset, None)
        _DL_BUFFER['ticker_map'] = tm
        new_cr = _DL_BUFFER.get('close_returns')
        new_op = _DL_BUFFER.get('original_prices')

    ns = _read_user_json(_u)
    if asset in ns:
        del ns[asset]
        # Scrittura ATOMICA (mai json.dump diretto su current.json: non atomico →
        # con scritture concorrenti corrompe il file). _mark_personale() più sotto
        # ripristina il campo _tipo.
        _atomic_json_write(_user_json_path(_u), ns, validate=False)

    user_pkl = _user_cache_path(_u, fn)
    try:
        if user_pkl.exists():
            data = pickle.load(open(user_pkl, 'rb'))
            for key, val in list(data.items()):
                if isinstance(val, pd.DataFrame) and asset in val.columns:
                    data[key] = val.drop(columns=[asset])
            data['ticker_map'] = tm
        else:
            data = {'close_returns': new_cr, 'original_prices': new_op,
                    'ticker_map': tm, 'saved_at': datetime.now().strftime('%d/%m/%Y %H:%M')}
        _atomic_pkl_write(user_pkl, data)
    except Exception as e:
        print(f'⚠ delete_asset_inline pkl: {e}')

    # Modifica del file (rimozione titolo) → diventa Personale e si persiste.
    _mark_personale(_u)
    new_options = [o for o in (options or []) if o['value'] != asset]
    new_stock = new_cr.to_json(date_format='iso', orient='split') if new_cr is not None else no_update
    return new_stock, new_options, (current_clicks or 0) + 1, _list_files_with_personale(), '__personale__'


# ─────────────────────────────────────────────────────────────────────────────
# Callback: aggiungi asset inline
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('inline-ticker-input',     'value',    allow_duplicate=True),
    Output('inline-desc-input',       'value',    allow_duplicate=True),
    Output('inline-add-status',       'children', allow_duplicate=True),
    Output('refresh-poll-interval',   'disabled', allow_duplicate=True),
    Output('refresh-poll-interval',   'n_intervals', allow_duplicate=True),
    Output('refresh-data-btn',        'disabled', allow_duplicate=True),
    Output('file-selector',           'options',  allow_duplicate=True),
    Output('file-selector',           'value',    allow_duplicate=True),
    Input('inline-add-btn',           'n_clicks'),
    State('inline-ticker-input',      'value'),
    State('inline-desc-input',        'value'),
    State('inline-valuta-dropdown',   'value'),
    State('file-selector',            'value'),
    State('stock-data',               'data'),
    State('original-prices-data',     'data'),
    State('ticker-map-store',         'data'),
    prevent_initial_call=True,
)
def add_asset_inline(n, ticker, desc, valuta, filename, sd_json, op_json, tm_store):
    if not n or not ticker or not ticker.strip():
        raise PreventUpdate
    ticker = ticker.strip().upper()
    desc   = (desc or ticker).strip()
    valuta = valuta or 'EUR'
    _u = _get_username()
    fn = filename if filename and filename != '__personale__' else 'ETF.xlsx'

    # Riallinea current.json con ciò che è MOSTRATO a schermo (store del grid):
    # all'avvio il display può venire dal buffer ETF (es. 29 asset) mentre
    # current.json è rimasto a una sessione precedente (es. 21). Sincronizzando
    # qui, l'accodo parte sempre dallo stesso identico set visualizzato.
    # reset_state=False → preserva pesi P1/P2/P3 e selezioni già impostati.
    try:
        from io import StringIO as _SIO
        if sd_json and op_json:
            _disp_cr = pd.read_json(_SIO(sd_json), orient='split')
            _disp_op = pd.read_json(_SIO(op_json), orient='split')
            if not _disp_cr.empty and not _disp_op.empty:
                _write_user_json(_disp_cr, _disp_op, dict(tm_store or {}), username=_u)
    except Exception as _e:
        print(f"⚠ resync current.json pre-add: {_e}", flush=True)

    # Asset ESISTENTI dalla fonte unica current.json (ora allineata al display),
    # così l'accodo non perde mai gli asset già presenti. Fallback al buffer.
    try:
        _cur = json.load(open(_user_json_path(_u)))
        existing_tm = {a: (v.get('ticker') or a)
                       for a, v in _cur.items() if isinstance(v, dict)}
    except Exception:
        existing_tm = {}
    if not existing_tm:
        with _DL_LOCK:
            existing_tm = dict(_DL_BUFFER.get('ticker_map', {}))

    if desc in existing_tm:
        return no_update, no_update, f'⚠ "{desc}" già presente', no_update, no_update, no_update, no_update, no_update

    start     = (pd.Timestamp.today() - pd.DateOffset(years=10)).strftime('%Y-%m-%d')
    all_descr = list(existing_tm.keys()) + [desc]
    # Aggiungere un titolo modifica il file → diventa Personale (un solo file:
    # marco _tipo dentro current.json, niente copie/altri file).
    _active_file_store['is_personale'] = True
    _active_file_store['filename'] = '__personale__'
    _set_tipo('personale', username=_u)
    threading.Thread(
        target=_do_gestisci_download,
        args=([ticker], [desc], [valuta], start, _u, fn, all_descr),
        kwargs={'persist_personale': True},
        daemon=True,
    ).start()
    return '', '', f'⏳ Download {ticker}…', False, 0, True, _list_files_with_personale(), '__personale__'


# ─────────────────────────────────────────────────────────────────────────────
# Callback: chiudi modale aggiornamento
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('progress-modal-overlay', 'style', allow_duplicate=True),
    Input('progress-modal-close', 'n_clicks'),
    prevent_initial_call=True,
)
def close_progress_modal(n):
    if not n:
        raise PreventUpdate
    return _MODAL_HIDDEN


# ─────────────────────────────────────────────────────────────────────────────
# Callback: salva dati (download Excel)
# ─────────────────────────────────────────────────────────────────────────────
# Clientside: mostra overlay immediatamente al click (callback primario per download-overlay.style)
app.clientside_callback(
    """
    function(n) {
        if (!n || n === 0) return {display: 'none'};
        return {
            display: 'flex', position: 'fixed', top: '0', left: '0',
            width: '100%', height: '100%',
            background: 'rgba(26,58,92,0.45)',
            zIndex: '2000', justifyContent: 'center', alignItems: 'center'
        };
    }
    """,
    Output('download-overlay', 'style'),
    Input('save-data-button',  'n_clicks'),
    prevent_initial_call=True,
)

@app.callback(
    Output('download-data',    'data'),
    Output('download-status',  'children'),
    Output('download-overlay', 'style', allow_duplicate=True),
    Input('save-data-button',  'n_clicks'),
    State('original-prices-data', 'data'),
    prevent_initial_call=True,
)
def salva_dati(n_clicks, original_prices_data):
    if not n_clicks or n_clicks == 0:
        raise PreventUpdate
    try:
        if not dc.read_current(_get_username()):
            return (no_update,
                    html.Div('⚠ Nessun dato disponibile — clicca prima ⟳ Aggiorna',
                             style={'color': '#e67e22', 'font-size': '11px'}),
                    _MODAL_HIDDEN)
        # Esporta dal file UNICO current.json (fogli Asset + Prezzi) via modulo condiviso
        return (
            dcc.send_bytes(lambda b: b.write(dc.export_bytes(_get_username())),
                           'dati_portafoglio.xlsx'),
            html.Div('✓ File scaricato', style={'color': 'green', 'font-size': '11px'}),
            _MODAL_HIDDEN,
        )
    except Exception as e:
        return no_update, html.Div(f'Errore: {e}', style={'color': 'red'}), _MODAL_HIDDEN


# ─────────────────────────────────────────────────────────────────────────────
# Callback: importa portafogli dalla Frontiera Efficiente
# ─────────────────────────────────────────────────────────────────────────────
_FE_PORT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'sessions', 'fe_portfolio.json')

@app.callback(
    Output('import-frontier-confirm', 'displayed'),
    Output('import-frontier-msg', 'children'),
    Input('import-frontier-btn', 'n_clicks'),
    prevent_initial_call=True,
)
def ask_import_frontier(n):
    if not n:
        raise PreventUpdate
    if not os.path.exists(_FE_PORT_FILE):
        return False, '⚠ Nessun portafoglio salvato dalla Frontiera'
    try:
        with open(_FE_PORT_FILE) as f:
            data = json.load(f)
        saved_at = data.get('saved_at', '')
        ports = list(data.get('portfolios', {}).keys())
        return True, f'Trovato: {", ".join(ports)} del {saved_at}'
    except Exception as e:
        return False, f'⚠ Errore lettura: {e}'


@app.callback(
    Output('weights-store-P1', 'data', allow_duplicate=True),
    Output('weights-store-P2', 'data', allow_duplicate=True),
    Output('weights-store-P3', 'data', allow_duplicate=True),
    Output({'type': 'weight-input', 'index': ALL}, 'value', allow_duplicate=True),
    Output('import-frontier-msg', 'children', allow_duplicate=True),
    Input('import-frontier-confirm', 'submit_n_clicks'),
    State({'type': 'weight-input', 'index': ALL}, 'id'),
    State({'type': 'weight-input', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def do_import_frontier(submit, all_ids, all_vals):
    if not submit:
        raise PreventUpdate
    try:
        with open(_FE_PORT_FILE) as f:
            data = json.load(f)
        ports = data.get('portfolios', {})
        w1 = ports.get('F1', {})
        w2 = ports.get('F2', {})
        w3 = ports.get('F3', {})

        new_vals = []
        for inp_id, cur_val in zip(all_ids, all_vals):
            idx = inp_id['index']  # es. 'P1-Az. ACWI'
            if idx.startswith('P1-'):
                asset = idx[3:]
                new_vals.append(w1.get(asset, 0) if w1 else cur_val)
            elif idx.startswith('P2-'):
                asset = idx[3:]
                new_vals.append(w2.get(asset, 0) if w2 else cur_val)
            elif idx.startswith('P3-'):
                asset = idx[3:]
                new_vals.append(w3.get(asset, 0) if w3 else cur_val)
            else:
                new_vals.append(cur_val)

        imported = [k for k in ['F1','F2','F3'] if k in ports]
        return w1 or no_update, w2 or no_update, w3 or no_update, new_vals, f'✓ Importato: {", ".join(imported)}'
    except Exception as e:
        return no_update, no_update, no_update, all_vals, f'⚠ Errore: {e}'


# ═════════════════════════════════════════════════════════════════════════════
# IMPORTA / ESPORTA PORTAFOGLIO — profili condivisi (sessions_manager)
# ═════════════════════════════════════════════════════════════════════════════
_PIO_OVERLAY = {'position': 'fixed', 'top': '0', 'left': '0', 'width': '100%',
                'height': '100%', 'z-index': '9000', 'background': 'rgba(0,0,0,0.45)',
                'align-items': 'center', 'justify-content': 'center'}

@app.callback(
    Output('pio-overlay',     'style'),
    Output('pio-exp-profile', 'options'),
    Output('pio-imp-profile', 'options'),
    Input('port-io-btn', 'n_clicks'),
    Input('pio-close',   'n_clicks'),
    prevent_initial_call=True,
)
def toggle_pio(open_n, close_n):
    if callback_context.triggered_id == 'port-io-btn':
        opts = [{'label': a, 'value': a} for a in _sm.list_analyses(_get_username())]
        return {**_PIO_OVERLAY, 'display': 'flex'}, opts, opts
    return {**_PIO_OVERLAY, 'display': 'none'}, no_update, no_update


# Reset dei campi del dialogo ad ogni apertura (evita stato vecchio)
@app.callback(
    Output('pio-mode',        'value',    allow_duplicate=True),
    Output('pio-imp-profile', 'value',    allow_duplicate=True),
    Output('pio-imp-status',  'children', allow_duplicate=True),
    Output('pio-exp-status',  'children', allow_duplicate=True),
    Output('pio-exp-new',     'value',    allow_duplicate=True),
    Output('pio-exp-profile', 'value',    allow_duplicate=True),
    Input('port-io-btn', 'n_clicks'),
    prevent_initial_call=True,
)
def pio_reset(n):
    if not n:
        raise PreventUpdate
    return 'export', None, '', '', '', None


@app.callback(
    Output('pio-export-view', 'style'),
    Output('pio-import-view', 'style'),
    Input('pio-mode', 'value'),
)
def pio_switch_view(mode):
    if mode == 'import':
        return {'display': 'none'}, {'display': 'block'}
    return {'display': 'block'}, {'display': 'none'}


@app.callback(
    Output('pio-exp-status',  'children'),
    Output('pio-exp-profile', 'options', allow_duplicate=True),
    Output('pio-imp-profile', 'options', allow_duplicate=True),
    Output('pio-exp-new',     'value'),
    Input('pio-exp-btn', 'n_clicks'),
    State('pio-exp-col',     'value'),
    State({'type': 'weight-input', 'index': ALL}, 'id'),
    State({'type': 'weight-input', 'index': ALL}, 'value'),
    State('pio-exp-profile', 'value'),
    State('pio-exp-new',     'value'),
    prevent_initial_call=True,
)
def pio_export(n, col, all_ids, all_vals, ana_sel, ana_new):
    if not n:
        raise PreventUpdate
    _u = _get_username()
    # Nome analisi: nuovo (prevale) o esistente da sovrascrivere
    name = (ana_new or '').strip() or (ana_sel or '').strip()
    if not name:
        return '⚠ Scrivi un nome nuovo o scegli un\'analisi da sovrascrivere', no_update, no_update, no_update

    col = col if col in ('P1', 'P2', 'P3') else 'P1'
    # 1) FONTE DI VERITÀ: current.json — sincronizzato da tutte le tab
    #    (la griglia del Portafoglio è obsoleta se hai calcolato in Frontiera).
    ns = _read_user_json(_u) or {}
    weights = {a: float(v.get(col, 0) or 0) for a, v in ns.items() if v.get(col, 0)}
    # 2) Fallback: griglia corrente (se current.json non ha pesi per quella colonna)
    if not weights:
        for inp_id, val in zip(all_ids or [], all_vals or []):
            idx = inp_id['index']
            if idx.startswith(col + '-') and val:
                try:
                    weights[idx[3:]] = float(val)
                except (TypeError, ValueError):
                    pass
    print(f"[PIO-EXPORT] col={col} | current.json+griglia={len(weights)} asset "
          f"{list(weights.items())[:4]}", flush=True)
    if not weights:
        return f'⚠ La colonna {col} non ha pesi da esportare', no_update, no_update, no_update

    # Meta autosufficiente: ticker+valuta di ogni asset (da current.json) per poter
    # ri-aggiungere/riscaricare gli asset mancanti al momento dell'import.
    ns = _read_user_json(_u) or {}
    meta = {a: {'ticker': (ns.get(a, {}) or {}).get('ticker', ''),
                'valuta': (ns.get(a, {}) or {}).get('currency', 'EUR')}
            for a in weights}
    ok = _sm.save_analysis(_u, name, weights, meta=meta)
    opts = [{'label': a, 'value': a} for a in _sm.list_analyses(_u)]
    if ok:
        return (f'✅ Colonna {col} salvata nell\'analisi "{name}" ({len(weights)} asset)',
                opts, opts, '')
    return '⚠ Errore durante l\'esportazione', no_update, no_update, no_update


def _import_add_missing_thread(username, dl_tickers, dl_descr, dl_valuta,
                               all_descr, target, full_weights, keep):
    """Thread: scarica+merge gli asset mancanti, poi applica i pesi alla colonna target."""
    start = (pd.Timestamp.today() - pd.DateOffset(years=10)).strftime('%Y-%m-%d')
    try:
        if dl_tickers:
            _do_gestisci_download(dl_tickers, dl_descr, dl_valuta, start,
                                  username, 'ETF.xlsx', all_descr=all_descr)
        w = {'P1': dict(keep.get('P1', {})), 'P2': dict(keep.get('P2', {})),
             'P3': dict(keep.get('P3', {}))}
        w[target] = full_weights
        _update_user_json(weights=w, username=username)
        print(f"✓ Import con download: +{len(dl_tickers)} asset, pesi su {target}")
    except Exception as e:
        print(f"⚠ Import add-missing fallito: {e}")
        with _DL_LOCK:
            _DL_STATE['status'] = 'error'


@app.callback(
    Output('weights-store-P1', 'data', allow_duplicate=True),
    Output('weights-store-P2', 'data', allow_duplicate=True),
    Output('weights-store-P3', 'data', allow_duplicate=True),
    Output({'type': 'weight-input', 'index': ALL}, 'value', allow_duplicate=True),
    Output('pio-imp-status', 'children'),
    Output('pio-overlay', 'style', allow_duplicate=True),
    Output('refresh-poll-interval',  'disabled',    allow_duplicate=True),
    Output('refresh-poll-interval',  'n_intervals', allow_duplicate=True),
    Output('progress-modal-overlay', 'style',       allow_duplicate=True),
    Output('refresh-data-btn',       'disabled',    allow_duplicate=True),
    Input('pio-imp-btn', 'n_clicks'),
    State('pio-imp-profile', 'value'),
    State('pio-imp-target',  'value'),
    State('weights-store-P1', 'data'),
    State('weights-store-P2', 'data'),
    State('weights-store-P3', 'data'),
    State({'type': 'weight-input', 'index': ALL}, 'id'),
    State({'type': 'weight-input', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def pio_import(n, analysis, target, p1d, p2d, p3d, all_ids, all_vals):
    _NU4 = (no_update, no_update, no_update, no_update)  # poll/progress (caso immediato)
    if not n:
        raise PreventUpdate
    if not analysis:
        return (no_update, no_update, no_update, no_update,
                "⚠ Scegli un'analisi da importare", no_update, *_NU4)
    _u = _get_username()
    target = target if target in ('P1', 'P2', 'P3') else 'P1'
    imported = {a: float(v) for a, v in (_sm.get_analysis(_u, analysis) or {}).items()}
    if not imported:
        return (no_update, no_update, no_update, no_update,
                '⚠ Analisi vuota', no_update, *_NU4)

    # Asset attualmente nel dataset: fonte di verità current.json (la griglia può
    # essere obsoleta se hai lavorato in un altro tab). Fallback sulle celle griglia.
    current_assets = set((_read_user_json(_u) or {}).keys())
    if not current_assets:
        current_assets = {d['index'][3:] for d in (all_ids or []) if d['index'].startswith('P1-')}
    missing = [a for a in imported if a not in current_assets]

    # ── CASO 1: tutti gli asset sono presenti → carica subito ─────────────────
    if not missing:
        slot = {'P1': dict(p1d or {}), 'P2': dict(p2d or {}), 'P3': dict(p3d or {})}
        slot[target] = imported
        new_vals = []
        for inp_id, curv in zip(all_ids or [], all_vals or []):
            idx = inp_id['index']
            if idx.startswith(target + '-'):
                new_vals.append(imported.get(idx[3:], 0))
            else:
                new_vals.append(curv)
        _update_user_json(weights={'P1': slot['P1'], 'P2': slot['P2'], 'P3': slot['P3']},
                          username=_u)
        msg = f'✓ Analisi "{analysis}" importata in {target} ({len(imported)} asset)'
        return (slot['P1'], slot['P2'], slot['P3'], new_vals, msg,
                {**_PIO_OVERLAY, 'display': 'none'}, *_NU4)

    # ── CASO 2: alcuni asset mancano dal dataset → ri-aggiungili e scaricali ──
    meta = _sm.get_analysis_meta(_u, analysis) or {}
    dl_descr   = [a for a in missing if (meta.get(a) or {}).get('ticker')]
    dl_tickers = [meta[a]['ticker'] for a in dl_descr]
    dl_valuta  = [(meta.get(a) or {}).get('valuta', 'EUR') for a in dl_descr]
    no_ticker  = [a for a in missing if not (meta.get(a) or {}).get('ticker')]

    if not dl_tickers:
        # Nessun ticker disponibile (analisi vecchia senza meta): carica solo i presenti
        present = {a: w for a, w in imported.items() if a in current_assets}
        slot = {'P1': dict(p1d or {}), 'P2': dict(p2d or {}), 'P3': dict(p3d or {})}
        slot[target] = present
        new_vals = [present.get(d['index'][3:], 0) if d['index'].startswith(target + '-')
                    else cv for d, cv in zip(all_ids or [], all_vals or [])]
        _update_user_json(weights={'P1': slot['P1'], 'P2': slot['P2'], 'P3': slot['P3']},
                          username=_u)
        msg = (f'⚠ {len(missing)} asset non nel portafoglio e senza ticker salvato: '
               f'caricati solo i {len(present)} presenti. Ri-esporta l\'analisi per includerli.')
        return (slot['P1'], slot['P2'], slot['P3'], new_vals, msg,
                {**_PIO_OVERLAY, 'display': 'none'}, *_NU4)

    # Mantieni le altre colonne; la target verrà riempita coi pesi completi dopo il download
    keep = {'P1': dict(p1d or {}), 'P2': dict(p2d or {}), 'P3': dict(p3d or {})}
    keep[target] = {}
    all_descr = list(dict.fromkeys(list(current_assets) + dl_descr))

    _cl_clear(_u)  # così il poll usa _DL_STATE (non il buffer cliente)
    with _DL_LOCK:
        _DL_STATE.update({'status': 'running', 'current': 0,
                          'total': len(dl_tickers), 'errors': []})
    threading.Thread(
        target=_import_add_missing_thread,
        args=(_u, dl_tickers, dl_descr, dl_valuta, all_descr, target, imported, keep),
        daemon=True,
    ).start()

    extra = f' ({len(no_ticker)} senza ticker, esclusi)' if no_ticker else ''
    msg = f'⏳ Aggiungo {len(dl_tickers)} asset mancanti e scarico i dati…{extra}'
    # Non imposto i pesi ora: il download + poll aggiornano griglia e pesi al termine
    return (no_update, no_update, no_update, no_update, msg,
            {**_PIO_OVERLAY, 'display': 'none'},
            False, 0, _MODAL_SHOWN, True)


# ── Rinomina l'analisi selezionata ───────────────────────────────────────────
@app.callback(
    Output('pio-exp-profile', 'options', allow_duplicate=True),
    Output('pio-imp-profile', 'options', allow_duplicate=True),
    Output('pio-imp-profile', 'value',   allow_duplicate=True),
    Output('pio-rename-input', 'value'),
    Output('pio-manage-status', 'children'),
    Input('pio-rename-btn', 'n_clicks'),
    State('pio-imp-profile', 'value'),
    State('pio-rename-input', 'value'),
    prevent_initial_call=True,
)
def pio_rename(n, old, new):
    if not n:
        raise PreventUpdate
    if not old:
        return no_update, no_update, no_update, no_update, "⚠ Seleziona prima un'analisi da rinominare"
    new = (new or '').strip()
    if not new:
        return no_update, no_update, no_update, no_update, '⚠ Scrivi il nuovo nome'
    _u = _get_username()
    ok = _sm.rename_analysis(_u, old, new)
    opts = [{'label': a, 'value': a} for a in _sm.list_analyses(_u)]
    if ok:
        return opts, opts, new, '', f'✏️ "{old}" rinominata in "{new}"'
    return no_update, no_update, no_update, no_update, '⚠ Rinomina non riuscita'


# ── Cancella l'analisi selezionata ───────────────────────────────────────────
@app.callback(
    Output('pio-exp-profile', 'options', allow_duplicate=True),
    Output('pio-imp-profile', 'options', allow_duplicate=True),
    Output('pio-imp-profile', 'value',   allow_duplicate=True),
    Output('pio-manage-status', 'children', allow_duplicate=True),
    Input('pio-del-btn', 'n_clicks'),
    State('pio-imp-profile', 'value'),
    prevent_initial_call=True,
)
def pio_delete(n, name):
    if not n:
        raise PreventUpdate
    if not name:
        return no_update, no_update, no_update, "⚠ Seleziona prima un'analisi da cancellare"
    _u = _get_username()
    ok = _sm.delete_analysis(_u, name)
    opts = [{'label': a, 'value': a} for a in _sm.list_analyses(_u)]
    if ok:
        return opts, opts, None, f'🗑 Analisi "{name}" cancellata'
    return no_update, no_update, no_update, '⚠ Cancellazione non riuscita'


# ── Cancella l'analisi selezionata DALLA VISTA ESPORTA ───────────────────────
@app.callback(
    Output('pio-exp-profile', 'options', allow_duplicate=True),
    Output('pio-imp-profile', 'options', allow_duplicate=True),
    Output('pio-exp-profile', 'value',   allow_duplicate=True),
    Output('pio-exp-status',  'children', allow_duplicate=True),
    Input('pio-exp-del-btn', 'n_clicks'),
    State('pio-exp-profile', 'value'),
    prevent_initial_call=True,
)
def pio_export_delete(n, name):
    if not n:
        raise PreventUpdate
    if not name:
        return no_update, no_update, no_update, "⚠ Seleziona prima un'analisi da cancellare nel menu sopra"
    _u = _get_username()
    ok = _sm.delete_analysis(_u, name)
    opts = [{'label': a, 'value': a} for a in _sm.list_analyses(_u)]
    if ok:
        return opts, opts, None, f'🗑 Analisi "{name}" cancellata'
    return no_update, no_update, no_update, '⚠ Cancellazione non riuscita'


# ── Azzera i pesi P1/P2/P3 (svuota i portafogli correnti) ─────────────────────
@app.callback(
    Output('weights-store-P1', 'data', allow_duplicate=True),
    Output('weights-store-P2', 'data', allow_duplicate=True),
    Output('weights-store-P3', 'data', allow_duplicate=True),
    Output({'type': 'weight-input', 'index': ALL}, 'value', allow_duplicate=True),
    Output('pio-manage-status', 'children', allow_duplicate=True),
    Output('pio-overlay', 'style', allow_duplicate=True),
    Input('pio-reset-cols-btn', 'n_clicks'),
    State({'type': 'weight-input', 'index': ALL}, 'id'),
    prevent_initial_call=True,
)
def pio_reset_cols(n, all_ids):
    if not n:
        raise PreventUpdate
    all_ids = all_ids or []
    new_vals = [0 for _ in all_ids]  # azzera tutte le celle P1/P2/P3
    _update_user_json(weights={'P1': {}, 'P2': {}, 'P3': {}}, username=_get_username())
    return ({}, {}, {}, new_vals, '🧹 P1/P2/P3 azzerati',
            {**_PIO_OVERLAY, 'display': 'none'})


# ─────────────────────────────────────────────────────────────────────────────
# Callback: scarica template ticker Excel
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('download-template', 'data'),
    Input('btn-download-template', 'n_clicks'),
    prevent_initial_call=True,
)
def download_template(n_clicks):
    # Template generato dal modulo condiviso data_core (uguale in Analisi Tattica)
    return dcc.send_bytes(lambda b: b.write(dc.template_bytes()), 'template_portafoglio.xlsx')


# ─────────────────────────────────────────────────────────────────────────────
# Callback: benchmark dropdown popolato dopo caricamento dati
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('benchmark-selector', 'options'),
    Output('benchmark-selector', 'value'),
    Input('asset-checklist', 'data'),
    State('benchmark-selector', 'value'),
)
def update_benchmark_options(options_tickers, current_value):
    if not options_tickers:
        return [], None
    portfolio_opts = [
        {'label': '── Portafoglio P1', 'value': 'Port1'},
        {'label': '── Portafoglio P2', 'value': 'Port2'},
        {'label': '── Portafoglio P3', 'value': 'Port3'},
    ]
    all_options = portfolio_opts + options_tickers
    if current_value and any(opt['value'] == current_value for opt in all_options):
        return all_options, current_value
    return all_options, options_tickers[0]['value'] if options_tickers else None


# ─────────────────────────────────────────────────────────────────────────────
# Callback: mostra/nascondi hint Update
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('update-hint', 'style'),
    Input('stock-data',              'data'),
    Input('update-portfolio-button', 'n_clicks'),
)
def toggle_update_hint(stock_data, update_clicks):
    _shown  = {'display': 'block', 'font-size': '9px', 'color': '#c0392b',
                'font-weight': '600', 'padding': '2px 5px 4px 5px',
                'background': '#fdf2f0', 'border-left': '3px solid #c0392b',
                'margin-bottom': '4px', 'border-radius': '0 4px 4px 0'}
    _hidden = {**_shown, 'display': 'none'}
    ctx = callback_context
    triggered = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''
    if triggered == 'update-portfolio-button' and update_clicks:
        return _hidden
    if triggered == 'stock-data' and stock_data:
        return _shown
    return _hidden


# ─────────────────────────────────────────────────────────────────────────────
# Callback: griglia pesi e asset
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('weights-grid-container', 'children'),
    Output('asset-count-display',    'children'),
    Input('update-portfolio-button', 'n_clicks'),
    State('stock-data',    'data'),
    State('asset-checklist', 'data'),
    State({'type': 'graph-select-checkbox',  'index': ALL}, 'value'),
    State({'type': 'ir-select-checkbox',     'index': ALL}, 'value'),
    State({'type': 'sharpe-select-checkbox', 'index': ALL}, 'value'),
    State({'type': 'tev-select-checkbox',    'index': ALL}, 'value'),
    State({'type': 'dd-select-checkbox',     'index': ALL}, 'value'),
    State({'type': 'vol-select-checkbox',    'index': ALL}, 'value'),
    State({'type': 'var90-select-checkbox',  'index': ALL}, 'value'),
    State({'type': 'var95-select-checkbox',  'index': ALL}, 'value'),
    State('weights-store-P1', 'data'),
    State('weights-store-P2', 'data'),
    State('weights-store-P3', 'data'),
    State('benchmark-selector', 'value'),
    State('ir-window-input',    'value'),
    State('ir-filter-radio',    'value'),
    prevent_initial_call=False,
)
def generate_asset_and_weight_inputs(update_clicks, stock_data_json, options_tickers,
                                      graph_vals, ir_vals, sharpe_vals, tev_vals,
                                      dd_vals, vol_vals, var90_vals, var95_vals,
                                      saved_p1, saved_p2, saved_p3,
                                      benchmark_value, ir_window, ir_filter):
    _placeholder = html.Div(
        'Carica i dati per visualizzare gli asset',
        style={'color': '#888', 'font-style': 'italic', 'font-size': '11px', 'padding': '12px 8px'}
    )
    # Se stock-data non è nello store, prova direttamente dal buffer in memoria
    if not stock_data_json:
        with _DL_LOCK:
            buf = dict(_DL_BUFFER)
        cr = buf.get('close_returns')
        if cr is not None and not cr.empty:
            stock_data_json = cr.to_json(date_format='iso', orient='split')
            if not options_tickers:
                options_tickers = [{'label': c, 'value': c} for c in cr.columns]
    if not stock_data_json or not options_tickers:
        return [_placeholder], ''

    # Ripristina selezioni precedenti
    saved_selected = []
    for grp in (graph_vals or []) + (ir_vals or []) + (sharpe_vals or []) + \
               (tev_vals or []) + (dd_vals or []) + (vol_vals or []) + \
               (var90_vals or []) + (var95_vals or []):
        saved_selected.extend(grp)

    asset_names = [option['value'] for option in options_tickers]
    saved_p1 = saved_p1 or {}
    saved_p2 = saved_p2 or {}
    saved_p3 = saved_p3 or {}

    # Ripristina dalla FONTE UNICA (current.json) in modo indipendente: la
    # selezione 📊 (checked) va riletta anche quando ci sono pesi, altrimenti
    # tornando da un'altra tab (es. Correlazioni) le spunte sparirebbero.
    _ns = None
    if not saved_selected:
        _ns = _read_user_json()
        saved_selected = [d for d, v in _ns.items() if v.get('checked')]
    if not saved_p1 and not saved_p2 and not saved_p3:
        if _ns is None:
            _ns = _read_user_json()
        saved_p1 = {d: v.get('P1', 0) for d, v in _ns.items()}
        saved_p2 = {d: v.get('P2', 0) for d, v in _ns.items()}
        saved_p3 = {d: v.get('P3', 0) for d, v in _ns.items()}

    # Calcola AKRatio per colorazione etichette
    assets_above_threshold = set()
    if ir_filter and ir_filter != 'all' and benchmark_value:
        df_all = _get_df(stock_data_json)
        window = ir_window if (ir_window and ir_window > 0) else 30
        if df_all is not None and benchmark_value in df_all.columns:
            threshold = 0.0 if ir_filter == 'gt_0' else -1.0
            benchmark_returns = df_all[benchmark_value].dropna()
            for asset in asset_names:
                if asset == benchmark_value or asset not in df_all.columns:
                    continue
                tail = pd.concat([df_all[asset].dropna(), benchmark_returns], axis=1).dropna().iloc[-window:]
                if len(tail) < window:
                    continue
                ir_val = calculate_rolling_information_ratio(
                    tail.iloc[:, 0], tail.iloc[:, 1], window=window).iloc[-1]
                if pd.notna(ir_val) and ir_val > threshold:
                    assets_above_threshold.add(asset)

    n_total = len(asset_names)
    n_above = len(assets_above_threshold)
    if ir_filter and ir_filter != 'all':
        label = '> 0' if ir_filter == 'gt_0' else '> −1'
        count_text = [
            html.Span(f'{n_above}', style={'font-weight': 'bold', 'color': '#c0392b', 'font-size': '12px'}),
            html.Span(f' / {n_total} asset', style={'color': '#555'}),
            html.Span(f'  in rosso (AKRatio {label})', style={'color': '#c0392b', 'font-style': 'italic'}),
        ]
    else:
        count_text = [
            html.Span(f'{n_total}', style={'font-weight': 'bold', 'color': '#1a3a5c', 'font-size': '12px'}),
            html.Span(f' / {n_total} asset', style={'color': '#555'}),
        ]

    rows = []

    def _chk(chk_id, opt_val, sel_val, w):
        return html.Div(
            dcc.Checklist(id=chk_id,
                          options=[{'label': '', 'value': opt_val}],
                          value=sel_val,
                          inputStyle={'width': '10px', 'height': '10px',
                                      'cursor': 'pointer', 'margin': '0'},
                          style={'display': 'flex', 'justify-content': 'center',
                                 'align-items': 'center', 'width': '100%'},
                          className='asset-checkbox'),
            style={'width': w, 'height': '22px', 'display': 'flex',
                   'align-items': 'center', 'justify-content': 'center'})

    for asset in asset_names:
        def create_weight_input(portfolio_index, a=asset):
            port_key  = f'P{portfolio_index}'
            saved_val = {1: saved_p1, 2: saved_p2, 3: saved_p3}[portfolio_index].get(a, 0)
            return dcc.Input(
                id={'type': 'weight-input', 'index': f'{port_key}-{a}'},
                type='number', value=saved_val, min=0, max=100, step=0.1, placeholder='0',
                style={'width': '96%', 'text-align': 'right', 'font-size': '11px',
                       'height': '20px', 'padding': '1px 3px', 'box-sizing': 'border-box'}
            )

        asset_val   = [asset]                       if asset in saved_selected else []
        ir_val      = [f'{asset}_InformationRatio'] if f'{asset}_InformationRatio' in saved_selected else []
        sharpe_val  = [f'{asset}_Sharpe']           if f'{asset}_Sharpe'           in saved_selected else []
        tev_val     = [f'{asset}_TEV']              if f'{asset}_TEV'              in saved_selected else []
        dd_val      = [f'{asset}_DD']               if f'{asset}_DD'               in saved_selected else []
        vol_val     = [f'{asset}_Vol']              if f'{asset}_Vol'              in saved_selected else []
        var90_val   = [f'{asset}_VaR90']            if f'{asset}_VaR90'            in saved_selected else []
        var95_val   = [f'{asset}_VaR95']            if f'{asset}_VaR95'            in saved_selected else []
        _label_color = '#c0392b' if asset in assets_above_threshold else '#1a3a5c'

        row_content = html.Div([
            html.Div(
                html.Span(asset, style={
                    'color': _label_color, 'fontWeight': 'bold',
                    'overflow': 'hidden', 'whiteSpace': 'nowrap',
                    'textOverflow': 'ellipsis', 'maxWidth': '100%', 'fontSize': '9px',
                }),
                **{'data-tooltip': asset, 'data-tooltip-color': _label_color},
                style={'width': '19%', 'height': '28px', 'display': 'flex',
                       'alignItems': 'center', 'paddingLeft': '4px',
                       'overflow': 'hidden', 'position': 'relative', 'cursor': 'default'},
            ),
            _chk({'type': 'graph-select-checkbox',  'index': asset}, asset,                       asset_val,  '3%'),
            html.Div(create_weight_input(1), className='weight-input-cell', style={'width': '12%'}),
            html.Div(create_weight_input(2), className='weight-input-cell', style={'width': '12%'}),
            html.Div(create_weight_input(3), className='weight-input-cell', style={'width': '12%'}),
            _chk({'type': 'ir-select-checkbox',     'index': asset}, f'{asset}_InformationRatio', ir_val,     '5%'),
            _chk({'type': 'sharpe-select-checkbox', 'index': asset}, f'{asset}_Sharpe',           sharpe_val, '5%'),
            _chk({'type': 'tev-select-checkbox',    'index': asset}, f'{asset}_TEV',              tev_val,    '5%'),
            _chk({'type': 'dd-select-checkbox',     'index': asset}, f'{asset}_DD',               dd_val,     '5%'),
            _chk({'type': 'vol-select-checkbox',    'index': asset}, f'{asset}_Vol',              vol_val,    '5%'),
            _chk({'type': 'var90-select-checkbox',  'index': asset}, f'{asset}_VaR90',            var90_val,  '5%'),
            _chk({'type': 'var95-select-checkbox',  'index': asset}, f'{asset}_VaR95',            var95_val,  '5%'),
            html.Button('×', id={'type': 'delete-asset-btn', 'index': asset}, n_clicks=0,
                        style={'fontSize': '11px', 'width': '16px', 'height': '16px',
                               'background': 'none', 'border': '1px solid #ddd',
                               'color': '#bbb', 'cursor': 'pointer', 'borderRadius': '2px',
                               'padding': '0', 'flexShrink': '0', 'lineHeight': '14px',
                               'marginLeft': '2px'}),
        ], style={'display': 'flex', 'align-items': 'center', 'border-bottom': '1px dotted #eee'})
        rows.append(row_content)

    # Righe portafogli P1/P2/P3
    def _pchk(chk_id, opt_val, sel_val, w):
        return html.Div(
            dcc.Checklist(id=chk_id,
                          options=[{'label': '', 'value': opt_val}],
                          value=sel_val,
                          inputStyle={'width': '10px', 'height': '10px',
                                      'cursor': 'pointer', 'margin': '0'},
                          style={'display': 'flex', 'justify-content': 'center',
                                 'align-items': 'center', 'width': '100%'}),
            style={'width': w, 'height': '22px', 'display': 'flex',
                   'align-items': 'center', 'justify-content': 'center'})

    for portfolio_num in [1, 2, 3]:
        portfolio_name  = f'Port{portfolio_num}'
        port_val        = [portfolio_name]                              if portfolio_name                              in saved_selected else []
        ir_port_val     = [f'{portfolio_name}_InformationRatio']       if f'{portfolio_name}_InformationRatio'       in saved_selected else []
        sharpe_port_val = [f'{portfolio_name}_Sharpe']                 if f'{portfolio_name}_Sharpe'                 in saved_selected else []
        tev_port_val    = [f'{portfolio_name}_TEV']                    if f'{portfolio_name}_TEV'                    in saved_selected else []
        dd_port_val     = [f'{portfolio_name}_DD']                     if f'{portfolio_name}_DD'                     in saved_selected else []
        vol_port_val    = [f'{portfolio_name}_Vol']                    if f'{portfolio_name}_Vol'                    in saved_selected else []
        var90_port_val  = [f'{portfolio_name}_VaR90']                  if f'{portfolio_name}_VaR90'                  in saved_selected else []
        var95_port_val  = [f'{portfolio_name}_VaR95']                  if f'{portfolio_name}_VaR95'                  in saved_selected else []

        portfolio_row = html.Div([
            html.Div(
                html.Span(portfolio_name, style={
                    'color': '#0066cc', 'fontWeight': 'bold',
                    'overflow': 'hidden', 'whiteSpace': 'nowrap',
                    'textOverflow': 'ellipsis', 'maxWidth': '100%', 'fontSize': '9px',
                }),
                **{'data-tooltip': portfolio_name, 'data-tooltip-color': '#0066cc'},
                style={'width': '19%', 'height': '28px', 'display': 'flex',
                       'alignItems': 'center', 'paddingLeft': '4px',
                       'overflow': 'hidden', 'position': 'relative', 'cursor': 'default'}),
            _pchk({'type': 'graph-select-checkbox',  'index': portfolio_name}, portfolio_name,                       port_val,        '3%'),
            html.Div('', style={'width': '12%'}),
            html.Div('', style={'width': '12%'}),
            html.Div('', style={'width': '12%'}),
            _pchk({'type': 'ir-select-checkbox',     'index': portfolio_name}, f'{portfolio_name}_InformationRatio', ir_port_val,     '5%'),
            _pchk({'type': 'sharpe-select-checkbox', 'index': portfolio_name}, f'{portfolio_name}_Sharpe',           sharpe_port_val, '5%'),
            _pchk({'type': 'tev-select-checkbox',    'index': portfolio_name}, f'{portfolio_name}_TEV',              tev_port_val,    '5%'),
            _pchk({'type': 'dd-select-checkbox',     'index': portfolio_name}, f'{portfolio_name}_DD',               dd_port_val,     '5%'),
            _pchk({'type': 'vol-select-checkbox',    'index': portfolio_name}, f'{portfolio_name}_Vol',              vol_port_val,    '5%'),
            _pchk({'type': 'var90-select-checkbox',  'index': portfolio_name}, f'{portfolio_name}_VaR90',            var90_port_val,  '5%'),
            _pchk({'type': 'var95-select-checkbox',  'index': portfolio_name}, f'{portfolio_name}_VaR95',            var95_port_val,  '5%'),
        ], style={'display': 'flex', 'align-items': 'center', 'border-bottom': '1px dotted #eee',
                  'background-color': '#f0f0f0'})
        rows.append(portfolio_row)

    return rows, count_text


# ─────────────────────────────────────────────────────────────────────────────
# Callback: raccoglie asset selezionati nello store globale
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('global-assets-selected', 'data'),
    Input({'type': 'graph-select-checkbox', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def collect_selected_assets(all_values):
    selected = [v[0] for v in all_values if v]
    _update_user_json(checked=selected)
    return selected


# ─────────────────────────────────────────────────────────────────────────────
# Callbacks: pulsanti cancella colonna
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output({'type': 'graph-select-checkbox', 'index': ALL}, 'value'),
    Input('deselect-all-tickers', 'n_clicks'),
    State({'type': 'graph-select-checkbox', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def deselect_all_tickers(n, vals):
    if not n: raise PreventUpdate
    return [[] for _ in vals]

@app.callback(
    Output({'type': 'ir-select-checkbox', 'index': ALL}, 'value'),
    Input('deselect-all-ir', 'n_clicks'),
    State({'type': 'ir-select-checkbox', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def deselect_all_ir(n, vals):
    if not n: raise PreventUpdate
    return [[] for _ in vals]

@app.callback(
    Output({'type': 'sharpe-select-checkbox', 'index': ALL}, 'value'),
    Input('deselect-all-sharpe', 'n_clicks'),
    State({'type': 'sharpe-select-checkbox', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def deselect_all_sharpe(n, vals):
    if not n: raise PreventUpdate
    return [[] for _ in vals]

@app.callback(
    Output({'type': 'tev-select-checkbox', 'index': ALL}, 'value'),
    Input('deselect-all-tev', 'n_clicks'),
    State({'type': 'tev-select-checkbox', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def deselect_all_tev(n, vals):
    if not n: raise PreventUpdate
    return [[] for _ in vals]

@app.callback(
    Output({'type': 'dd-select-checkbox', 'index': ALL}, 'value'),
    Input('deselect-all-dd', 'n_clicks'),
    State({'type': 'dd-select-checkbox', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def deselect_all_dd(n, vals):
    if not n: raise PreventUpdate
    return [[] for _ in vals]

@app.callback(
    Output({'type': 'vol-select-checkbox', 'index': ALL}, 'value'),
    Input('deselect-all-vol', 'n_clicks'),
    State({'type': 'vol-select-checkbox', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def deselect_all_vol(n, vals):
    if not n: raise PreventUpdate
    return [[] for _ in vals]

@app.callback(
    Output({'type': 'var90-select-checkbox', 'index': ALL}, 'value'),
    Input('deselect-all-var90', 'n_clicks'),
    State({'type': 'var90-select-checkbox', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def deselect_all_var90(n, vals):
    if not n: raise PreventUpdate
    return [[] for _ in vals]

@app.callback(
    Output({'type': 'var95-select-checkbox', 'index': ALL}, 'value'),
    Input('deselect-all-var95', 'n_clicks'),
    State({'type': 'var95-select-checkbox', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def deselect_all_var95(n, vals):
    if not n: raise PreventUpdate
    return [[] for _ in vals]

@app.callback(
    Output({'type': 'weight-input', 'index': ALL}, 'value'),
    Input('reset-p1-tab1', 'n_clicks'),
    Input('reset-p2-tab1', 'n_clicks'),
    Input('reset-p3-tab1', 'n_clicks'),
    State({'type': 'weight-input', 'index': ALL}, 'id'),
    State({'type': 'weight-input', 'index': ALL}, 'value'),
    prevent_initial_call=True,
)
def reset_portfolio_weights(n1, n2, n3, all_ids, all_vals):
    ctx = callback_context
    triggered = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''
    prefix = {'reset-p1-tab1': 'P1-', 'reset-p2-tab1': 'P2-', 'reset-p3-tab1': 'P3-'}.get(triggered)
    if not prefix:
        raise PreventUpdate
    return [0 if inp_id['index'].startswith(prefix) else val
            for inp_id, val in zip(all_ids, all_vals)]


# ─────────────────────────────────────────────────────────────────────────────
# Callback: somme pesi
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('sum-p1-display', 'children'),
    Output('sum-p2-display', 'children'),
    Output('sum-p3-display', 'children'),
    Output('sum-p1-display', 'style'),
    Output('sum-p2-display', 'style'),
    Output('sum-p3-display', 'style'),
    Output('weights-store-P1', 'data'),
    Output('weights-store-P2', 'data'),
    Output('weights-store-P3', 'data'),
    Input({'type': 'weight-input', 'index': ALL}, 'value'),
    State({'type': 'weight-input', 'index': ALL}, 'id'),
    State('weights-store-P1', 'data'),
    State('weights-store-P2', 'data'),
    State('weights-store-P3', 'data'),
    prevent_initial_call=True,
)
def update_portfolio_weights(all_input_values, all_input_ids, p1_data, p2_data, p3_data):
    p1 = dict(p1_data or {})
    p2 = dict(p2_data or {})
    p3 = dict(p3_data or {})

    for val, inp_id in zip(all_input_values, all_input_ids):
        idx = inp_id['index']
        if idx.startswith('P1-'):
            p1[idx[3:]] = val or 0
        elif idx.startswith('P2-'):
            p2[idx[3:]] = val or 0
        elif idx.startswith('P3-'):
            p3[idx[3:]] = val or 0

    sum1 = sum(v for v in p1.values() if v)
    sum2 = sum(v for v in p2.values() if v)
    sum3 = sum(v for v in p3.values() if v)

    def _style(s):
        base = {'width': '12%', 'text-align': 'right', 'padding-right': '5px',
                'font-size': '10px', 'font-weight': 'bold', 'box-sizing': 'border-box'}
        base['color'] = '#1b7a34' if abs(s - 100) < 0.01 else '#d62728'
        return base

    _update_user_json(weights={'P1': p1, 'P2': p2, 'P3': p3})
    return (f'{sum1:.1f}%', f'{sum2:.1f}%', f'{sum3:.1f}%',
            _style(sum1), _style(sum2), _style(sum3),
            p1, p2, p3)


# ─────────────────────────────────────────────────────────────────────────────
# Callback: seleziona colonna dal click sul grafico
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('selected-column', 'value'),
    Input('update-portfolio-button', 'n_clicks'),
    Input('delete-column-button',    'n_clicks'),
    Input('controls-and-graph',      'clickData'),
    State('controls-and-graph',      'figure'),
    prevent_initial_call=True,
)
def update_selected_column(update_clicks, delete_clicks, clickData, figure):
    ctx = callback_context
    triggered_id = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''
    if triggered_id in ('update-portfolio-button', 'delete-column-button'):
        return ''
    if triggered_id == 'controls-and-graph' and clickData and clickData['points']:
        curve_number = clickData['points'][0]['curveNumber']
        if figure and 'data' in figure and curve_number < len(figure['data']):
            return figure['data'][curve_number].get('name', '')
    return ''


# ─────────────────────────────────────────────────────────────────────────────
# Callback: date picker → tab1-slider-store (grafico reagisce al cambio date)
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('dr-start-tab1',     'date'),
    Output('dr-end-tab1',       'date'),
    Output('dr-label-tab1',     'children'),
    Output('tab1-slider-store', 'data'),
    Input('stock-data',         'data'),
    Input('dr-start-tab1',      'date'),
    Input('dr-end-tab1',        'date'),
    prevent_initial_call=False,
)
def sync_date_range(stock_data, start_date, end_date):
    ctx = callback_context
    triggered_id = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''

    if not stock_data:
        raise PreventUpdate

    df = _get_df(stock_data)
    if df is None or df.empty:
        raise PreventUpdate

    ts_min = int(df.index.min().timestamp())
    ts_max = int(df.index.max().timestamp())

    if triggered_id == 'stock-data':
        # Inizializzazione: 10 anni fa → oggi, clampati al range dei dati
        today         = pd.Timestamp.today().normalize()
        ten_years_ago = today - pd.DateOffset(years=10)
        s = max(ts_min, int(ten_years_ago.timestamp()))
        e = min(ts_max, int(today.timestamp()))
        d_start = pd.Timestamp(s, unit='s').strftime('%Y-%m-%d')
        d_end   = pd.Timestamp(e, unit='s').strftime('%Y-%m-%d')
        label   = (f"{pd.Timestamp(s, unit='s').strftime('%d/%m/%Y')} — "
                   f"{pd.Timestamp(e, unit='s').strftime('%d/%m/%Y')}")
        return d_start, d_end, label, [s, e]

    # L'utente ha cambiato una delle due date: non riscrivere i picker (evita loop)
    try:
        s = int(pd.Timestamp(start_date).timestamp()) if start_date else ts_min
        e = int(pd.Timestamp(end_date).timestamp())   if end_date   else ts_max
    except Exception:
        s, e = ts_min, ts_max

    s = max(ts_min, min(s, ts_max))
    e = max(ts_min, min(e, ts_max))

    label = (f"{pd.Timestamp(s, unit='s').strftime('%d/%m/%Y')} — "
             f"{pd.Timestamp(e, unit='s').strftime('%d/%m/%Y')}")

    return no_update, no_update, label, [s, e]


# ─────────────────────────────────────────────────────────────────────────────
# Callback: aggiorna grafico principale
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('controls-and-graph',        'figure'),
    Output('date-values',               'children'),
    Output('insufficient-data-tickers', 'value'),

    Input('update-portfolio-button', 'n_clicks'),
    Input('delete-column-button',    'n_clicks'),
    Input('controls-and-graph',      'clickData'),

    State('dr-start-tab1',           'date'),
    State('dr-end-tab1',             'date'),
    State('tab1-slider-store',       'data'),
    State('global-assets-selected',  'data'),
    State('benchmark-selector',      'value'),
    State('ir-window-input',         'value'),
    State('ak-ma-input',             'value'),
    State('weights-store-P1',        'data'),
    State('weights-store-P2',        'data'),
    State('weights-store-P3',        'data'),
    State({'type': 'ir-select-checkbox',     'index': ALL}, 'value'),
    State({'type': 'sharpe-select-checkbox', 'index': ALL}, 'value'),
    State({'type': 'tev-select-checkbox',    'index': ALL}, 'value'),
    State({'type': 'dd-select-checkbox',     'index': ALL}, 'value'),
    State({'type': 'vol-select-checkbox',    'index': ALL}, 'value'),
    State({'type': 'var90-select-checkbox',  'index': ALL}, 'value'),
    State({'type': 'var95-select-checkbox',  'index': ALL}, 'value'),
    State('vol-window-input',        'value'),
    State('insufficient-data-store', 'data'),
    State('selected-column',         'value'),
    State('stock-data',              'data'),
    prevent_initial_call=True,
)
def update_graph(update_clicks, delete_clicks, clickData, picker_start, picker_end, date_range, selected_assets,
                 benchmark_value, ir_window, ak_ma_window,
                 weights_p1_data, weights_p2_data, weights_p3_data,
                 all_ir_checkbox_values, all_sharpe_checkbox_values, all_tev_checkbox_values,
                 all_dd_checkbox_values, all_vol_checkbox_values,
                 all_var90_checkbox_values, all_var95_checkbox_values, vol_window,
                 insufficient_data, selected_column, stock_data):

    ctx = callback_context
    if not ctx.triggered:
        raise PreventUpdate

    triggered_id = ctx.triggered[0]['prop_id'].split('.')[0]

    if not update_clicks and not delete_clicks and not clickData:
        raise PreventUpdate

    if not stock_data:
        with _DL_LOCK:
            buf = dict(_DL_BUFFER)
        cr = buf.get('close_returns')
        if cr is not None and not cr.empty:
            stock_data = cr.to_json(date_format='iso', orient='split')
        else:
            return {}, 'Nessun dato caricato — clicca ⟳ Aggiorna', ''

    def _collect(vals):
        result = []
        for v in (vals or []):
            if v:
                result.extend(v)
        return result

    selected_irs    = _collect(all_ir_checkbox_values)
    selected_sharpe = _collect(all_sharpe_checkbox_values)
    selected_tev    = _collect(all_tev_checkbox_values)
    selected_dd     = _collect(all_dd_checkbox_values)
    selected_vol    = _collect(all_vol_checkbox_values)
    selected_var90  = _collect(all_var90_checkbox_values)
    selected_var95  = _collect(all_var95_checkbox_values)

    vol_window = vol_window if (vol_window and vol_window > 0) else 30
    ir_window  = ir_window  if (ir_window  and ir_window  > 0) else 30

    close_returns = _get_df(stock_data)

    # Priorità ai valori diretti dei DatePicker (evita race condition con tab1-slider-store)
    if picker_start and picker_end:
        start_date  = pd.to_datetime(picker_start)
        end_date    = pd.to_datetime(picker_end)
        filtered_df = close_returns.loc[start_date:end_date]
    elif date_range and len(date_range) == 2:
        start_date  = pd.to_datetime(date_range[0], unit='s')
        end_date    = pd.to_datetime(date_range[1], unit='s')
        filtered_df = close_returns.loc[start_date:end_date]
    else:
        filtered_df = close_returns
        start_date  = close_returns.index.min()
        end_date    = close_returns.index.max()

    asset_columns      = list(filtered_df.columns)
    df_with_portfolios = filtered_df.copy()

    weights_data_map = {
        'Port1': weights_p1_data,
        'Port2': weights_p2_data,
        'Port3': weights_p3_data,
    }

    def calc_portfolio(df, name, weights_dict, asset_cols):
        weights_dict = weights_dict or {}
        total_w = sum(weights_dict.values())
        normalized = {}
        if total_w > 0:
            for a, w in weights_dict.items():
                if a in asset_cols and w and w > 0:
                    normalized[a] = w / 100.0
        port = pd.Series(0.0, index=df.index, name=name)
        if normalized:
            for a, w in normalized.items():
                if a in df.columns:
                    port += df[a] * w
        return port

    for p_name, w_dict in weights_data_map.items():
        df_with_portfolios[p_name] = calc_portfolio(
            df_with_portfolios, p_name, w_dict, asset_columns)

    if benchmark_value is None or benchmark_value not in df_with_portfolios.columns:
        benchmark_col = df_with_portfolios.columns[0] if not df_with_portfolios.empty else None
    else:
        benchmark_col = benchmark_value

    if benchmark_col is None:
        return {}, 'Nessun dato caricato', ''

    benchmark_returns = df_with_portfolios[benchmark_col]
    benchmark_cumulative_returns = (1 + benchmark_returns).cumprod() - 1

    # Calcola IR/Sharpe/TEV solo per gli asset effettivamente selezionati
    _need_ir     = {s.replace('_InformationRatio', '') for s in selected_irs    if s}
    _need_sharpe = {si.replace('_Sharpe', '')          for si in selected_sharpe if si}
    _need_tev    = {ti.replace('_TEV', '')             for ti in selected_tev    if ti}

    _extra = {}
    for col in _need_ir:
        if col in df_with_portfolios.columns and col != benchmark_col:
            _extra[f'{col}_InformationRatio'] = calculate_rolling_information_ratio(
                df_with_portfolios[col], benchmark_returns, window=ir_window)
    for col in _need_sharpe:
        if col in df_with_portfolios.columns:
            _extra[f'{col}_SharpeRatio'] = calculate_rolling_sharpe_ratio(
                df_with_portfolios[col], window=ir_window)
    for col in _need_tev:
        if col in df_with_portfolios.columns and col != benchmark_col:
            _extra[f'{col}_TEV'] = calculate_tracking_error_volatility(
                df_with_portfolios[col], benchmark_returns, window=ir_window)

    if _extra:
        df_final = pd.concat([df_with_portfolios,
                               pd.DataFrame(_extra, index=df_with_portfolios.index)], axis=1)
    else:
        df_final = df_with_portfolios

    # Guardia: nessun dato nell'intervallo → grafico vuoto invece di errore
    if df_final is None or len(df_final.index) == 0:
        empty = go.Figure()
        empty.add_annotation(text='Nessun dato nell\'intervallo selezionato',
                             xref='paper', yref='paper', x=0.5, y=0.5,
                             showarrow=False, font=dict(size=13, color='#888'))
        empty.update_layout(paper_bgcolor='white', plot_bgcolor='#f8faff',
                            margin=dict(t=30, b=20, l=40, r=20))
        return empty

    # Versione ridotta per il rendering (max 500 punti per trace)
    df_plot = _thin(df_final)
    benchmark_cumulative_returns = _thin(benchmark_cumulative_returns)

    fig = make_subplots(
        rows=7, cols=1, shared_xaxes=False, vertical_spacing=0.03,
        row_heights=[0.26, 0.12, 0.12, 0.12, 0.12, 0.13, 0.13],
        specs=[[{"secondary_y": False}]] * 7,
        subplot_titles=('Cumulative Returns', 'AKRatio', 'Sharpe Ratio',
                        'TEV', 'DrawDown', 'Volatilità', 'VaR (90% / 95%)')
    )

    for row in range(1, 8):
        fig.add_trace(
            go.Scatter(x=[df_final.index[0], df_final.index[-1]], y=[0, 0],
                       mode='lines', line=dict(color='rgba(0,0,0,0)'),
                       showlegend=False, hoverinfo='skip'),
            row=row, col=1)

    selected_assets = selected_assets or []

    # Subplot 1: Cumulative Returns
    if benchmark_col in selected_assets:
        bench_name = f'{benchmark_col} Cum. Returns'
        is_sel = (selected_column == bench_name)
        fig.add_trace(go.Scatter(x=benchmark_cumulative_returns.index,
                                  y=benchmark_cumulative_returns,
                                  name=bench_name,
                                  line=dict(color='red', width=8 if is_sel else 2),
                                  legend='legend'), row=1, col=1)
        color_index = 1
    else:
        color_index = 0

    for col in selected_assets:
        if col in df_with_portfolios.columns and col != benchmark_col:
            series = df_with_portfolios[col]
            if col.startswith('Port'):
                pnum    = int(col[-1])
                w_dict  = {1: weights_p1_data, 2: weights_p2_data, 3: weights_p3_data}.get(pnum, {}) or {}
                active  = [a for a, w in w_dict.items() if w and w > 0 and a in filtered_df.columns]
                if active:
                    first_valid = filtered_df[active].dropna(how='any').index.min()
                    if pd.notna(first_valid):
                        series = series.loc[first_valid:]
            cum_ret    = _thin((1 + series).cumprod() - 1)
            trace_name = f'{col} Cum. Returns'
            is_sel     = (selected_column == trace_name)
            if is_sel:
                line_dict = dict(color='red', width=8)
            else:
                tc = color_palette[color_index % len(color_palette)]
                line_dict = dict(color=tc)
                if col.startswith('Port'):
                    line_dict.update({'width': 4, 'dash': 'solid'})
            fig.add_trace(go.Scatter(x=cum_ret.index, y=cum_ret, name=trace_name,
                                      line=line_dict, legend='legend'), row=1, col=1)
            color_index += 1

    # Mappa colore per asset (evita scan O(n²) in ogni subplot successivo)
    _color_map = {}
    for _t in fig.data:
        _nm = _t.name or ''
        if ' Cum. Returns' in _nm:
            _asset = _nm.replace(' Cum. Returns', '')
            try:
                _c = _t.line.color
                if _c:
                    _color_map[_asset] = _c
            except AttributeError:
                pass

    # Subplot 2: AKRatio
    _ak_ma = int(ak_ma_window) if ak_ma_window and int(ak_ma_window) > 1 else 1
    color_index = 0
    for col_ir in selected_irs:
        if col_ir in df_plot.columns:
            orig = col_ir.replace('_InformationRatio', '')
            is_sel = (selected_column == col_ir)
            if is_sel:
                line_dict = dict(color='red', width=8)
            else:
                tc = _color_map.get(orig, color_palette[color_index % len(color_palette)])
                line_dict = dict(color=tc, width=4 if orig.startswith('Port') else 2.5)
            y_vals = df_plot[col_ir].rolling(_ak_ma, min_periods=1).mean() if _ak_ma > 1 else df_plot[col_ir]
            fig.add_trace(go.Scatter(x=df_plot.index, y=y_vals,
                                      name=col_ir.replace('_InformationRatio', '_AKRatio'),
                                      line=line_dict, legend='legend2'), row=2, col=1)
            color_index += 1
    if selected_irs:
        fig.add_trace(go.Scatter(x=df_plot.index, y=[0]*len(df_plot), name='Zero Line (IR)',
                                  line=dict(color='red', dash='solid', width=2),
                                  showlegend=True, legend='legend2'), row=2, col=1)

    # Subplot 3: Sharpe
    for si in selected_sharpe:
        if si:
            an = si.replace('_Sharpe', '')
            sc = f'{an}_SharpeRatio'
            if sc in df_plot.columns:
                tn    = f'{an} Sharpe Ratio'
                is_sel = (selected_column == tn)
                if is_sel:
                    line_dict = dict(color='red', width=8)
                else:
                    tc = _color_map.get(an, color_palette[color_index % len(color_palette)])
                    line_dict = dict(color=tc, dash='solid',
                                     width=4 if an.startswith('Port') else 2.5)
                y_vals = df_plot[sc].rolling(_ak_ma, min_periods=1).mean() if _ak_ma > 1 else df_plot[sc]
                fig.add_trace(go.Scatter(x=df_plot.index, y=y_vals, name=tn,
                                          line=line_dict, legend='legend3'), row=3, col=1)

    # Subplot 4: TEV
    for ti in selected_tev:
        if ti:
            an = ti.replace('_TEV', '')
            tc_nm = f'{an}_TEV'
            if tc_nm in df_plot.columns:
                tn    = f'{an} TEV'
                is_sel = (selected_column == tn)
                if is_sel:
                    line_dict = dict(color='red', width=8)
                else:
                    tc = _color_map.get(an, color_palette[color_index % len(color_palette)])
                    line_dict = dict(color=tc, dash='dash',
                                     width=4 if an.startswith('Port') else 2.5)
                y_vals = df_plot[tc_nm].rolling(_ak_ma, min_periods=1).mean() if _ak_ma > 1 else df_plot[tc_nm]
                fig.add_trace(go.Scatter(x=df_plot.index, y=y_vals, name=tn,
                                          line=line_dict, legend='legend4'), row=4, col=1)

    # Subplot 5: DrawDown
    for di in selected_dd:
        if di:
            an = di.replace('_DD', '')
            if an in df_with_portfolios.columns:
                dds   = _thin(calculate_drawdown(df_with_portfolios[an]))
                tn    = f'{an} DrawDown'
                is_sel = (selected_column == tn)
                if is_sel:
                    line_dict = dict(color='red', width=8)
                    fillcolor = 'rgba(200,0,0,0.08)'
                else:
                    tc = _color_map.get(an, color_palette[color_index % len(color_palette)])
                    line_dict = dict(color=tc, dash='dot',
                                     width=4 if an.startswith('Port') else 2.5)
                    fillcolor = (tc.replace('rgb', 'rgba').replace(')', ',0.10)')
                                 if tc and tc.startswith('rgb') else 'rgba(200,0,0,0.08)')
                fig.add_trace(go.Scatter(x=dds.index, y=dds, name=tn,
                                          line=line_dict, legend='legend5',
                                          fill='tozeroy', fillcolor=fillcolor), row=5, col=1)

    # Subplot 6: Volatilità
    for vi in selected_vol:
        if vi:
            an = vi.replace('_Vol', '')
            if an in df_with_portfolios.columns:
                vs    = _thin(_rolling_volatility(df_with_portfolios[an], vol_window))
                tn    = f'{an} Volatilità'
                is_sel = (selected_column == tn)
                if is_sel:
                    line_dict = dict(color='red', width=8)
                else:
                    tc = _color_map.get(an, color_palette[color_index % len(color_palette)])
                    line_dict = dict(color=tc, width=4 if an.startswith('Port') else 2.5)
                fig.add_trace(go.Scatter(x=vs.index, y=vs, name=tn,
                                          line=line_dict, legend='legend6'), row=6, col=1)

    # Subplot 7: VaR
    for v90 in selected_var90:
        if v90:
            an = v90.replace('_VaR90', '')
            if an in df_with_portfolios.columns:
                vs    = _thin(calculate_historical_cvar(df_with_portfolios[an], vol_window, 0.10))
                tn    = f'{an} VaR90'
                is_sel = (selected_column == tn)
                if is_sel:
                    line_dict = dict(color='red', width=8)
                else:
                    tc = _color_map.get(an, color_palette[color_index % len(color_palette)])
                    line_dict = dict(color=tc, width=4 if an.startswith('Port') else 2.5,
                                     dash='solid')
                fig.add_trace(go.Scatter(x=vs.index, y=vs, name=tn,
                                          line=line_dict, legend='legend7'), row=7, col=1)

    for v95 in selected_var95:
        if v95:
            an = v95.replace('_VaR95', '')
            if an in df_with_portfolios.columns:
                vs    = _thin(calculate_historical_cvar(df_with_portfolios[an], vol_window, 0.05))
                tn    = f'{an} VaR95'
                is_sel = (selected_column == tn)
                if is_sel:
                    line_dict = dict(color='red', width=8)
                else:
                    tc = _color_map.get(an, color_palette[color_index % len(color_palette)])
                    line_dict = dict(color=tc, width=4 if an.startswith('Port') else 2.5,
                                     dash='dot')
                fig.add_trace(go.Scatter(x=vs.index, y=vs, name=tn,
                                          line=line_dict, legend='legend7'), row=7, col=1)

    # ── Tooltip personalizzato per ogni linea ────────────────────────────────
    # Riquadro con bordo del colore della linea, sfondo bianco, testo nero.
    # Riga 1: data + valore  |  Riga 2: nome serie per esteso.
    for _tr in fig.data:
        try:
            _lc = _tr.line.color
        except AttributeError:
            continue
        if not _lc or _lc == 'rgba(0,0,0,0)':
            continue
        _nm = _tr.name or ''
        if not _nm:
            continue
        _fmt = '.1%' if any(k in _nm for k in ('Cum. Returns', 'DrawDown', 'VaR')) else '.4f'
        _tr.update(
            hovertemplate=(
                f'%{{x|%d/%m/%Y}}   %{{y:{_fmt}}}<br>'
                f'{_nm}'
                '<extra></extra>'
            ),
            hoverlabel=dict(
                bgcolor='white',
                bordercolor=_lc,
                font=dict(color='black', size=11),
                namelength=0,
            ),
        )

    for row in range(1, 8):
        fig.update_xaxes(title_text='', row=row, col=1)
    fig.update_yaxes(title_text='Cumulative Returns', row=1, col=1)
    fig.update_yaxes(title_text='AKRatio',            row=2, col=1)
    fig.update_yaxes(title_text='Sharpe Ratio',       row=3, col=1)
    fig.update_yaxes(title_text='TEV',                row=4, col=1)
    fig.update_yaxes(title_text='DrawDown',           row=5, col=1)
    fig.update_yaxes(title_text='Volatilità (ann.)',  row=6, col=1)
    fig.update_yaxes(title_text='VaR',                row=7, col=1)

    fig.update_layout(
        uirevision='graph',
        hovermode='closest',
        height=1900, showlegend=True,
        legend=dict(title=dict(text='<b>Asset</b>', font=dict(size=11)),
                    orientation='v', yanchor='top', y=1.0, xanchor='left', x=1.01,
                    bgcolor='rgba(255,255,255,0.85)', bordercolor='#aed6f1', borderwidth=1),
        legend2=dict(title=dict(text='<b>AKRatio</b>', font=dict(size=11)),
                     orientation='v', yanchor='top', y=0.72, xanchor='left', x=1.01,
                     bgcolor='rgba(255,255,255,0.85)', bordercolor='#aed6f1', borderwidth=1),
        legend3=dict(title=dict(text='<b>Sharpe Ratio</b>', font=dict(size=11)),
                     orientation='v', yanchor='top', y=0.58, xanchor='left', x=1.01,
                     bgcolor='rgba(255,255,255,0.85)', bordercolor='#aed6f1', borderwidth=1),
        legend4=dict(title=dict(text='<b>TEV</b>', font=dict(size=11)),
                     orientation='v', yanchor='top', y=0.44, xanchor='left', x=1.01,
                     bgcolor='rgba(255,255,255,0.85)', bordercolor='#aed6f1', borderwidth=1),
        legend5=dict(title=dict(text='<b>DrawDown</b>', font=dict(size=11)),
                     orientation='v', yanchor='top', y=0.30, xanchor='left', x=1.01,
                     bgcolor='rgba(255,255,255,0.85)', bordercolor='#aed6f1', borderwidth=1),
        legend6=dict(title=dict(text='<b>Volatilità</b>', font=dict(size=11)),
                     orientation='v', yanchor='top', y=0.16, xanchor='left', x=1.01,
                     bgcolor='rgba(255,255,255,0.85)', bordercolor='#aed6f1', borderwidth=1),
        legend7=dict(title=dict(text='<b>VaR — (solido 90%, punteg. 95%)</b>', font=dict(size=11)),
                     orientation='v', yanchor='top', y=0.02, xanchor='left', x=1.01,
                     bgcolor='rgba(255,255,255,0.85)', bordercolor='#aed6f1', borderwidth=1),
        margin=dict(b=20, t=60, r=220),
        autosize=True,
        xaxis=dict(range=[start_date, end_date]),
        xaxis2=dict(range=[start_date, end_date]),
        xaxis3=dict(range=[start_date, end_date]),
        xaxis4=dict(range=[start_date, end_date]),
        xaxis5=dict(range=[start_date, end_date]),
        xaxis6=dict(range=[start_date, end_date]),
        xaxis7=dict(range=[start_date, end_date]),
    )

    date_values = (f"Intervallo di date: {start_date.strftime('%d-%m-%Y')} — "
                   f"{end_date.strftime('%d-%m-%Y')}")
    insuff_text = (f"Ticker con dati insufficienti: {', '.join(insufficient_data)}"
                   if insufficient_data else "")

    return fig, date_values, insuff_text


# ─────────────────────────────────────────────────────────────────────────────
# Callback: sessione — toggle pannello
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('session-panel', 'style'),
    Input('session-toggle-btn', 'n_clicks'),
    State('session-panel', 'style'),
    prevent_initial_call=True,
)
def toggle_session_panel(n, current_style):
    if current_style and current_style.get('display') == 'none':
        return {'display': 'block', 'position': 'absolute', 'top': '70px', 'right': '10px',
                'z-index': '1000', 'background': 'white', 'border': '1px solid #ccc',
                'border-radius': '8px', 'box-shadow': '0 4px 20px rgba(0,0,0,0.15)',
                'padding': '16px 20px', 'width': '720px', 'max-width': '95vw'}
    return {'display': 'none'}


# ─────────────────────────────────────────────────────────────────────────────
# Callback: sessione — aggiorna lista
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('session-list-container', 'children'),
    Input('session-refresh-btn',    'n_clicks'),
    Input('session-save-btn',       'n_clicks'),
    Input('session-delete-trigger', 'data'),
    Input('session-toggle-btn',     'n_clicks'),
    prevent_initial_call=False,
)
def refresh_session_list(*_):
    sessions = list_sessions()
    if not sessions:
        return html.Div('Nessuna sessione salvata.',
                        style={'font-size': '11px', 'color': '#aaa',
                               'padding': '10px', 'text-align': 'center'})
    return [_build_session_row(r) for r in sessions]


# ─────────────────────────────────────────────────────────────────────────────
# Callback: sessione — salva
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('session-save-status', 'children'),
    Output('session-save-status', 'style'),
    Input('session-save-btn', 'n_clicks'),
    State('session-name-input', 'value'),
    State('session-desc-input', 'value'),
    *[State(sid, 'data') for sid in CLIENT_SESSION_STORES],
    prevent_initial_call=True,
)
def save_session_cb(n_clicks, name, desc, *store_values):
    if not n_clicks:
        raise PreventUpdate
    store_data = {sid: val for sid, val in zip(CLIENT_SESSION_STORES, store_values)
                  if val is not None}
    if not store_data:
        return ('⚠ Nessun dato da salvare. Carica prima i dati.',
                {'color': '#e65100', 'font-size': '10px', 'margin-top': '5px'})
    rec = save_session(name=name or '', description=desc or '', store_data=store_data)
    return (f"✅ Salvata: \"{rec['name']}\" ({rec['size_kb']} KB)",
            {'color': '#1b7a34', 'font-size': '10px', 'margin-top': '5px'})


# ─────────────────────────────────────────────────────────────────────────────
# Callback: sessione — elimina
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('session-delete-trigger', 'data'),
    Input({'type': 'session-delete-btn', 'index': ALL}, 'n_clicks'),
    prevent_initial_call=True,
)
def delete_session_cb(all_clicks):
    ctx = callback_context
    if not ctx.triggered:
        raise PreventUpdate
    triggered = ctx.triggered[0]
    if not triggered['value']:
        raise PreventUpdate
    try:
        id_dict    = json.loads(triggered['prop_id'].split('.')[0])
        session_id = id_dict['index']
    except Exception:
        raise PreventUpdate
    delete_session(session_id)
    return str(session_id)


# ─────────────────────────────────────────────────────────────────────────────
# Callback: sessione — seleziona per il caricamento
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('session-selected-id',  'data'),
    Output('session-load-trigger', 'data'),
    Input({'type': 'session-load-btn', 'index': ALL}, 'n_clicks'),
    prevent_initial_call=True,
)
def select_session(all_clicks):
    ctx = callback_context
    if not ctx.triggered:
        raise PreventUpdate
    triggered = ctx.triggered[0]
    if not triggered['value']:
        raise PreventUpdate
    try:
        id_dict    = json.loads(triggered['prop_id'].split('.')[0])
        session_id = id_dict['index']
    except Exception:
        raise PreventUpdate
    return session_id, session_id


@app.callback(
    *[Output(sid, 'data', allow_duplicate=True) for sid in CLIENT_SESSION_STORES],
    Input('session-load-trigger', 'data'),
    prevent_initial_call=True,
)
def load_session_cb(session_id):
    if not session_id:
        raise PreventUpdate
    store_data = load_session(session_id)
    return tuple(store_data.get(sid, no_update) for sid in CLIENT_SESSION_STORES)


# ═════════════════════════════════════════════════════════════════════════════
# PANNELLO 📁 FILE — selettore default + file personali + salva con nome
# ═════════════════════════════════════════════════════════════════════════════

def _build_fp_row(f):
    """Riga di un file personale nella lista del pannello File."""
    fn   = f['filename']
    name = f.get('saved_name', fn)
    date = f.get('saved_at', '') or f.get('modified', '')
    size = f.get('size_kb', '?')
    btn = {'border': 'none', 'border-radius': '3px', 'cursor': 'pointer',
           'font-size': '10px', 'padding': '3px 8px', 'font-weight': 'bold'}
    return html.Div([
        html.Div([
            html.Span(name, style={'font-weight': 'bold', 'font-size': '11px',
                                   'color': '#1a3a5c'}),
            html.Br(),
            html.Span(f'🕐 {date}  ·  {size} KB',
                      style={'font-size': '9px', 'color': '#888'}),
        ], style={'flex': '1', 'min-width': '0'}),
        html.Div([
            html.Button('📂 Carica', id={'type': 'fp-load-btn', 'index': fn},
                        n_clicks=0,
                        style={**btn, 'background': '#1a3a5c', 'color': 'white',
                               'margin-right': '4px'}),
            html.Button('🗑', id={'type': 'fp-del-btn', 'index': fn}, n_clicks=0,
                        style={**btn, 'background': '#c0392b', 'color': 'white'}),
        ], style={'display': 'flex', 'align-items': 'center', 'flex-shrink': '0',
                  'margin-left': '10px'}),
    ], style={'display': 'flex', 'align-items': 'center', 'padding': '6px 8px',
              'border-bottom': '1px solid #eee', 'background': 'white',
              'border-radius': '3px', 'margin-bottom': '3px'})


# ── Toggle pannello ───────────────────────────────────────────────────────────
@app.callback(
    Output('file-panel', 'style'),
    Input('file-panel-btn', 'n_clicks'),
    Input('fp-close-btn',   'n_clicks'),
    State('file-panel', 'style'),
    prevent_initial_call=True,
)
def toggle_file_panel(n, n_close, cur):
    # Chiudi esplicito → nascondi
    if callback_context.triggered_id == 'fp-close-btn':
        return {'display': 'none'}
    if cur and cur.get('display') == 'none':
        return {'display': 'block', 'position': 'absolute', 'top': '70px', 'left': '10px',
                'z-index': '1000', 'background': 'white', 'border': '1px solid #ccc',
                'border-radius': '8px', 'box-shadow': '0 4px 20px rgba(0,0,0,0.15)',
                'padding': '16px 20px', 'width': '620px', 'max-width': '95vw'}
    return {'display': 'none'}


# ── Lista file personali ────────────────────────────────────────────────────
@app.callback(
    Output('fp-file-list', 'children'),
    Input('fp-refresh-btn',   'n_clicks'),
    Input('fp-save-btn',      'n_clicks'),
    Input('file-panel-btn',   'n_clicks'),
    Input('fp-delete-trigger', 'data'),
    prevent_initial_call=False,
)
def refresh_fp_list(*_):
    files = _sm.list_user_files(_get_username())
    if not files:
        return html.Div('Nessun file personale salvato.',
                        style={'font-size': '11px', 'color': '#aaa',
                               'padding': '10px', 'text-align': 'center'})
    return [_build_fp_row(f) for f in files]


# ── Salva con nome ──────────────────────────────────────────────────────────
@app.callback(
    Output('fp-save-status', 'children'),
    Output('fp-save-status', 'style'),
    Input('fp-save-btn',     'n_clicks'),
    State('fp-save-name',    'value'),
    *[State(sid, 'data') for sid in CLIENT_SESSION_STORES],
    prevent_initial_call=True,
)
def fp_save_named(n, name, *store_values):
    if not n:
        raise PreventUpdate
    _err = {'color': '#e65100', 'font-size': '10px', 'margin-top': '5px'}
    _ok  = {'color': '#1b7a34', 'font-size': '10px', 'margin-top': '5px'}
    if not (name and name.strip()):
        return '⚠ Inserisci un nome per il file.', _err

    _u = _get_username()
    stores = {sid: val for sid, val in zip(CLIENT_SESSION_STORES, store_values)}
    with _DL_LOCK:
        cr       = _DL_BUFFER.get('close_returns')
        op       = _DL_BUFFER.get('original_prices')
        tm       = dict(_DL_BUFFER.get('ticker_map', {}))
        vm       = dict(_DL_BUFFER.get('valuta_map', {}))
        saved_at = _DL_BUFFER.get('saved_at', '')
    # Fallback: ricostruisci cr dai 9 store se il buffer è vuoto
    if cr is None and stores.get('stock-data'):
        try:
            cr = pd.read_json(io.StringIO(stores['stock-data']), orient='split')
        except Exception:
            cr = None
    if cr is None:
        return '⚠ Nessun dato da salvare. Carica prima un file.', _err

    data = {
        'close_returns':   cr,
        'original_prices': op,
        'ticker_map':      tm,
        'valuta_map':      vm,
        'saved_at':        saved_at,
        '_stores':         stores,
    }
    path = _sm.save_named(_u, data, name.strip())
    return f'✅ Salvato: {path.name}', _ok


# ── Stage caricamento file (default o personale) con warning ─────────────────
@app.callback(
    Output('pending-fileload-store', 'data'),
    Output('fileload-modal-overlay', 'style'),
    Input({'type': 'fp-default-btn', 'index': ALL}, 'n_clicks'),
    Input({'type': 'fp-load-btn',    'index': ALL}, 'n_clicks'),
    prevent_initial_call=True,
)
def stage_file_load(default_clicks, load_clicks):
    ctx = callback_context
    if not ctx.triggered:
        raise PreventUpdate
    trig = ctx.triggered[0]
    if not trig['value']:
        raise PreventUpdate
    try:
        id_dict = json.loads(trig['prop_id'].split('.')[0])
    except Exception:
        raise PreventUpdate
    kind = 'default' if id_dict['type'] == 'fp-default-btn' else 'personal'
    needs_confirm = _sm.has_unsaved_changes(_get_username())
    payload = {'kind': kind, 'key': id_dict['index'],
               'await': needs_confirm, 'ts': time.time()}
    return payload, (_EDITOR_SHOWN if needs_confirm else _EDITOR_HIDDEN)


@app.callback(
    Output('pending-fileload-store', 'data', allow_duplicate=True),
    Output('fileload-modal-overlay', 'style', allow_duplicate=True),
    Input('fileload-no', 'n_clicks'),
    Input('fileload-x',  'n_clicks'),
    prevent_initial_call=True,
)
def cancel_file_load(n_no, n_x):
    if not (n_no or n_x):
        raise PreventUpdate
    return None, _EDITOR_HIDDEN


# Chiudi il modal al click di "Sì, carica" (il caricamento lo fa execute_file_load)
@app.callback(
    Output('fileload-modal-overlay', 'style', allow_duplicate=True),
    Input('fileload-yes', 'n_clicks'),
    prevent_initial_call=True,
)
def _close_fileload_modal(n):
    if not n:
        raise PreventUpdate
    return _EDITOR_HIDDEN


# ── Esegui caricamento file ──────────────────────────────────────────────────
@app.callback(
    Output('stock-data',              'data',     allow_duplicate=True),
    Output('original-prices-data',    'data',     allow_duplicate=True),
    Output('asset-checklist',         'data',     allow_duplicate=True),
    Output('ticker-map-store',        'data',     allow_duplicate=True),
    Output('data-last-updated',       'children', allow_duplicate=True),
    Output('upload-status',           'children', allow_duplicate=True),
    Output('weights-store-P1',        'data',     allow_duplicate=True),
    Output('weights-store-P2',        'data',     allow_duplicate=True),
    Output('weights-store-P3',        'data',     allow_duplicate=True),
    Output('global-assets-selected',  'data',     allow_duplicate=True),
    Output('insufficient-data-store', 'data',     allow_duplicate=True),
    Output('update-portfolio-button', 'n_clicks', allow_duplicate=True),
    Input('pending-fileload-store',   'data'),
    Input('fileload-yes',             'n_clicks'),
    State('update-portfolio-button',  'n_clicks'),
    prevent_initial_call=True,
)
def execute_file_load(pending, submit_n, cur_clicks):
    ctx = callback_context
    trig_id = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else ''
    if not pending:
        raise PreventUpdate
    # Serviva conferma: non eseguire sul cambio dello store, attendi il click "Sì"
    if trig_id == 'pending-fileload-store' and pending.get('await'):
        raise PreventUpdate
    if trig_id == 'fileload-yes' and not submit_n:
        raise PreventUpdate

    _u   = _get_username()
    kind = pending.get('kind')
    key  = pending.get('key')
    _cl_clear(_u)
    _active_file_store['is_personale'] = (kind == 'personal')

    # ── Carica i dati ─────────────────────────────────────────────────────────
    if kind == 'default':
        data = _sm.load_default(key)
    else:
        data = _sm.load_named(_u, key)
    if not data:
        raise PreventUpdate

    cr     = data.get('close_returns')
    op     = data.get('original_prices')
    tm     = dict(data.get('ticker_map', {}))
    vm     = dict(data.get('valuta_map', {}))
    sa     = data.get('saved_at', '')
    stores = data.get('_stores', {}) or {}

    # Ricostruisci cr/op dai 9 store se mancano i DataFrame
    if cr is None and stores.get('stock-data'):
        try:
            cr = pd.read_json(io.StringIO(stores['stock-data']), orient='split')
        except Exception:
            cr = None
    if op is None and stores.get('original-prices-data'):
        try:
            op = pd.read_json(io.StringIO(stores['original-prices-data']), orient='split')
        except Exception:
            op = None
    if cr is None:
        raise PreventUpdate

    # Aggiorna buffer condiviso
    with _DL_LOCK:
        _DL_BUFFER.clear()
        _DL_BUFFER.update({'close_returns': cr, 'original_prices': op,
                           'ticker_map': tm, 'valuta_map': vm, 'saved_at': sa})

    # Output store
    stock_json = stores.get('stock-data') or cr.to_json(date_format='iso', orient='split')
    op_json    = (stores.get('original-prices-data')
                  or (op.to_json(date_format='iso', orient='split') if op is not None else None))
    options    = stores.get('asset-checklist') or [{'label': c, 'value': c} for c in cr.columns]
    tmap_out   = stores.get('ticker-map-store') or tm

    if kind == 'personal':
        p1 = stores.get('weights-store-P1', {}) or {}
        p2 = stores.get('weights-store-P2', {}) or {}
        p3 = stores.get('weights-store-P3', {}) or {}
        gsel  = stores.get('global-assets-selected', []) or []
        insuf = stores.get('insufficient-data-store', []) or []
        _write_user_json(cr, op, tm, vm, reset_state=True)
        _update_user_json(weights={'P1': p1, 'P2': p2, 'P3': p3}, username=_u)
        label = f'✓ {len(options)} asset — {pending.get("key")}'
    else:
        p1, p2, p3 = {}, {}, {}
        gsel, insuf = [], []
        _write_user_json(cr, op, tm, vm, reset_state=True)
        label = f'✓ {len(options)} asset — {key}'

    # Salva la nuova sessione di lavoro condivisa
    _sm.save_working(_u, {'close_returns': cr, 'original_prices': op,
                          'ticker_map': tm, 'valuta_map': vm, 'saved_at': sa,
                          '_stores': stores}, source=key)

    return (
        stock_json, op_json, options, tmap_out,
        f'Aggiornati: {sa}' if sa else '',
        html.Div(label, style={'color': '#007755', 'font-size': '11px'}),
        p1, p2, p3, gsel, insuf,
        (cur_clicks or 0) + 1,
    )


@app.callback(
    Output('fp-delete-trigger', 'data'),
    Input({'type': 'fp-del-btn', 'index': ALL}, 'n_clicks'),
    prevent_initial_call=True,
)
def delete_fp_file(all_clicks):
    ctx = callback_context
    if not ctx.triggered:
        raise PreventUpdate
    trig = ctx.triggered[0]
    if not trig['value']:
        raise PreventUpdate
    try:
        id_dict  = json.loads(trig['prop_id'].split('.')[0])
        filename = id_dict['index']
    except Exception:
        raise PreventUpdate
    _sm.delete_user_file(_get_username(), filename)
    return f'{filename}:{time.time()}'


# ── Marca sessione modificata su cambio pesi (per il warning) ────────────────
@app.callback(
    Output('sm-dirty-sink', 'data'),
    Input('weights-store-P1', 'data'),
    Input('weights-store-P2', 'data'),
    Input('weights-store-P3', 'data'),
    prevent_initial_call=True,
)
def mark_session_dirty(p1, p2, p3):
    _sm.mark_modified(_get_username())
    return no_update


# ─────────────────────────────────────────────────────────────────────────────
# Startup: carica tutti i pkl esistenti, scarica in background quelli mancanti
# ─────────────────────────────────────────────────────────────────────────────
def _startup_load():
    global _DL_STATE, _DL_BUFFER

    # Carica ETF (file attivo di default) nel buffer principale
    if _MARKET_DATA_FILE.exists():
        try:
            with open(_MARKET_DATA_FILE, 'rb') as f:
                data = pickle.load(f)
            with _DL_LOCK:
                _DL_BUFFER.update(data)
                _DL_STATE['status']  = 'done'
                _DL_STATE['current'] = 1
                _DL_STATE['total']   = 1
            print(f"✓ Dati ETF caricati da disco — {data.get('saved_at', '?')}")
        except Exception as e:
            print(f"⚠ Lettura market_data.pkl fallita: {e}")

    # Carica ARIMA dal file separato (se esiste)
    _arima_pkl = _arima_cache_path('ETF.xlsx')
    if _arima_pkl.exists():
        try:
            with open(_arima_pkl, 'rb') as f:
                arima_data = pickle.load(f)
            with _DL_LOCK:
                if arima_data.get('arima'):
                    _DL_BUFFER['arima']             = arima_data['arima']
                    _DL_BUFFER['arima_computed_at'] = arima_data.get('arima_computed_at', '')
            print(f"✓ ARIMA caricato da {_arima_pkl.name}")
        except Exception as e:
            print(f"⚠ Lettura {_arima_pkl.name} fallita: {e}")

    # Scarica in background tutti i file xlsx per cui manca il pkl
    def _bg_all():
        start = (pd.Timestamp.today() - pd.DateOffset(years=10)).strftime('%Y-%m-%d')
        xlsx_files = sorted(_FILES_DIR.glob('*.xlsx')) if _FILES_DIR.exists() else [Path(_XLSX)]
        for xlsx_path in xlsx_files:
            filename  = xlsx_path.name
            cache_pkl = _file_cache_path(filename)
            if Path(cache_pkl).exists():
                print(f"✓ Cache {filename} già presente — skip download")
                continue
            try:
                print(f"▶ Download iniziale {filename}…")
                tickers, descr, valuta = _build_ticker_list(filename)
                is_etf = (filename == 'ETF.xlsx')
                _do_download(tickers, descr, valuta, start,
                             cache_file=cache_pkl, update_buffer=is_etf)
                print(f"✓ Download completato: {filename}")
            except Exception as e:
                print(f"⚠ Download iniziale {filename} fallito: {e}")

    threading.Thread(target=_bg_all, daemon=True).start()

_startup_load()



# ─────────────────────────────────────────────────────────────────────────────
# ARIMA+GARCH notturno: calcola mu/cov e li scrive in market_data.pkl['arima']
# Il formato è quello letto da frontiera-efficiente/app.py → _read_arima_cache()
# ─────────────────────────────────────────────────────────────────────────────
def _compute_arima_garch(returns_df, window=250):
    """ARIMA(p,0,q) best-AIC + GARCH(1,1). Restituisce (mu_series, cov_df) annualizzati."""
    import warnings
    try:
        from statsmodels.tsa.arima.model import ARIMA as _ARIMA
    except ImportError:
        print("⚠ statsmodels non installato — ARIMA saltato")
        return None, None
    try:
        from arch import arch_model as _arch
        _has_arch = True
    except ImportError:
        _has_arch = False

    cols = [c for c in returns_df.columns if returns_df[c].dropna().shape[0] >= 60]
    mu_d, sig_d, resid_d = {}, {}, {}

    for i, col in enumerate(cols):
        # Conversione a rendimenti logaritmici: più stazionari e normali
        s_arith = returns_df[col].dropna().tail(window)
        s = np.log1p(s_arith)  # ln(1 + r_aritm) = rendimento logaritmico giornaliero

        best_aic  = np.inf
        best_mu   = float(s.mean())
        best_resid = (s - s.mean()).values

        for p in range(3):
            for q in range(3):
                if p == 0 and q == 0:
                    continue
                try:
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore')
                        m = _ARIMA(s, order=(p, 0, q)).fit()
                    if m.aic < best_aic:
                        best_aic   = m.aic
                        best_resid = m.resid.values
                        # Usa la previsione a 252 passi (come il cono di previsione ARIMA):
                        # evita l'instabilità di const/(1-sum(AR)) per processi near unit root
                        fc = m.get_forecast(steps=252).predicted_mean
                        best_mu = float(fc.mean())
                except Exception:
                    pass
        mu_d[col]    = best_mu   # media logaritmica giornaliera incondizionata
        resid_d[col] = best_resid

        sig_fallback = float(np.std(best_resid)) or float(s.std())
        if _has_arch and len(best_resid) > 20:
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    gm = _arch(best_resid * 100, vol='Garch', p=1, q=1, rescale=False)
                    gf = gm.fit(disp='off')
                    pr = gf.params
                    o, a, b = pr['omega'], pr['alpha[1]'], pr['beta[1]']
                    if a + b < 1.0:
                        sig_garch = float(np.sqrt(o / (1 - a - b))) / 100
                        # safeguard: rifiuta stime GARCH fuori range (< 1/10 o > 10x lo std storico)
                        sig_d[col] = sig_garch if sig_fallback / 10 < sig_garch < sig_fallback * 10 else sig_fallback
                    else:
                        sig_d[col] = sig_fallback
            except Exception:
                sig_d[col] = sig_fallback
        else:
            sig_d[col] = sig_fallback
        print(f"  [{i+1}/{len(cols)}] {col}: mu_log={mu_d[col]:.5f} vol_log={sig_d[col]:.5f}")

    if not mu_d:
        return None, None

    min_len = min(len(resid_d[c]) for c in cols)
    r_mat   = np.column_stack([resid_d[c][:min_len] for c in cols])
    corr    = np.nan_to_num(np.corrcoef(r_mat.T), nan=0.0)
    np.fill_diagonal(corr, 1.0)
    vols    = np.array([sig_d[c] for c in cols])
    D       = np.diag(vols)
    cov_d   = D @ corr @ D + 1e-8 * np.eye(len(cols))

    # Correzione di Jensen: converte rendimento log atteso in rendimento aritmetico atteso
    # E[R_aritm] = exp(mu_log * 252 + 0.5 * sigma_log^2 * 252) - 1
    def _jensen(mu_log, sig_log, s_arith_fallback):
        exp_arg = mu_log * 252 + 0.5 * sig_log**2 * 252
        if not np.isfinite(exp_arg) or exp_arg > 3.0:  # cap a ~1900% — oltre è numericamente rotto
            return float(s_arith_fallback.mean() * 252)
        return float(np.exp(exp_arg) - 1)

    mu_annual  = pd.Series(
        {c: _jensen(mu_d[c], sig_d[c], returns_df[c].dropna().tail(window)) for c in cols}
    )
    cov_annual = pd.DataFrame(cov_d * 252, index=cols, columns=cols)
    return mu_annual, cov_annual


def _save_arima_to_pkl(mu_series, cov_df, target_pkl=None):
    """Salva ARIMA nel file separato (NON dentro il pkl principale)."""
    ts = datetime.now().strftime('%d/%m/%Y %H:%M')
    arima_data = {
        'mu':          mu_series.to_dict(),
        'cov':         cov_df.to_dict(),
        'computed_at': ts,
    }
    # target_pkl qui è il percorso del FILE ARIMA separato (non il pkl principale)
    arima_pkl = Path(target_pkl) if target_pkl else _arima_cache_path('ETF.xlsx')
    try:
        existing = {}
        if arima_pkl.exists():
            with open(arima_pkl, 'rb') as f:
                existing = pickle.load(f)
        existing['arima']             = arima_data
        existing['arima_computed_at'] = ts
        _atomic_pkl_write(arima_pkl, existing)
        print(f"✓ ARIMA salvato in {arima_pkl.name} — {ts}")
    except Exception as e:
        print(f"⚠ Salvataggio ARIMA fallito ({arima_pkl.name}): {e}")
    # Aggiorna buffer in memoria (solo per il file ETF principale)
    if target_pkl is None or 'market_data_arima' in str(arima_pkl):
        with _DL_LOCK:
            _DL_BUFFER['arima']             = arima_data
            _DL_BUFFER['arima_computed_at'] = ts


# ─────────────────────────────────────────────────────────────────────────────
# Scheduler: dati alle 18:30 (lun-ven), ARIMA+GARCH a mezzanotte (lun-sab)
# ─────────────────────────────────────────────────────────────────────────────
def _scheduled_update():
    """Aggiornamento notturno: scarica dati per tutti i file in Files/."""
    start = (pd.Timestamp.today() - pd.DateOffset(years=10)).strftime('%Y-%m-%d')
    active_file = _active_file_store.get('filename', 'ETF.xlsx')
    xlsx_files = sorted(_FILES_DIR.glob('*.xlsx')) if _FILES_DIR.exists() else []
    if not xlsx_files:
        xlsx_files = [Path(_XLSX)]
    for xlsx_path in xlsx_files:
        filename = xlsx_path.name
        try:
            tickers, descr, valuta = _build_ticker_list(filename)
            cache = _file_cache_path(filename)
            is_active = (filename == active_file)
            print(f"⏰ Aggiornamento {filename}: {len(tickers)} ticker")
            _do_download(tickers, descr, valuta, start,
                         cache_file=cache, update_buffer=is_active)
        except Exception as e:
            print(f"⚠ Aggiornamento {filename} fallito: {e}")


def _scheduled_arima():
    """Calcolo ARIMA+GARCH a mezzanotte su tutti i file in Files/."""
    xlsx_files = sorted(_FILES_DIR.glob('*.xlsx')) if _FILES_DIR.exists() else [Path(_XLSX)]
    for xlsx_path in xlsx_files:
        filename = xlsx_path.name
        cache_pkl = _file_cache_path(filename)
        arima_pkl = _arima_cache_path(filename)
        try:
            if not Path(cache_pkl).exists():
                print(f"⚠ ARIMA {filename}: cache non trovata, skip")
                continue
            with open(cache_pkl, 'rb') as f:
                cached = pickle.load(f)
            ret = cached.get('close_returns')
            if ret is None or (hasattr(ret, 'empty') and ret.empty):
                print(f"⚠ ARIMA {filename}: nessun dato disponibile")
                continue
            print(f"🌙 ARIMA+GARCH {filename} — {len(ret.columns)} asset…")
            mu, cov = _compute_arima_garch(ret, window=250)
            if mu is not None:
                _save_arima_to_pkl(mu, cov, target_pkl=arima_pkl)
                print(f"✓ ARIMA+GARCH completato: {filename}")
            else:
                print(f"⚠ ARIMA {filename}: calcolo non riuscito")
        except Exception as e:
            print(f"⚠ ARIMA notturno {filename} fallito: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Callback: reset asset ai default
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('upload-status',           'children',  allow_duplicate=True),
    Output('asset-checklist',         'data',      allow_duplicate=True),
    Output('stock-data',              'data',      allow_duplicate=True),
    Output('original-prices-data',    'data',      allow_duplicate=True),
    Output('ticker-map-store',        'data',      allow_duplicate=True),
    Output('data-last-updated',       'children',  allow_duplicate=True),
    Output('weights-store-P1',        'data',      allow_duplicate=True),
    Output('weights-store-P2',        'data',      allow_duplicate=True),
    Output('weights-store-P3',        'data',      allow_duplicate=True),
    Output('update-portfolio-button', 'n_clicks',  allow_duplicate=True),
    Output('file-selector',           'options',   allow_duplicate=True),
    Output('file-selector',           'value',     allow_duplicate=True),
    Input('reset-default-btn',        'n_clicks'),
    State('update-portfolio-button',  'n_clicks'),
    prevent_initial_call=True,
)
def reset_to_default(n, cur_clicks):
    if not n:
        raise PreventUpdate
    _err = (no_update,) * 11
    if not _MARKET_DATA_FILE.exists():
        return ('⚠ File di default non trovato', *_err)
    try:
        with open(_MARKET_DATA_FILE, 'rb') as f:
            d = pickle.load(f)
        cr = d.get('close_returns')
        op = d.get('original_prices')
        tm = d.get('ticker_map', {})
        vm = d.get('valuta_map', {})
        saved_at = d.get('saved_at', '')
    except Exception as e:
        return (f'⚠ Errore: {e}', *_err)
    if cr is None or op is None:
        return ('⚠ Dati non trovati nel file di default', *_err)
    with _DL_LOCK:
        _DL_BUFFER.update({'close_returns': cr, 'original_prices': op,
                           'ticker_map': tm, 'valuta_map': vm, 'saved_at': saved_at})
    _write_user_json(cr, op, tm, vm, reset_state=True)
    _active_file_store['is_personale'] = False
    options  = [{'label': col, 'value': col} for col in cr.columns]
    last_upd = f'Aggiornati: {saved_at}' if saved_at else ''
    return (
        html.Div(f'✓ {len(options)} asset — file di default',
                 style={'color': '#007755', 'font-size': '11px'}),
        options,
        cr.to_json(date_format='iso', orient='split'),
        op.to_json(date_format='iso', orient='split'),
        tm,
        last_upd,
        {}, {}, {},
        (cur_clicks or 0) + 1,
        _list_files(), 'ETF.xlsx',
    )


# ─────────────────────────────────────────────────────────────────────────────
# Callback: ripristina file-selector al caricamento pagina
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output('file-selector', 'options', allow_duplicate=True),
    Output('file-selector', 'value',   allow_duplicate=True),
    Input('nav-reload', 'data'),
    prevent_initial_call=True,
)
def restore_file_selector(nav_reload):
    # La selezione rispecchia il tipo del file di lavoro (un solo file: current.json).
    # Se è "personale" → 👤 Personale; altrimenti default ETF.
    _u = _get_username()
    is_pers = (_read_tipo(_u) == 'personale')
    opts = _list_files_with_personale() if is_pers else _list_files()
    return opts, ('__personale__' if is_pers else 'ETF.xlsx')


try:
    from apscheduler.schedulers.background import BackgroundScheduler
    _scheduler = BackgroundScheduler(timezone='Europe/Rome')
    _scheduler.add_job(_scheduled_update, 'cron', hour=0, minute=0,
                       day_of_week='mon-sat', misfire_grace_time=3600)
    _scheduler.add_job(_scheduled_arima,  'cron', hour=0, minute=30,
                       day_of_week='mon-sat', misfire_grace_time=3600)
    _scheduler.start()
    print("✓ Scheduler avviato — dati mezzanotte, ARIMA 00:30 (lun-sab)")
except ImportError:
    print("⚠ apscheduler non installato — aggiornamento automatico disabilitato")
    print("  pip install apscheduler")

# ─────────────────────────────────────────────────────────────────────────────
# Callbacks: Converti ISIN
# ─────────────────────────────────────────────────────────────────────────────

@app.callback(
    Output('isin-modal-overlay', 'style'),
    Input('isin-open-btn',  'n_clicks'),
    Input('isin-close-btn', 'n_clicks'),
    State('isin-modal-overlay', 'style'),
    prevent_initial_call=True,
)
def _toggle_isin_modal(open_n, close_n, style):
    _BASE = {'position': 'fixed', 'top': '0', 'left': '0', 'width': '100%', 'height': '100%',
             'z-index': '9000', 'background': 'rgba(0,0,0,0.45)',
             'align-items': 'center', 'justify-content': 'center'}
    tid = callback_context.triggered[0]['prop_id'].split('.')[0] if callback_context.triggered else ''
    return {**_BASE, 'display': 'flex'} if tid == 'isin-open-btn' else {**_BASE, 'display': 'none'}


app.clientside_callback(
    "function(ts){ return ts ? null : window.dash_clientside.no_update; }",
    Output('isin-upload', 'contents', allow_duplicate=True),
    Input('isin-req-id',  'data'),
    prevent_initial_call=True,
)

@app.callback(
    Output('isin-req-id',       'data'),
    Output('isin-poll',         'disabled'),
    Output('isin-progress-text','children'),
    Output('isin-dl-btn',       'style'),
    Output('isin-load-status',  'children'),
    Input('isin-upload',        'contents'),
    State('isin-upload',        'filename'),
    prevent_initial_call=True,
)
def _start_isin_conversion(contents, filename):
    if not contents:
        raise PreventUpdate
    import uuid, base64
    _u   = _get_username()
    rid  = str(uuid.uuid4())[:8]
    _, b64 = contents.split(',', 1)
    raw  = base64.b64decode(b64)
    with _ISIN_LOCK:
        _ISIN_STATE.update({
            'running': True, 'done': False, 'req_id': rid,
            'progress': 'Avvio conversione…', 'n_done': 0, 'n_total': 0,
            'result_bytes': None, 'excluded': [],
            'tickers': [], 'descr': [], 'valuta': [], 'pesi': [],
        })
    threading.Thread(target=_run_isin_conversion, args=(raw, _u, rid), daemon=True).start()
    _BTN_HIDE = {'display': 'none'}
    return rid, False, f'⏳ Avvio conversione {filename}…', _BTN_HIDE, ''


@app.callback(
    Output('isin-progress-text', 'children',     allow_duplicate=True),
    Output('isin-poll',          'disabled',     allow_duplicate=True),
    Output('isin-dl-btn',        'style',        allow_duplicate=True),
    Output('isin-load-status',   'children',     allow_duplicate=True),
    Output('stock-data',         'data',         allow_duplicate=True),
    Output('asset-checklist',    'data',         allow_duplicate=True),
    Output('update-portfolio-button', 'n_clicks', allow_duplicate=True),
    Input('isin-poll',           'n_intervals'),
    State('isin-req-id',         'data'),
    State('update-portfolio-button', 'n_clicks'),
    prevent_initial_call=True,
)
def _poll_isin(_, req_id, cur_clicks):
    _NU = no_update
    _BTN_SHOW = {'display': 'inline-block', 'font-size': '11px', 'padding': '6px 14px',
                 'background': '#1b7a34', 'color': 'white', 'border': 'none',
                 'border-radius': '4px', 'cursor': 'pointer',
                 'font-weight': 'bold', 'margin-right': '8px'}
    _BTN_HIDE = {'display': 'none'}

    if not req_id:
        raise PreventUpdate
    with _ISIN_LOCK:
        s = dict(_ISIN_STATE)
    if s.get('req_id') != req_id:
        raise PreventUpdate

    progress = s.get('progress', '')
    n_done   = s.get('n_done', 0)
    n_total  = s.get('n_total', 1)
    pct      = int(n_done / max(n_total, 1) * 100)
    txt      = f'{progress} ({pct}%)' if n_total > 0 else progress

    if not s.get('done'):
        return txt, False, _BTN_HIDE, '', _NU, _NU, _NU

    # Done
    load_msg = _NU
    new_stock = _NU
    new_opts  = _NU
    new_clicks = _NU

    tickers = s.get('tickers', [])
    if tickers:
        with _DL_LOCK:
            cr = _DL_BUFFER.get('close_returns')
        if cr is not None:
            new_stock  = cr.to_json(date_format='iso', orient='split')
            new_opts   = [{'label': c, 'value': c} for c in cr.columns]
            new_clicks = (cur_clicks or 0) + 1
        n_excl   = len(s.get('excluded', []))
        load_msg = (f'⏳ Download dati in corso per {len(tickers)} ticker…'
                    if cr is None else
                    f'✓ Dati caricati — {len(tickers)} titoli'
                    + (f', {n_excl} esclusi' if n_excl else ''))
    else:
        load_msg = s.get('progress', '')

    return txt, True, _BTN_SHOW, load_msg, new_stock, new_opts, new_clicks


@app.callback(
    Output('isin-download-data', 'data'),
    Input('isin-dl-btn', 'n_clicks'),
    State('isin-req-id', 'data'),
    prevent_initial_call=True,
)
def _download_isin_excel(n, req_id):
    if not n or not req_id:
        raise PreventUpdate
    with _ISIN_LOCK:
        s = dict(_ISIN_STATE)
    if s.get('req_id') != req_id or not s.get('result_bytes'):
        raise PreventUpdate
    return dcc.send_bytes(s['result_bytes'], 'portafoglio_convertito.xlsx')


register_style_analysis_callbacks(app)

# ─────────────────────────────────────────────────────────────────────────────
server = app.server   # esposto per gunicorn

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8051))
    app.run(debug=False, port=port, host='0.0.0.0')

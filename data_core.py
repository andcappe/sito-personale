"""
data_core.py — Logica dati CONDIVISA (un solo posto, richiamato da più app).

Tutto ruota attorno al file UNICO per utente: sessions/<utente>/current.json
(dataset + pesi P1/P2/P3) e sessions/<utente>/analyses.json (analisi salvate).

Espone: lettura/scrittura current.json, download prezzi con conversione valuta
in EUR, aggiungi-asset (accoda a current.json), template/esporta Excel, e
salva/carica "tutto il lavoro" (snapshot di current.json + analyses.json).

NON contiene componenti o callback Dash: è una libreria pura, importabile da
qualunque app (Analisi Tattica, Portafoglio, …) senza effetti collaterali.
"""
import os
import io
import json
from pathlib import Path

import pandas as pd

ROOT = Path(os.path.dirname(os.path.abspath(__file__)))


# ─── Storage persistente (S3/R2) ──────────────────────────────────────────────
def cloud_push(path):
    """Replica il file sullo storage persistente (R2) se configurato. Best-effort."""
    try:
        import cloud_storage
        cloud_storage.push(path)
    except Exception:
        pass


# ─── Utente / percorsi ────────────────────────────────────────────────────────
def get_username():
    try:
        from flask import session as _fs
        return _fs.get('username') or 'anon'
    except Exception:
        return 'anon'


def current_path(username=None):
    u = username or get_username()
    d = ROOT / 'sessions' / u
    d.mkdir(parents=True, exist_ok=True)
    return d / 'current.json'


def analyses_path(username=None):
    u = username or get_username()
    return ROOT / 'sessions' / u / 'analyses.json'


# ─── current.json: lettura / scrittura ───────────────────────────────────────
def read_current(username=None):
    try:
        with open(current_path(username)) as f:
            raw = json.load(f)
        # current.json contiene solo voci-asset (dict). Eventuali chiavi meta
        # (es. "_tipo": "personale"|"default:ETF") vengono ignorate qui, così
        # ogni iterazione a valle vede solo asset.
        return {k: v for k, v in raw.items() if isinstance(v, dict)}
    except Exception:
        return {}


def read_meta(username=None):
    """Legge le chiavi meta (non-asset) di current.json, es. {'_tipo': ...}."""
    try:
        with open(current_path(username)) as f:
            raw = json.load(f)
        return {k: v for k, v in raw.items() if not isinstance(v, dict)}
    except Exception:
        return {}


def write_current(data, username=None):
    """Scrittura atomica di current.json + replica su storage persistente."""
    path = current_path(username)
    try:
        tmp = str(path) + '.tmp'
        with open(tmp, 'w') as f:
            json.dump(data, f)
        os.replace(tmp, path)
        cloud_push(path)
        return True
    except Exception as e:
        print(f"⚠ [data_core] scrittura current.json fallita: {e}", flush=True)
        return False


def write_json_atomic(path, data):
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = str(path) + '.tmp'
        with open(tmp, 'w') as f:
            json.dump(data, f)
        os.replace(tmp, path)
        cloud_push(path)
        return True
    except Exception as e:
        print(f"⚠ [data_core] scrittura {getattr(path, 'name', path)} fallita: {e}", flush=True)
        return False


def read_analyses(username=None):
    try:
        with open(analyses_path(username)) as f:
            return json.load(f)
    except Exception:
        return {}


# ─── Viste sui dati ───────────────────────────────────────────────────────────
def asset_options(username=None):
    return [{'label': a, 'value': a} for a in read_current(username).keys()]


def build_dataset(username=None):
    """
    Ricostruisce (close_returns, original_prices, ticker_map) dal file UNICO
    current.json. Usato per mostrare i dati senza dipendere dai buffer in memoria.
    """
    data = read_current(username)
    pcols, rcols, tm = {}, {}, {}
    for a, v in data.items():
        if not isinstance(v, dict):
            continue
        dates = v.get('dates')
        if not dates:
            continue
        try:
            idx = pd.to_datetime(dates)
        except Exception:
            continue
        if v.get('prices') and len(v['prices']) == len(dates):
            pcols[a] = pd.Series(v['prices'], index=idx)
        if v.get('returns') and len(v['returns']) == len(dates):
            rcols[a] = pd.Series(v['returns'], index=idx)
        tm[a] = v.get('ticker') or a
    op = pd.DataFrame(pcols).sort_index() if pcols else None
    cr = pd.DataFrame(rcols).sort_index() if rcols else None
    return cr, op, tm


def build_prices(username=None):
    """DataFrame dei PREZZI da current.json (chiave 'prices' per asset)."""
    data = read_current(username)
    cols = {}
    for asset, v in data.items():
        if not isinstance(v, dict):
            continue
        dates, prices = v.get('dates'), v.get('prices')
        if not dates or not prices or len(dates) != len(prices):
            continue
        try:
            cols[asset] = pd.Series(prices, index=pd.to_datetime(dates))
        except Exception:
            continue
    if not cols:
        return None
    try:
        return pd.DataFrame(cols).sort_index()
    except Exception:
        return None


# ─── Download + conversione valuta (come Analisi di Portafoglio) ──────────────
def fx_series(name, start):
    """Serie del cambio (es. EURUSD=X) per la conversione in EUR."""
    import yfinance as yf
    try:
        fx = yf.download(name, start=start, auto_adjust=True, progress=False)
        if fx is None or len(fx) == 0:
            return None
        c = fx['Close']
        if isinstance(c, pd.DataFrame):
            c = c.iloc[:, 0]
        return c.ffill()
    except Exception:
        return None


def download_series(ticker, currency='EUR'):
    """
    Scarica i prezzi (Close adj) da Yahoo e li CONVERTE in EUR in base alla valuta
    (USD→/EURUSD, GBP→/EURGBP). Ritorna pd.Series o None.
    """
    import yfinance as yf
    start = (pd.Timestamp.today() - pd.DateOffset(years=10)).strftime('%Y-%m-%d')
    try:
        df = yf.download(ticker, start=start, auto_adjust=True, progress=False)
    except Exception:
        return None
    if df is None or len(df) == 0:
        return None
    close = df['Close']
    if isinstance(close, pd.DataFrame):
        close = close.iloc[:, 0]
    px = close.ffill().dropna()
    cur = (currency or 'EUR').upper()
    if cur == 'USD':
        fx = fx_series('EURUSD=X', start)
        if fx is not None:
            px = (px / fx.reindex(px.index).ffill()).dropna()
    elif cur == 'GBP':
        fx = fx_series('EURGBP=X', start)
        if fx is not None:
            px = (px / fx.reindex(px.index).ffill()).dropna()
    return px


def add_asset_to_current(ticker, description, currency='EUR', username=None):
    """Scarica un ticker e lo ACCODA a current.json (file unico). → (ok, messaggio)."""
    ticker      = (ticker or '').strip()
    description = (description or '').strip() or ticker
    currency    = (currency or 'EUR').strip() or 'EUR'
    if not ticker:
        return False, "⚠ Inserisci un ticker"
    px = download_series(ticker, currency)
    if px is None or len(px) < 30:
        return False, f"⚠ Nessun dato per '{ticker}' (ticker corretto?)"
    rets = px.pct_change(fill_method=None)
    dates = [d.strftime('%Y-%m-%d') for d in px.index]
    entry = {
        'ticker':   ticker,
        'currency': currency,
        'dates':    dates,
        'prices':   [round(float(v), 4) if pd.notna(v) else None for v in px],
        'returns':  [round(float(v), 6) if pd.notna(v) else None for v in rets],
        'checked':  False, 'P1': 0, 'P2': 0, 'P3': 0,
    }
    data = read_current(username)
    is_new = description not in data
    data[description] = entry
    if not write_current(data, username):
        return False, "⚠ Errore salvataggio current.json"
    return True, f"✓ {description} ({ticker}) {'aggiunto' if is_new else 'aggiornato'} — {len(px)} prezzi"


# ─── Template / Esporta Excel ─────────────────────────────────────────────────
def template_bytes():
    """Template Excel ticker (TICKER/DESCRIZIONE/VALUTA/Peso %) formattato."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Portafoglio'
    headers    = ['TICKER', 'DESCRIZIONE', 'VALUTA', 'Peso %']
    col_widths = [14, 32, 10, 10]
    examples   = [
        ['ISAC.L',   'Az. ACWI',            'USD', ''],
        ['SWDA.MI',  'Az. World',           'EUR', ''],
        ['CSSPX.MI', 'Az. USA SP500',       'EUR', ''],
        ['EIMI.MI',  'Az. Emerging Market', 'EUR', ''],
        ['NVDA',     'NVIDIA Corporation',  'USD', ''],
    ]
    hdr_fill = PatternFill('solid', fgColor='1A3A5C')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    hdr_aln  = Alignment(horizontal='center', vertical='center')
    alt_fill = PatternFill('solid', fgColor='EEF4FF')
    whi_fill = PatternFill('solid', fgColor='FFFFFF')
    thin     = Side(style='thin', color='C0D0E8')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_aln, border
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 18
    for ri, row_data in enumerate(examples, 2):
        fill = alt_fill if ri % 2 == 0 else whi_fill
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill, c.border = fill, border
            c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[ri].height = 16
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def export_bytes(username=None):
    data = read_current(username)
    rows = [{'DESCRIZIONE': k, 'TICKER': v.get('ticker', ''),
             'VALUTA': v.get('currency', 'EUR'),
             'N_PREZZI': len(v.get('prices', [])),
             'ULTIMO_PREZZO': (v.get('prices') or [None])[-1]}
            for k, v in data.items() if isinstance(v, dict)]
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        pd.DataFrame(rows).to_excel(w, index=False, sheet_name='Asset')
        prices = build_prices(username)
        if prices is not None:
            prices.to_excel(w, sheet_name='Prezzi')
    out.seek(0)
    return out.read()


# ─── "File": salva/ricarica TUTTO il lavoro (current.json + analyses.json) ────
def profili_dir(username=None):
    u = username or get_username()
    d = ROOT / 'sessions' / u / 'tattica_profili'
    d.mkdir(parents=True, exist_ok=True)
    return d


def list_profili(username=None):
    out = []
    for p in sorted(profili_dir(username).glob('*.json'),
                    key=lambda x: x.stat().st_mtime, reverse=True):
        try:
            kb = max(1, round(p.stat().st_size / 1024))
        except Exception:
            kb = 0
        out.append({'name': p.name, 'label': p.stem, 'kb': kb})
    return out


def save_profilo(name, username=None):
    import re as _re
    name = (name or '').strip()
    if not name:
        return False, "⚠ Scrivi un nome per il lavoro"
    current = read_current(username)
    if not current:
        return False, "⚠ Nessun dato da salvare"
    snapshot = {'_format': 'tattica_v1',
                'current': current,
                'analyses': read_analyses(username)}
    safe = _re.sub(r'[^A-Za-z0-9_\- ]+', '_', name).strip()[:40] or 'lavoro'
    if write_json_atomic(profili_dir(username) / f"{safe}.json", snapshot):
        return True, f"✓ Salvato tutto il lavoro: '{safe}'"
    return False, "⚠ Errore salvataggio"


def load_profilo(filename, username=None):
    if not filename:
        return False, "⚠ Scegli un lavoro salvato"
    try:
        snap = json.load(open(profili_dir(username) / filename))
    except Exception:
        return False, "⚠ File non leggibile"
    if isinstance(snap, dict) and '_format' in snap:
        current  = snap.get('current', {}) or {}
        analyses = snap.get('analyses', None)
    else:
        current, analyses = (snap or {}), None
    ok = write_current(current, username)
    if analyses is not None:
        write_json_atomic(analyses_path(username), analyses)
    if ok:
        extra = f", {len(analyses)} analisi" if analyses else ""
        return True, f"✓ Caricato tutto: {len(current)} asset{extra}"
    return False, "⚠ Errore caricamento"


def delete_profilo(filename, username=None):
    if not filename:
        return False, "⚠ Niente da cancellare"
    try:
        (profili_dir(username) / filename).unlink()
        return True, "🗑 Lavoro cancellato"
    except Exception as e:
        return False, f"⚠ Errore: {e}"
